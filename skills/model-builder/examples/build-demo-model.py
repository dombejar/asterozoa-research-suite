#!/usr/bin/env python3
"""build-demo-model.py — worked example: generate a complete Asterozoa-caliber model.

Canonical D6 reference: the property-first-then-regroup pattern is implemented in
build_property_level_example(wb). It writes a P&L by Property block and a P&L by
Segment (Regroup) block where every segment formula is a SUM of property rows (never
a direct Actuals link), verified by check_row cells that evaluate to 0 after recalc.

Fictional company: Meridian Resorts Group, Inc. (NASDAQ: MRG) — a regional casino
operator with a ramping flagship (Harborview), a stable legacy asset (Legacy Riverboat),
and a contemplated development project (Canyon Crossing). ALL NUMBERS ARE FICTIONAL.

Demonstrates every architectural element of the skill:
  scenario toggles (3 financing scenarios x Base/Bull ramp) - per-property P&L build -
  development module - EBITDA->CFO engine with Check rows - per-instrument debt schedules
  with day-count interest - covenant test - SOTP valuation with MOIC/IRR (both case columns
  fully live via the per-case cash walk; zero circular references) - cap structure with
  YIELD - Asterozoa branding throughout.
  D6 addition: build_property_level_example appends a property-level regroup demo block
  after the main engine content (rows 127+), writes row metadata to a sidecar JSON
  file (<workbook-stem>.d6meta.json) next to the output workbook.

Run:    python3 build-demo-model.py [out.xlsx]
Then:   python3 ../scripts/recalc.py out.xlsx --in-place
        python3 ../scripts/audit_model.py out.xlsx --sweep \
            --toggle "Key Assumptions!D6=1,2,3" --toggle "Key Assumptions!D7=Base,Bull" \
            --output "Valuation Summary!G34" --output "Valuation Summary!H34" \
            --default "Key Assumptions!D6=2" --default "Key Assumptions!D7=Base"
"""
import datetime as dt
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))

import openpyxl  # noqa: E402
from model_kit import (  # noqa: E402
    FMT_DATE, FMT_MULT, FMT_NUM, FMT_NUM1, FMT_PCT1, FMT_PCT2, FMT_USD2, FMT_YN,
    brand_band, brand_footer, brand_title, check_row, choose_live, disclaimer_tab,
    input_block, margin_layout, period_headers, set_formula, set_input, set_label,
    set_link, set_total, spacer,
)

COMPANY = "Meridian Resorts Group, Inc. (NASDAQ: MRG) — DEMO / FICTIONAL DATA"
SRC = "Demo data — fictional, for skill illustration"

# ---- fictional actuals (FY23, FY24, FY25), internally consistent by construction
REV = {"Harborview": [0, 70, 120], "Canyon": [0, 0, 0], "Legacy": [85, 82, 80]}
EBITDA = {"Harborview": [-5, 14, 32], "Canyon": [0, 0, 0], "Legacy": [22, 21, 20],
          "Corporate": [-7, -8, -8]}
DWC = [2, 3, 2]
CASH_TAX = [-1, -2, -3]
CASH_INT = [-28, -28, -29]
CAPEX = [-60, -25, -16]
BORROW = [10, 0, 0]              # RCF draw FY23
BEG_CASH_FY23 = 120

TOT_REV = [sum(v[i] for v in REV.values()) for i in range(3)]
TOT_EBITDA = [sum(v[i] for v in EBITDA.values()) for i in range(3)]
CFO = [TOT_EBITDA[i] + DWC[i] + CASH_TAX[i] + CASH_INT[i] for i in range(3)]
END_CASH = []
beg = BEG_CASH_FY23
for i in range(3):
    end = beg + CFO[i] + CAPEX[i] + BORROW[i]
    END_CASH.append(end)
    beg = end

AN_COLS = ["D", "E", "F", "G", "H", "I", "J"]          # FY23 .. FY29E
ACT, FC = AN_COLS[:3], AN_COLS[3:]
LBL = ["FY23", "FY24", "FY25", "FY26E", "FY27E", "FY28E", "FY29E"]


def build_property_level_example(wb, wb_path: Path):
    """Write a property-level P&L block and a segment regroup block into the Model sheet.

    D6 canonical reference: the property-first-then-regroup pattern.

    This function:
    - Detects the last occupied row in the Model sheet dynamically (scans backward
      from row 200) to avoid hardcoded row collisions with the debt schedule.
    - Writes a section header, a property block (Harborview/Canyon Crossing/Legacy
      Riverboat revenue and EBITDA, each referencing existing Model rows via set_formula
      with same-sheet formula strings -- NOT set_link, which is for cross-sheet refs).
    - Writes a segment regroup block (Gaming Operations/Legacy Operations) using SUM
      formulas over the property block rows above. Segment rows NEVER source from Actuals
      directly (chain-of-custody rule, architecture.md section 9).
    - Writes check_row cells confirming segment totals == property totals.
    - Writes row metadata as JSON to a sidecar file (<workbook-stem>.d6meta.json)
      next to the output workbook so test_d6_regroup.py can locate all rows without
      hardcoding. No hidden sheet is created.
    - Does NOT touch freeze_panes or any sheet other than Model.

    Returns: dict with all row numbers written (same as the sidecar JSON).

    Risk R1 mitigation: last-row scan logged to stdout.
    Risk R4 mitigation: sidecar JSON separates formula-read from value-read passes.
    Risk R5 mitigation: freeze_panes is never touched here.
    """
    import json as _json

    m = wb["Model"]

    # -- Dynamic last-row scan (R1 mitigation) --
    # Scan backward from sentinel row 200 across all data columns to find last occupied row.
    last_row = 0
    for scan_row in range(200, 0, -1):
        for col in AN_COLS + ["B", "C", "Z"]:
            if m[f"{col}{scan_row}"].value is not None:
                last_row = max(last_row, scan_row)
        if last_row >= scan_row:
            break  # first row with content found; no rows above it can be higher
    print(f"  [D6] last occupied Model row (dynamic scan): {last_row}")

    start = last_row + 3  # two-row gap spacer

    # -- Section header --
    set_label(m, f"B{start}", "P&L by Property -- Detailed Regroup Demo (D6)", bold=True)

    # -- Property block: Revenue rows --
    # Existing Model property revenue rows: Harborview=49, Canyon Crossing=50, Legacy Riverboat=51
    prop_rev_start = start + 1
    set_label(m, f"B{prop_rev_start}", "Property Revenues (D6 Demo)", bold=True)
    pr_harborview = prop_rev_start + 1
    pr_canyon = prop_rev_start + 2
    pr_legacy = prop_rev_start + 3
    for row, src_row, label in (
        (pr_harborview, 49, "Harborview"),
        (pr_canyon, 50, "Canyon Crossing"),
        (pr_legacy, 51, "Legacy Riverboat"),
    ):
        set_label(m, f"B{row}", label)
        for c in AN_COLS:
            set_formula(m, f"{c}{row}", f"=+{c}{src_row}", fmt=FMT_NUM1)

    prop_rev_total = pr_legacy + 1
    set_label(m, f"B{prop_rev_total}", "Total Property Revenue", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}{prop_rev_total}",
                  f"=SUM({c}{pr_harborview}:{c}{pr_legacy})", fmt=FMT_NUM1)
    prop_rev_end = prop_rev_total

    # -- Property block: EBITDA rows --
    # Existing Model EBITDA rows: Harborview=59, Canyon Crossing=60, Legacy Riverboat=61
    prop_eb_start = prop_rev_end + 2
    set_label(m, f"B{prop_eb_start}", "Property EBITDA (D6 Demo)", bold=True)
    pe_harborview = prop_eb_start + 1
    pe_canyon = prop_eb_start + 2
    pe_legacy = prop_eb_start + 3
    for c in AN_COLS:
        set_formula(m, f"{c}{pe_harborview}", f"=+{c}59", fmt=FMT_NUM1)
        set_formula(m, f"{c}{pe_canyon}", f"=+{c}60", fmt=FMT_NUM1)
        set_formula(m, f"{c}{pe_legacy}", f"=+{c}61", fmt=FMT_NUM1)
    set_label(m, f"B{pe_harborview}", "Harborview")
    set_label(m, f"B{pe_canyon}", "Canyon Crossing")
    set_label(m, f"B{pe_legacy}", "Legacy Riverboat")

    prop_eb_total = pe_legacy + 1
    set_label(m, f"B{prop_eb_total}", "Total Property EBITDA", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}{prop_eb_total}",
                  f"=SUM({c}{pe_harborview}:{c}{pe_legacy})", fmt=FMT_NUM1)
    prop_eb_end = prop_eb_total

    # -- Segment regroup block --
    # Segment rows SUM the property rows above -- NEVER sourced from Actuals directly.
    seg_start = prop_eb_end + 2
    set_label(m, f"B{seg_start}", "P&L by Segment (Regroup)", bold=True)

    # Revenue segment rows
    seg_rev_start = seg_start
    sg_rev_gaming = seg_start + 1  # Gaming = Harborview + Canyon Crossing
    sg_rev_legacy = seg_start + 2  # Legacy = Legacy Riverboat
    seg_rev_total = seg_start + 3
    set_label(m, f"B{sg_rev_gaming}", "Gaming Operations")
    set_label(m, f"B{sg_rev_legacy}", "Legacy Operations")
    set_label(m, f"B{seg_rev_total}", "Total Segment Revenue", bold=True)
    for c in AN_COLS:
        # SUM of property revenue rows (pr_harborview, pr_canyon) -- not Actuals
        set_formula(m, f"{c}{sg_rev_gaming}",
                    f"=SUM({c}{pr_harborview}:{c}{pr_canyon})", fmt=FMT_NUM1)
        # Legacy segment = SUM of Legacy property row -- not Actuals
        set_formula(m, f"{c}{sg_rev_legacy}",
                    f"=SUM({c}{pr_legacy}:{c}{pr_legacy})", fmt=FMT_NUM1)
        set_total(m, f"{c}{seg_rev_total}",
                  f"=SUM({c}{sg_rev_gaming}:{c}{sg_rev_legacy})", fmt=FMT_NUM1)
    seg_rev_end = seg_rev_total

    # EBITDA segment rows
    seg_eb_start = seg_rev_end + 1
    sg_eb_gaming = seg_eb_start + 1
    sg_eb_legacy = seg_eb_start + 2
    seg_eb_total = seg_eb_start + 3
    set_label(m, f"B{seg_eb_start}", "Segment EBITDA (Regroup)", bold=True)
    set_label(m, f"B{sg_eb_gaming}", "Gaming Operations")
    set_label(m, f"B{sg_eb_legacy}", "Legacy Operations")
    set_label(m, f"B{seg_eb_total}", "Total Segment EBITDA", bold=True)
    for c in AN_COLS:
        set_formula(m, f"{c}{sg_eb_gaming}",
                    f"=SUM({c}{pe_harborview}:{c}{pe_canyon})", fmt=FMT_NUM1)
        set_formula(m, f"{c}{sg_eb_legacy}",
                    f"=SUM({c}{pe_legacy}:{c}{pe_legacy})", fmt=FMT_NUM1)
        set_total(m, f"{c}{seg_eb_total}",
                  f"=SUM({c}{sg_eb_gaming}:{c}{sg_eb_legacy})", fmt=FMT_NUM1)
    seg_eb_end = seg_eb_total

    # -- Check rows: segment totals must equal property totals --
    rev_check = seg_eb_end + 1
    ebitda_check = seg_eb_end + 2
    check_row(m, rev_check, "B", AN_COLS, f"{{c}}{seg_rev_total}", f"{{c}}{prop_rev_total}")
    check_row(m, ebitda_check, "B", AN_COLS, f"{{c}}{seg_eb_total}", f"{{c}}{prop_eb_total}")

    # -- Write row metadata to sidecar JSON file --
    meta = {
        "prop_rev_start": prop_rev_start,
        "prop_rev_end": prop_rev_end,
        "prop_eb_start": prop_eb_start,
        "prop_eb_end": prop_eb_end,
        "seg_rev_start": seg_rev_start,
        "seg_rev_end": seg_rev_end,
        "seg_eb_start": seg_eb_start,
        "seg_eb_end": seg_eb_end,
        "rev_check": rev_check,
        "ebitda_check": ebitda_check,
    }
    sidecar_path = Path(wb_path).with_suffix("").with_name(
        Path(wb_path).stem + ".d6meta.json"
    )
    sidecar_path.write_text(_json.dumps(meta, indent=2))
    print(f"  [D6] metadata written to sidecar: {sidecar_path}")
    return meta


def build(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ================================================================ Actuals
    a = wb.create_sheet("Actuals")
    margin_layout(a, label_width=42)
    brand_title(a, COMPANY, "Actuals", rule_to_col="F")
    set_label(a, "B7", "Period Ended:")
    period_headers(a, 7, 4, dt.date(2023, 12, 31), 3, labels=LBL[:3], label_row=8)
    set_label(a, "B8", "($ in Millions)")

    r = 10
    set_label(a, f"B{r}", "P&L by Property", bold=True); r += 1
    set_label(a, f"B{r}", "Revenues", bold=True); r += 1
    rev_rows = {}
    for name in ("Harborview", "Canyon", "Legacy"):
        set_label(a, f"B{r}", {"Canyon": "Canyon Crossing",
                               "Legacy": "Legacy Riverboat"}.get(name, name), indent=0)
        for i, c in enumerate(ACT):
            set_input(a, f"{c}{r}", REV[name][i], fmt=FMT_NUM1, source=SRC)
        rev_rows[name] = r; r += 1
    set_label(a, f"B{r}", "Total Revenues", bold=True)
    for c in ACT:
        set_total(a, f"{c}{r}", f"=SUM({c}{rev_rows['Harborview']}:{c}{r-1})", fmt=FMT_NUM1)
    tot_rev_row = r; r += 2

    set_label(a, f"B{r}", "Adj EBITDA", bold=True); r += 1
    eb_rows = {}
    for name in ("Harborview", "Canyon", "Legacy", "Corporate"):
        set_label(a, f"B{r}", {"Canyon": "Canyon Crossing",
                               "Legacy": "Legacy Riverboat"}.get(name, name))
        for i, c in enumerate(ACT):
            set_input(a, f"{c}{r}", EBITDA[name][i], fmt=FMT_NUM1, source=SRC)
        eb_rows[name] = r; r += 1
    set_label(a, f"B{r}", "Total Adj EBITDA", bold=True)
    for c in ACT:
        set_total(a, f"{c}{r}", f"=SUM({c}{eb_rows['Harborview']}:{c}{r-1})", fmt=FMT_NUM1)
    tot_eb_row = r; r += 2

    set_label(a, f"B{r}", "Cash Flow (as reported)", bold=True); r += 1
    cf_map = {}
    for label, vals in (("Change in working capital", DWC),
                        ("Cash paid for income taxes", CASH_TAX),
                        ("Cash interest", CASH_INT),
                        ("Net cash provided by operating activities", CFO),
                        ("Capital expenditures", CAPEX),
                        ("Debt borrowings", BORROW),
                        ("Ending cash and equivalents", END_CASH)):
        set_label(a, f"B{r}", label)
        for i, c in enumerate(ACT):
            set_input(a, f"{c}{r}", vals[i], fmt=FMT_NUM1, source=SRC)
        cf_map[label] = r; r += 1
    a.freeze_panes = "C9"

    # ================================================================ Key Assumptions
    ka = wb.create_sheet("Key Assumptions", 0)
    margin_layout(ka, label_width=34, notes_col="L", notes_width=46)
    brand_title(ka, COMPANY, "Key Assumptions", rule_to_col="L", tab="assumption")
    set_label(ka, "B6", "Scenario Toggle:")
    set_input(ka, "D6", 2, fmt="General", source="User toggle (1-3)")
    set_label(ka, "B7", "P&L Ramp Assumptions:")
    set_input(ka, "D7", "Base", fmt="General", source="User toggle (Base/Bull)")
    set_label(ka, "B9", "Scenarios:", bold=True)
    for i, desc in enumerate((
            "No development. Run existing assets; no new financing.",
            "Develop Canyon Crossing. $150M new notes at 10.5% on 6/30/26.",
            "Develop + refinance: scenario 2 plus refi of 9.00% notes at YE27 with $300M at 8.50%."), start=10):
        set_label(ka, f"B{i}", str(i - 9)); set_label(ka, f"C{i}", desc)
    set_label(ka, "B14", "Ramp Cases:", bold=True)
    set_label(ka, "B15", "Base"); set_input(ka, "C15", 1, fmt="General", source="enum")
    set_label(ka, "B16", "Bull"); set_input(ka, "C16", 2, fmt="General", source="enum")

    set_label(ka, "E18", "Live", bold=True)
    for col, s in (("G", "Scenario 1"), ("H", "Scenario 2"), ("I", "Scenario 3")):
        set_label(ka, f"{col}18", s, bold=True)
        ka.column_dimensions[col].width = 13
    set_label(ka, "L18", "Notes", bold=True)
    ka.column_dimensions["E"].width = 13
    spacer(ka, "F")

    def ka_row(r, label, vals, fmt=FMT_NUM1, note=None):
        set_label(ka, f"B{r}", label)
        for col, v in zip(("G", "H", "I"), vals):
            set_input(ka, f"{col}{r}", v, fmt=fmt, source=SRC)
        choose_live(ka, f"E{r}", "$D$6", [f"G{r}", f"H{r}", f"I{r}"], fmt=fmt)
        if note:
            set_label(ka, f"L{r}", note)

    set_label(ka, "B20", "Revolving Credit Facility", bold=True)
    ka_row(21, "RCF commitments", [30, 30, 30])
    ka_row(22, "Drawn", [10, 10, 10])
    ka_row(23, "SOFR (constant)", [0.04, 0.04, 0.04], fmt=FMT_PCT2)
    ka_row(24, "Margin over SOFR", [0.03, 0.03, 0.03], fmt=FMT_PCT2)
    ka_row(25, "Annual commitment fee", [0.005, 0.005, 0.005], fmt=FMT_PCT2)

    set_label(ka, "B27", "Existing 9.00% Senior Notes due 2029", bold=True)
    ka_row(28, "Amount", [300, 300, 300])
    ka_row(29, "Interest rate", [0.09, 0.09, 0.09], fmt=FMT_PCT2)
    ka_row(30, "Repayment date", [dt.date(2099, 12, 31), dt.date(2099, 12, 31),
                                  dt.date(2027, 12, 31)], fmt=FMT_DATE,
           note="2099 = held to maturity (beyond model horizon)")

    set_label(ka, "B32", "New Financing - 1 (development notes)", bold=True)
    ka_row(33, "Borrowing date", [dt.date(2099, 12, 31), dt.date(2026, 6, 30),
                                  dt.date(2026, 6, 30)], fmt=FMT_DATE)
    ka_row(34, "Borrowing amount", [0, 150, 150])
    ka_row(35, "Interest rate", [0, 0.105, 0.105], fmt=FMT_PCT2)
    ka_row(36, "Issuance costs", [0, 0.02, 0.02], fmt=FMT_PCT2)

    set_label(ka, "B38", "New Financing - 2 (refinancing)", bold=True)
    ka_row(39, "Borrowing date", [dt.date(2099, 12, 31), dt.date(2099, 12, 31),
                                  dt.date(2027, 12, 31)], fmt=FMT_DATE)
    ka_row(40, "Borrowing amount", [0, 0, 300])
    ka_row(41, "Interest rate", [0, 0, 0.085], fmt=FMT_PCT2)
    ka_row(42, "Issuance costs", [0, 0, 0.02], fmt=FMT_PCT2)

    set_label(ka, "B44", "Other", bold=True)
    ka_row(45, "Develop Canyon Crossing?", ["No", "Yes", "Yes"], fmt="General")
    ka_row(46, "Maintenance capex, % of revenue", [0.08, 0.08, 0.08], fmt=FMT_PCT1)
    ka_row(47, "Change in WC, % of revenue change", [0.02, 0.02, 0.02], fmt=FMT_PCT1)
    ka_row(48, "Cash tax rate (on EBITDA less cash interest)", [0.12, 0.12, 0.12],
           fmt=FMT_PCT1, note="Demo simplification — real models tax EBT")
    input_block(ka, 21, 7, 28, 3)
    ka.freeze_panes = "E19"

    # ================================================================ P&L Ramp Assumptions
    rp = wb.create_sheet("P&L Ramp Assumptions", 1)
    margin_layout(rp, label_width=30)
    brand_title(rp, COMPANY, "P&L Ramp Assumptions", rule_to_col="U", tab="assumption")
    set_label(rp, "B5", "Financing Scenario:")
    set_link(rp, "C5", "=+'Key Assumptions'!$D$6", fmt="General")
    set_label(rp, "B6", "P&L Ramp Assumptions:")
    set_link(rp, "C6", "='Key Assumptions'!$D$7", fmt="General")
    set_label(rp, "B8", "Cases:", bold=True)
    set_label(rp, "B9", "Base"); set_input(rp, "C9", 1, fmt="General", source="enum")
    set_label(rp, "B10", "Bull"); set_input(rp, "C10", 2, fmt="General", source="enum")

    set_label(rp, "G12", "Live", bold=True)
    set_label(rp, "M12", "Base Case", bold=True)
    set_label(rp, "R12", "Bull Case", bold=True)
    period_headers(rp, 13, 4, dt.date(2023, 12, 31), 7, labels=LBL, label_row=14)
    for i, c in enumerate(["M", "N", "O", "P"]):
        set_formula(rp, f"{c}13", f"=+{FC[i]}13", fmt=FMT_DATE)
        set_label(rp, f"{c}14", LBL[3 + i], bold=True)
        rp.column_dimensions[c].width = 9.7
    for i, c in enumerate(["R", "S", "T", "U"]):
        set_formula(rp, f"{c}13", f"=+{FC[i]}13", fmt=FMT_DATE)
        set_label(rp, f"{c}14", LBL[3 + i], bold=True)
        rp.column_dimensions[c].width = 9.7
    spacer(rp, "K"); spacer(rp, "L", 1.7); spacer(rp, "Q", 4.7)

    BASE, BULL = ["M", "N", "O", "P"], ["R", "S", "T", "U"]

    def live_choose(row, fc_idx, fmt=FMT_NUM1):
        c = FC[fc_idx]
        set_formula(rp, f"{c}{row}",
                    f"=CHOOSE(VLOOKUP($C$6,$B$9:$C$10,2,FALSE),"
                    f"{BASE[fc_idx]}{row},{BULL[fc_idx]}{row})", fmt=fmt)

    # Revenues
    set_label(rp, "B16", "Revenues", bold=True)
    props = [("Harborview", 17), ("Canyon Crossing", 18), ("Legacy Riverboat", 19)]
    grow = {"Harborview": ([0.15, 0.10, 0.08, 0.05], [0.20, 0.14, 0.10, 0.08]),
            "Legacy Riverboat": ([0.00, 0.00, -0.02, -0.02], [0.02, 0.02, 0.00, 0.00])}
    marg = {"Harborview": ([0.28, 0.29, 0.30, 0.31], [0.29, 0.31, 0.32, 0.33]),
            "Canyon Crossing": ([0, 0, 0.22, 0.28], [0, 0, 0.24, 0.30]),
            "Legacy Riverboat": ([0.25] * 4, [0.26] * 4)}
    # growth driver rows 23-25, margin rows 33-35; Canyon opening revenue input C24
    for name, r in props:
        set_label(rp, f"B{r}", name)
        for i, c in enumerate(ACT):
            set_link(rp, f"{c}{r}", f"=+Model!{c}{49 + (r - 17)}", fmt=FMT_NUM1)
        for i in range(4):
            live_choose(r, i)
    set_label(rp, "B20", "Total Revenues", bold=True)
    for c in AN_COLS + BASE + BULL:
        set_total(rp, f"{c}20", f"=SUM({c}17:{c}19)", fmt=FMT_NUM1)

    set_label(rp, "B22", "yoy growth % / drivers", bold=True)
    for name, r in (("Harborview", 23), ("Legacy Riverboat", 25)):
        set_label(rp, f"B{r}", name)
        for blk, case_i in ((BASE, 0), (BULL, 1)):
            for i, c in enumerate(blk):
                set_input(rp, f"{c}{r}", grow[name][case_i][i], fmt=FMT_PCT1, source=SRC)
    set_label(rp, "B24", "Canyon Crossing opening-year revenue")
    set_input(rp, "C24", 60, fmt=FMT_NUM1, source=SRC + " (FY28 opening, Base)")
    set_input(rp, "D24", 75, fmt=FMT_NUM1, source=SRC + " (FY28 opening, Bull)")
    set_label(rp, "B26", "Canyon yoy growth (post-opening)")
    for blk, vals in ((BASE, [0, 0, 0, 0.50]), (BULL, [0, 0, 0, 0.60])):
        for i, c in enumerate(blk):
            set_input(rp, f"{c}26", vals[i], fmt=FMT_PCT1, source=SRC)

    # case-block revenue formulas
    for blk, canyon_open_cell in ((BASE, "$C$24"), (BULL, "$D$24")):
        prev = ["F"] + blk[:-1]
        for i, c in enumerate(blk):
            set_formula(rp, f"{c}17", f"=+{prev[i]}17*(1+{c}23)", fmt=FMT_NUM1)   # Harborview
            set_formula(rp, f"{c}19", f"=+{prev[i]}19*(1+{c}25)", fmt=FMT_NUM1)   # Legacy
        # Canyon: zero until FY28 opening (gated by Develop flag), then growth
        set_formula(rp, f"{blk[0]}18", "=0*1", fmt=FMT_NUM1)
        set_formula(rp, f"{blk[1]}18", "=0*1", fmt=FMT_NUM1)
        set_link(rp, f"{blk[2]}18",
                 f"=+{canyon_open_cell}*('Key Assumptions'!$E$45=\"Yes\")", fmt=FMT_NUM1)
        set_formula(rp, f"{blk[3]}18", f"=+{blk[2]}18*(1+{blk[3]}26)", fmt=FMT_NUM1)

    # EBITDA
    set_label(rp, "B28", "Adj EBITDA", bold=True)
    for name, r in (("Harborview", 29), ("Canyon Crossing", 30), ("Legacy Riverboat", 31)):
        set_label(rp, f"B{r}", name)
        for i, c in enumerate(ACT):
            set_link(rp, f"{c}{r}", f"=+Actuals!{c}{eb_rows[name.split()[0]]}", fmt=FMT_NUM1)
        for blk in (BASE, BULL):
            for c in blk:
                set_formula(rp, f"{c}{r}", f"=+{c}{r - 12}*{c}{r + 5}", fmt=FMT_NUM1)
        for i in range(4):
            live_choose(r, i)
    set_label(rp, "B32", "Corporate")
    for i, c in enumerate(ACT):
        set_link(rp, f"{c}32", f"=+Actuals!{c}{eb_rows['Corporate']}", fmt=FMT_NUM1)
    for blk, vals in ((BASE, [-9, -9, -10, -10]), (BULL, [-9, -9, -10, -10])):
        for i, c in enumerate(blk):
            set_input(rp, f"{c}32", vals[i], fmt=FMT_NUM1, source=SRC)
    for i in range(4):
        live_choose(32, i)
    set_label(rp, "B33", "Total Adj EBITDA", bold=True)
    for c in AN_COLS + BASE + BULL:
        set_total(rp, f"{c}33", f"=SUM({c}29:{c}32)", fmt=FMT_NUM1)

    set_label(rp, "B34", "% margin", bold=True)
    for name, r in (("Harborview", 34), ("Canyon Crossing", 35), ("Legacy Riverboat", 36)):
        pass
    # margin driver rows 34-36 (blocks only; live computed on Model)
    for name, r in (("Harborview", 34), ("Canyon Crossing", 35), ("Legacy Riverboat", 36)):
        set_label(rp, f"B{r}", name)
        for blk, case_i in ((BASE, 0), (BULL, 1)):
            for i, c in enumerate(blk):
                set_input(rp, f"{c}{r}", marg[name][case_i][i], fmt=FMT_PCT1, source=SRC)
    for blk in (BASE, BULL):
        input_block(rp, 23, openpyxl.utils.column_index_from_string(blk[0]), 4, 4)
        input_block(rp, 34, openpyxl.utils.column_index_from_string(blk[0]), 3, 4)

    # ---- per-case cash walk: lets Valuation Summary show BOTH cases fully live with
    # ZERO circular references (debt/interest/dev-capex/financing are scenario-only,
    # so only the operating walk differs by case)
    set_label(rp, "B38", "Per-case cash walk (memo)", bold=True)
    set_label(rp, "B39", "Cash from operations (per case)")
    set_label(rp, "B40", "FCF after capex + financing (per case)")
    set_label(rp, "B41", "FY29E ending cash (per case)")
    set_label(rp, "B42", "FY29E net debt (per case)")
    for blk in (BASE, BULL):
        prev = ["F"] + blk[:-1]
        for i, c in enumerate(blk):
            fc = FC[i]
            set_link(rp, f"{c}39",
                     f"=+{c}33+({c}20-{prev[i]}20)*'Key Assumptions'!$E$47"
                     f"-MAX({c}33-Model!{fc}$118,0)*'Key Assumptions'!$E$48"
                     f"-Model!{fc}$118", fmt=FMT_NUM1)
            set_link(rp, f"{c}40",
                     f"=+{c}39-{c}20*'Key Assumptions'!$E$46"
                     f"+Model!{fc}$26+Model!{fc}$29+Model!{fc}$30", fmt=FMT_NUM1)
        set_link(rp, f"{blk[3]}41", f"=+Model!$F$35+SUM({blk[0]}40:{blk[3]}40)", fmt=FMT_NUM1)
        set_link(rp, f"{blk[3]}42", f"=+Model!$J$110-{blk[3]}41", fmt=FMT_NUM1)
    rp.freeze_panes = "D15"

    # ================================================================ Canyon Develop. Assumptions
    dv = wb.create_sheet("Canyon Develop. Assumptions", 2)
    margin_layout(dv, label_width=44)
    brand_title(dv, COMPANY, "Canyon Crossing Development Assumptions",
                rule_to_col="J", tab="assumption")
    set_label(dv, "B5", "Live Scenario:")
    set_link(dv, "C5", "=+'Key Assumptions'!$D$6", fmt="General")
    period_headers(dv, 7, 4, dt.date(2023, 12, 31), 7, labels=LBL, label_row=8)

    set_label(dv, "B10", "Live (selected scenario)", bold=True)
    live_rows = {"budget": 11, "capex": 12}
    set_label(dv, "B11", "Total development budget")
    set_label(dv, "B12", "Development capex (cash out)")
    # scenario blocks: rows 15 (s1), 19 (s2), 23 (s3): budget row + capex row
    blocks = {1: (15, [0, 0, 0, 0]), 2: (19, [-72, -90, -18, 0]), 3: (23, [-72, -90, -18, 0])}
    for s, (r0, phase) in blocks.items():
        set_label(dv, f"B{r0 - 1}", f"Scenario {s}", bold=True)
        set_label(dv, f"B{r0}", "Total development budget")
        set_input(dv, f"C{r0}", 0 if s == 1 else 180, fmt=FMT_NUM1,
                  source=SRC + " (hard 120 / soft 30 / contingency 18 / preopening 12)")
        set_label(dv, f"B{r0 + 1}", "Development capex (cash out)")
        for i, c in enumerate(FC):
            set_input(dv, f"{c}{r0 + 1}", phase[i], fmt=FMT_NUM1,
                      source=SRC + " (40%/50%/10% phasing)")
        input_block(dv, r0, 3, 2, 8)
    set_formula(dv, "C11", "=CHOOSE($C$5,C15,C19,C23)", fmt=FMT_NUM1)
    for c in FC:
        set_formula(dv, f"{c}12", f"=CHOOSE($C$5,{c}16,{c}20,{c}24)", fmt=FMT_NUM1)
    dv.freeze_panes = "D9"

    # ================================================================ Model (engine)
    m = wb.create_sheet("Model", 3)
    margin_layout(m, label_width=36, notes_col="Z", notes_width=46)
    brand_title(m, COMPANY, "Financial Model", rule_to_col="J")
    set_label(m, "B5", "Live Scenario:")
    set_link(m, "C5", "=+'Key Assumptions'!$D$6", fmt="General")
    set_label(m, "B6", "P&L Ramp Assumptions:")
    set_link(m, "C6", "='Key Assumptions'!$D$7", fmt="General")
    set_label(m, "B7", "Period Ended:")
    period_headers(m, 7, 4, dt.date(2023, 12, 31), 7, labels=LBL, label_row=8)
    set_label(m, "B8", "($ in Millions)")

    # ---- P&L and Cash Flow Summary
    set_label(m, "B10", "P&L and Cash Flow Summary", bold=True)
    set_label(m, "B12", "Revenues")
    for c in AN_COLS:
        set_formula(m, f"{c}12", f"=+{c}52", fmt=FMT_NUM1)
    set_label(m, "B13", "yoy, %")
    for i, c in enumerate(AN_COLS[1:], start=1):
        set_formula(m, f"{c}13", f'=+IFERROR({c}12/{AN_COLS[i-1]}12-1,"nm")', fmt=FMT_PCT1)
    set_label(m, "B15", "Adjusted EBITDA")
    for c in AN_COLS:
        set_formula(m, f"{c}15", f"=+{c}63", fmt=FMT_NUM1)
    set_label(m, "B16", "margin, %")
    for c in AN_COLS:
        set_formula(m, f"{c}16", f'=+IFERROR({c}15/{c}12,"nm")', fmt=FMT_PCT1)

    set_label(m, "B18", "Plus/(Less):")
    set_label(m, "B19", "Change in working capital")
    for i, c in enumerate(ACT):
        set_link(m, f"{c}19", f"=+Actuals!{c}{cf_map['Change in working capital']}", fmt=FMT_NUM1)
    for i, c in enumerate(FC):
        prev = AN_COLS[3 + i - 1]
        set_link(m, f"{c}19", f"=+({c}12-{prev}12)*'Key Assumptions'!$E$47", fmt=FMT_NUM1)
    set_label(m, "B20", "Cash paid for income taxes")
    for c in ACT:
        set_link(m, f"{c}20", f"=+Actuals!{c}{cf_map['Cash paid for income taxes']}", fmt=FMT_NUM1)
    for c in FC:
        set_link(m, f"{c}20", f"=-MAX({c}15+{c}21,0)*'Key Assumptions'!$E$48", fmt=FMT_NUM1)
    set_label(m, "B21", "Cash interest")
    for c in ACT:
        set_link(m, f"{c}21", f"=+Actuals!{c}{cf_map['Cash interest']}", fmt=FMT_NUM1)
    for c in FC:
        set_formula(m, f"{c}21", f"=-{c}118", fmt=FMT_NUM1)
    set_label(m, "B22", "Cash from operations", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}22", f"=SUM({c}15,{c}19:{c}21)", fmt=FMT_NUM1)
    check_row(m, 23, "B", ACT, "{c}22",
              f"Actuals!{{c}}{cf_map['Net cash provided by operating activities']}")

    set_label(m, "B25", "Maintenance capital expenditures")
    for c in ACT:
        set_link(m, f"{c}25", f"=+Actuals!{c}{cf_map['Capital expenditures']}", fmt=FMT_NUM1)
    for c in FC:
        set_link(m, f"{c}25", f"=-{c}12*'Key Assumptions'!$E$46", fmt=FMT_NUM1)
    set_label(m, "B26", "Development capex (Canyon Crossing)")
    for c in FC:
        set_link(m, f"{c}26", f"=+'Canyon Develop. Assumptions'!{c}12", fmt=FMT_NUM1)
    set_label(m, "B27", "Cash from investing", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}27", f"=SUM({c}25:{c}26)", fmt=FMT_NUM1)

    set_label(m, "B29", "Debt borrowings/(repayments)")
    for c in ACT:
        set_link(m, f"{c}29", f"=+Actuals!{c}{cf_map['Debt borrowings']}", fmt=FMT_NUM1)
    for c in FC:
        set_formula(m, f"{c}29", f"=+{c}112", fmt=FMT_NUM1)
    set_label(m, "B30", "Payment of debt issuance costs")
    for c in FC:
        set_formula(m, f"{c}30", f"=+{c}124", fmt=FMT_NUM1)
    set_label(m, "B31", "Cash from financing", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}31", f"=SUM({c}29:{c}30)", fmt=FMT_NUM1)

    set_label(m, "B33", "Net cash flow", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}33", f"=+{c}22+{c}27+{c}31", fmt=FMT_NUM1)
    set_label(m, "B34", "Beginning cash balance")
    set_input(m, "D34", BEG_CASH_FY23, fmt=FMT_NUM1, source=SRC)
    for i, c in enumerate(AN_COLS[1:], start=1):
        set_formula(m, f"{c}34", f"=+{AN_COLS[i-1]}35", fmt=FMT_NUM1)
    set_label(m, "B35", "Ending cash balance", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}35", f"=SUM({c}33:{c}34)", fmt=FMT_NUM1)
    check_row(m, 36, "B", ACT, "{c}35",
              f"Actuals!{{c}}{cf_map['Ending cash and equivalents']}")

    # ---- Liquidity / Leverage / Covenant
    set_label(m, "B38", "Liquidity, Leverage and Covenants", bold=True)
    set_label(m, "B39", "Cash and equivalents")
    for c in AN_COLS:
        set_formula(m, f"{c}39", f"=+{c}35", fmt=FMT_NUM1)
    set_label(m, "B40", "RCF availability")
    for c in AN_COLS:
        set_link(m, f"{c}40", "=+'Key Assumptions'!$E$21-'Key Assumptions'!$E$22", fmt=FMT_NUM1)
    set_label(m, "B41", "Total liquidity", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}41", f"=SUM({c}39:{c}40)", fmt=FMT_NUM1)
    set_label(m, "B43", "Total debt")
    for c in AN_COLS:
        set_formula(m, f"{c}43", f"=+{c}110", fmt=FMT_NUM1)
    set_label(m, "B44", "Net debt")
    for c in AN_COLS:
        set_formula(m, f"{c}44", f"=+{c}43-{c}39", fmt=FMT_NUM1)
    set_label(m, "B45", "Net debt / Adj EBITDA")
    for c in AN_COLS:
        set_formula(m, f"{c}45", f'=+IFERROR({c}44/{c}15,"nm")', fmt=FMT_MULT)
    set_label(m, "B46", "Covenant: Adj EBITDA exceeds drawn RCF?")
    for c in AN_COLS:
        set_link(m, f"{c}46", f"=({c}15>'Key Assumptions'!$E$22)*1", fmt=FMT_YN)

    # ---- P&L by Property
    set_label(m, "B48", "P&L by Property", bold=True)
    set_label(m, "B48", "P&L by Property", bold=True)
    for name, r in (("Harborview", 49), ("Canyon Crossing", 50), ("Legacy Riverboat", 51)):
        set_label(m, f"B{r}", name)
        for c in ACT:
            set_link(m, f"{c}{r}",
                     f"=+Actuals!{c}{rev_rows[name.split()[0]]}", fmt=FMT_NUM1)
        for c in FC:
            set_link(m, f"{c}{r}", f"=+'P&L Ramp Assumptions'!{c}{17 + (r - 49)}", fmt=FMT_NUM1)
    set_label(m, "B52", "Total Revenues", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}52", f"=SUM({c}49:{c}51)", fmt=FMT_NUM1)

    for name, r in (("Harborview", 59), ("Canyon Crossing", 60), ("Legacy Riverboat", 61),
                    ("Corporate", 62)):
        set_label(m, f"B{r}", name + (" Adj EBITDA" if r == 59 else ""))
        for c in ACT:
            set_link(m, f"{c}{r}", f"=+Actuals!{c}{eb_rows[name.split()[0]]}", fmt=FMT_NUM1)
        for c in FC:
            set_link(m, f"{c}{r}", f"=+'P&L Ramp Assumptions'!{c}{29 + (r - 59)}", fmt=FMT_NUM1)
    set_label(m, "B58", "Adj EBITDA by Property", bold=True)
    set_label(m, "B63", "Total Adj EBITDA", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}63", f"=SUM({c}59:{c}62)", fmt=FMT_NUM1)

    # ---- Debt schedules (forecast columns only; actuals carried as inputs)
    set_label(m, "B66", "Debt/Equity Schedule", bold=True)

    set_label(m, "B68", "Revolving Credit Facility", bold=True)
    set_label(m, "B69", "Drawn balance (constant in demo)")
    for c in AN_COLS:
        set_link(m, f"{c}69", "=+'Key Assumptions'!$E$22", fmt=FMT_NUM1)
    set_label(m, "B70", "Interest rate (SOFR + margin)")
    for c in FC:
        set_link(m, f"{c}70", "=+'Key Assumptions'!$E$23+'Key Assumptions'!$E$24", fmt=FMT_PCT2)
    set_label(m, "B71", "Cash interest + commitment fee")
    for c in FC:
        set_link(m, f"{c}71",
                 f"=+{c}69*{c}70+('Key Assumptions'!$E$21-'Key Assumptions'!$E$22)"
                 f"*'Key Assumptions'!$E$25", fmt=FMT_NUM1)

    set_label(m, "B73", "Existing 9.00% Senior Notes due 2029", bold=True)
    set_label(m, "B74", "Repayment date")
    set_link(m, "C74", "=+'Key Assumptions'!$E$30", fmt=FMT_DATE)
    set_label(m, "B75", "Beginning balance")
    set_link(m, "F75", "=+'Key Assumptions'!$E$28", fmt=FMT_NUM1)
    for i, c in enumerate(FC):
        prev = AN_COLS[3 + i - 1]
        set_formula(m, f"{c}75", f"=+{prev}77", fmt=FMT_NUM1)
    set_label(m, "B76", "Borrowings/(repayments)")
    for c in FC:
        set_formula(m, f"{c}76", f"=-{c}75*($C$74<={c}$7)*($C$74>{chr(ord(c)-1)}$7)",
                    fmt=FMT_NUM1)
    set_label(m, "B77", "Ending balance")
    set_formula(m, "F77", "=+F75", fmt=FMT_NUM1)
    for c in FC:
        set_formula(m, f"{c}77", f"=SUM({c}75:{c}76)", fmt=FMT_NUM1)
    set_label(m, "B78", "Cash interest")
    for c in FC:
        set_link(m, f"{c}78", f"=+{c}75*'Key Assumptions'!$E$29", fmt=FMT_NUM1)

    def new_financing_block(r0, label, date_cell, amt_cell, rate_cell, cost_cell):
        set_label(m, f"B{r0}", label, bold=True)
        set_label(m, f"B{r0+1}", "Borrowing date")
        set_link(m, f"C{r0+1}", f"=+'Key Assumptions'!{date_cell}", fmt=FMT_DATE)
        set_label(m, f"B{r0+2}", "Beginning balance")
        set_formula(m, f"F{r0+2}", "=0*1", fmt=FMT_NUM1)
        for i, c in enumerate(FC):
            prev = AN_COLS[3 + i - 1]
            set_formula(m, f"{c}{r0+2}", f"=+{prev}{r0+4}", fmt=FMT_NUM1)
        set_label(m, f"B{r0+3}", "Borrowings/(repayments)")
        for c in FC:
            set_link(m, f"{c}{r0+3}",
                     f"=+'Key Assumptions'!{amt_cell}*($C${r0+1}<={c}$7)"
                     f"*($C${r0+1}>{chr(ord(c)-1)}$7)", fmt=FMT_NUM1)
        set_label(m, f"B{r0+4}", "Ending balance")
        set_formula(m, f"F{r0+4}", "=0*1", fmt=FMT_NUM1)
        for c in FC:
            set_formula(m, f"{c}{r0+4}", f"=SUM({c}{r0+2}:{c}{r0+3})", fmt=FMT_NUM1)
        set_label(m, f"B{r0+5}", "Days outstanding in period")
        for c in FC:
            set_formula(m, f"{c}{r0+5}",
                        f"=IF(AND($C${r0+1}<={c}$7,$C${r0+1}>{chr(ord(c)-1)}$7),"
                        f"{c}$7-$C${r0+1},IF($C${r0+1}>{c}$7,0,365))", fmt=FMT_NUM)
        set_label(m, f"B{r0+6}", "Cash interest")
        for c in FC:
            set_link(m, f"{c}{r0+6}",
                     f"=+({c}{r0+2}+{c}{r0+3})*'Key Assumptions'!{rate_cell}"
                     f"*{c}{r0+5}/365", fmt=FMT_NUM1)
        set_label(m, f"B{r0+7}", "Issuance costs")
        for c in FC:
            set_link(m, f"{c}{r0+7}",
                     f"=-{c}{r0+3}*'Key Assumptions'!{cost_cell}*({c}{r0+3}>0)", fmt=FMT_NUM1)

    new_financing_block(80, "New Financing - 1 (development notes)",
                        "$E$33", "$E$34", "$E$35", "$E$36")
    new_financing_block(89, "New Financing - 2 (refinancing)",
                        "$E$39", "$E$40", "$E$41", "$E$42")

    # ---- Rollup
    set_label(m, "B98", "Debt Rollup", bold=True)
    set_label(m, "B99", "Ending balances:")
    for lbl, src_row, r in (("Revolving Credit Facility", 69, 100),
                            ("9.00% Senior Notes due 2029", 77, 101),
                            ("New Financing - 1", 84, 102),
                            ("New Financing - 2", 93, 103)):
        set_label(m, f"B{r}", lbl, indent=1)
        for c in (AN_COLS if src_row == 69 else ["F"] + FC):
            set_formula(m, f"{c}{r}", f"=+{c}{src_row}", fmt=FMT_NUM1)
    # actuals columns for notes/new fins (D,E): notes constant 300, new fin 0
    for c in ("D", "E"):
        set_link(m, f"{c}101", "=+'Key Assumptions'!$E$28", fmt=FMT_NUM1)
        set_formula(m, f"{c}102", "=0*1", fmt=FMT_NUM1)
        set_formula(m, f"{c}103", "=0*1", fmt=FMT_NUM1)
    set_label(m, "B110", "Total debt", bold=True)
    for c in AN_COLS:
        set_total(m, f"{c}110", f"=SUM({c}100:{c}103)", fmt=FMT_NUM1)

    set_label(m, "B112", "Net borrowings/(repayments) — forecast")
    for c in FC:
        set_formula(m, f"{c}112", f"=+{c}76+{c}83+{c}92", fmt=FMT_NUM1)
    set_label(m, "B118", "Total cash interest — forecast", bold=True)
    for c in FC:
        set_total(m, f"{c}118", f"=+{c}71+{c}78+{c}86+{c}95", fmt=FMT_NUM1)
    set_label(m, "B124", "Total issuance costs — forecast")
    for c in FC:
        set_formula(m, f"{c}124", f"=+{c}87+{c}96", fmt=FMT_NUM1)
    set_label(m, "Z48", "Demo simplifications: RCF drawn constant; no quarterly axis; "
                        "cash taxes off EBITDA-less-interest. Real builds follow the full "
                        "reference architecture.")
    m.freeze_panes = "D9"

    # ================================================================ Valuation Summary
    vs = wb.create_sheet("Valuation Summary", 4)
    margin_layout(vs, label_width=30)
    brand_title(vs, COMPANY, "Valuation Summary", rule_to_col="H", tab="output")
    set_label(vs, "B5", "Financing Scenario:")
    set_link(vs, "C5", "=+'Key Assumptions'!$D$6", fmt="General")
    set_label(vs, "B6", "P&L Ramp Assumptions:")
    set_link(vs, "C6", "='Key Assumptions'!$D$7", fmt="General")

    set_label(vs, "G8", "Base"); set_label(vs, "H8", "Bull")
    set_label(vs, "E10", "Actual FY25", bold=True)
    set_label(vs, "G10", "Projected FY'29", bold=True)
    brand_band(vs, 11, "E", "H")
    set_label(vs, "E11", "FY25"); set_label(vs, "G11", "Base"); set_label(vs, "H11", "Bull")
    for col in ("E", "G", "H"):
        vs.column_dimensions[col].width = 13.1
    spacer(vs, "D", 2.7); spacer(vs, "F", 2.7)

    # Both case columns pull LIVE from the ramp tab's always-computed case blocks —
    # no scenario-freeze pattern, no circular references, opens clean on any machine.
    set_label(vs, "B12", "EBITDA", bold=True)
    segs = (("Harborview", 13, "Model!$F59", "'P&L Ramp Assumptions'!P29",
             "'P&L Ramp Assumptions'!U29"),
            ("Canyon Crossing", 14, "Model!$F60", "'P&L Ramp Assumptions'!P30",
             "'P&L Ramp Assumptions'!U30"),
            ("Other Operations", 15, "Model!$F61+Model!$F62",
             "'P&L Ramp Assumptions'!P31+'P&L Ramp Assumptions'!P32",
             "'P&L Ramp Assumptions'!U31+'P&L Ramp Assumptions'!U32"))
    for name, r, ltm, base_pull, bull_pull in segs:
        set_label(vs, f"B{r}", name)
        set_link(vs, f"E{r}", f"=+{ltm}", fmt=FMT_NUM1)
        set_link(vs, f"G{r}", f"=+{base_pull}", fmt=FMT_NUM1)
        set_link(vs, f"H{r}", f"=+{bull_pull}", fmt=FMT_NUM1)
    set_label(vs, "B16", "Total EBITDA", bold=True)
    for cc in ("E", "G", "H"):
        set_total(vs, f"{cc}16", f"=SUM({cc}13:{cc}15)", fmt=FMT_NUM1)

    set_label(vs, "B18", "Multiple", bold=True)
    mults = {"G": (9.0, 9.0, 6.0), "H": (10.0, 10.0, 7.0)}
    for i, (name, r, *_) in enumerate(segs):
        set_label(vs, f"B{18 + i + 1}", name)
        set_label(vs, f"E{18 + i + 1}", "n/a")
        for cc in ("G", "H"):
            set_input(vs, f"{cc}{18 + i + 1}", mults[cc][i], fmt=FMT_MULT,
                      source="Asterozoa estimate — segment comps")
    set_label(vs, "B22", "Blended Multiple")
    for cc in ("G", "H"):
        set_formula(vs, f"{cc}22", f'=+IFERROR({cc}28/{cc}16,"nm")', fmt=FMT_MULT)
    set_link(vs, "E22", "=+'Cap Structure'!$C$19/E16", fmt=FMT_MULT)

    set_label(vs, "B24", "Enterprise Value", bold=True)
    for i, (name, r, *_) in enumerate(segs):
        set_label(vs, f"B{24 + i + 1}", name)
        set_label(vs, f"E{24 + i + 1}", "n/a")
        for cc in ("G", "H"):
            set_formula(vs, f"{cc}{24 + i + 1}", f"=+{cc}{13 + i}*{cc}{19 + i}", fmt=FMT_NUM)
    set_label(vs, "B28", "Total Enterprise Value", bold=True)
    for cc in ("G", "H"):
        set_total(vs, f"{cc}28", f"=SUM({cc}25:{cc}27)", fmt=FMT_NUM)
    set_link(vs, "E28", "=+'Cap Structure'!$C$19", fmt=FMT_NUM)

    set_label(vs, "B30", "Net Debt")
    set_link(vs, "E30", "=-Model!$F44", fmt=FMT_NUM)
    set_link(vs, "G30", "=-'P&L Ramp Assumptions'!P42", fmt=FMT_NUM)
    set_link(vs, "H30", "=-'P&L Ramp Assumptions'!U42", fmt=FMT_NUM)
    set_label(vs, "B31", "Equity Value", bold=True)
    for cc in ("E", "G", "H"):
        set_total(vs, f"{cc}31", f"=+{cc}28+{cc}30", fmt=FMT_NUM)

    set_label(vs, "B33", "Shares outstanding (mm)")
    set_input(vs, "E33", 30.0, fmt=FMT_NUM1, source="Demo — fictional share count")
    for cc in ("G", "H"):
        set_formula(vs, f"{cc}33", f"=+$E33+{cc}39", fmt=FMT_NUM1)
    set_label(vs, "B34", "Price per share")
    set_input(vs, "E34", 4.20, fmt=FMT_USD2, source="Demo — fictional price as of today")
    for cc in ("G", "H"):
        set_formula(vs, f"{cc}34", f"=+{cc}31/{cc}33", fmt=FMT_USD2)
    set_label(vs, "B35", "MOIC")
    for cc in ("G", "H"):
        set_formula(vs, f"{cc}35", f"=+{cc}34/$E$34", fmt=FMT_MULT)
    set_label(vs, "B36", "IRR (4-year hold)")
    for cc in ("G", "H"):
        set_formula(vs, f"{cc}36", f"=+{cc}35^(1/4)-1", fmt=FMT_PCT1)

    set_label(vs, "B39", "Assumed equity comp new share issuance")
    for cc in ("G", "H"):
        set_input(vs, f"{cc}39", 2.0, fmt=FMT_NUM1, source="Demo — dilution assumption")
    set_label(vs, "B41", "Note: Base and Bull columns are both fully live (per-case cash "
                         "walk on the P&L Ramp tab) — no circular references.")
    brand_footer(vs, 43)

    # ================================================================ Cap Structure
    cs = wb.create_sheet("Cap Structure", 5)
    margin_layout(cs, label_width=42)
    brand_title(cs, COMPANY, "Capital Structure", rule_to_col="K", tab="output")
    set_label(cs, "B5", "As of:")
    set_input(cs, "C5", dt.date(2025, 12, 31), fmt=FMT_DATE, source=SRC)
    for col, h in (("B", "(USD in Millions)"), ("C", "Amount"), ("D", "Price"),
                   ("E", "Mkt. Val."), ("F", "Interest"), ("G", "Maturity"),
                   ("H", "Rate"), ("I", "Yield"), ("J", "x Book"), ("K", "x Mkt")):
        set_label(cs, f"{col}6", h, bold=True)
    brand_band(cs, 6, "B", "K")
    for col, w in (("C", 10.6), ("D", 7), ("E", 11.7), ("F", 8), ("G", 10),
                   ("H", 8), ("I", 8), ("J", 7), ("K", 7)):
        cs.column_dimensions[col].width = w

    set_label(cs, "B8", "$30M Senior Secured Revolving Credit Facility (1)")
    set_link(cs, "C8", "=+'Key Assumptions'!$E$22", fmt=FMT_NUM1)
    set_label(cs, "D8", "n/a")
    set_formula(cs, "E8", "=+C8", fmt=FMT_NUM1)
    set_link(cs, "F8", "=+C8*('Key Assumptions'!$E$23+'Key Assumptions'!$E$24)", fmt=FMT_NUM1)
    set_input(cs, "G8", dt.date(2028, 6, 30), fmt=FMT_DATE, source=SRC)
    set_label(cs, "H8", "SOFR + 3%")
    set_label(cs, "B9", "9.00% Senior Notes due 2029")
    set_link(cs, "C9", "=+'Key Assumptions'!$E$28", fmt=FMT_NUM1)
    set_input(cs, "D9", 94.0, fmt=FMT_NUM1, source="Demo — fictional market quote")
    set_formula(cs, "E9", "=+C9*D9/100", fmt=FMT_NUM1)
    set_link(cs, "F9", "=+C9*'Key Assumptions'!$E$29", fmt=FMT_NUM1)
    set_input(cs, "G9", dt.date(2029, 6, 15), fmt=FMT_DATE, source=SRC)
    set_link(cs, "H9", "=+'Key Assumptions'!$E$29", fmt=FMT_PCT2)
    set_formula(cs, "I9", "=YIELD(TODAY()+2,G9,H9,D9,100,2)", fmt=FMT_PCT2)

    set_label(cs, "B11", "Total Debt", bold=True)
    for col in ("C", "E", "F"):
        set_total(cs, f"{col}11", f"=SUM({col}8:{col}9)", fmt=FMT_NUM1)
    set_formula(cs, "J11", "=+C11/$C$24", fmt=FMT_MULT)
    set_formula(cs, "K11", "=+E11/$C$24", fmt=FMT_MULT)
    set_label(cs, "B13", "Less: Cash and Equivalents")
    set_link(cs, "C13", "=-Model!$F35", fmt=FMT_NUM1)
    set_formula(cs, "E13", "=+C13", fmt=FMT_NUM1)
    set_label(cs, "B14", "Net Debt", bold=True)
    for col in ("C", "E"):
        set_total(cs, f"{col}14", f"=+{col}11+{col}13", fmt=FMT_NUM1)
    set_formula(cs, "J14", "=+C14/$C$24", fmt=FMT_MULT)
    set_formula(cs, "K14", "=+E14/$C$24", fmt=FMT_MULT)

    set_label(cs, "B16", "Market Cap")
    set_link(cs, "C16", "=+'Valuation Summary'!$E$33*'Valuation Summary'!$E$34", fmt=FMT_NUM1)
    set_formula(cs, "E16", "=+C16", fmt=FMT_NUM1)
    set_label(cs, "B19", "Enterprise Value", bold=True)
    set_label(cs, "B19", "Enterprise Value", bold=True)
    for col in ("C", "E"):
        set_total(cs, f"{col}19", f"=+{col}14+{col}16", fmt=FMT_NUM1)
    set_formula(cs, "J19", "=+C19/$C$24", fmt=FMT_MULT)
    set_formula(cs, "K19", "=+E19/$C$24", fmt=FMT_MULT)

    set_label(cs, "B22", "Operating Metrics", bold=True)
    set_label(cs, "B23", "FY25 Revenue")
    set_link(cs, "C23", f"=+Actuals!$F${tot_rev_row}", fmt=FMT_NUM1)
    set_label(cs, "B24", "FY25 Adj EBITDA")
    set_link(cs, "C24", f"=+Actuals!$F${tot_eb_row}", fmt=FMT_NUM1)

    set_label(cs, "B26", "Liquidity", bold=True)
    set_label(cs, "B27", "RCF Commitments")
    set_link(cs, "C27", "=+'Key Assumptions'!$E$21", fmt=FMT_NUM1)
    set_label(cs, "B28", "Less: Drawn")
    set_link(cs, "C28", "=-'Key Assumptions'!$E$22", fmt=FMT_NUM1)
    set_label(cs, "B29", "Plus: Cash and Equivalents")
    set_formula(cs, "C29", "=-C13", fmt=FMT_NUM1)
    set_label(cs, "B30", "Total Liquidity", bold=True)
    set_total(cs, "C30", "=SUM(C27:C29)", fmt=FMT_NUM1)

    set_label(cs, "B32", "Credit Metrics", bold=True)
    set_label(cs, "B33", "Gross Leverage")
    set_formula(cs, "C33", "=+C11/C24", fmt=FMT_MULT)
    set_label(cs, "B34", "Net Leverage")
    set_formula(cs, "C34", "=+C14/C24", fmt=FMT_MULT)
    set_label(cs, "B36", "Notes:")
    set_label(cs, "B37", "1. Demo facility. All instruments and prices are fictional.")
    brand_footer(cs, 39)

    # ================================================================ DISCLAIMER first
    disclaimer_tab(wb, 0)

    # ================================================================ D6 property-level example
    # Must be called after all Model sheet content is written and before wb.save().
    # Does not reset freeze_panes (Risk R5).
    build_property_level_example(wb, path)

    wb.save(path)
    print(f"written: {path}")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else str(
        Path(__file__).parent / "MRG-demo-output" / "MRG Model.xlsx")
    Path(out).parent.mkdir(parents=True, exist_ok=True)
    build(out)
