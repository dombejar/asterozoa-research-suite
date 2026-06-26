#!/usr/bin/env python3
"""audit_model.py — machine audit for Asterozoa models (see reference/audit-checklist.md).

Stages:
  structure   openpyxl pass on the FORMULA view: color protocol, hardcodes in calc zones,
              source comments, hidden sheets/rows/cols, external links, formula consistency
  values      openpyxl pass on RECALCED cached values: error literals + every 'Check' row ≈ 0
              (run scripts/recalc.py first, or pass --recalc to do it here)
  sweep       drive real Excel through every scenario × case combo: assert no errors per state,
              record key outputs, assert they move; ends on the default state and SAVES (this
              refreshes valuation freeze columns)

Usage:
  python3 audit_model.py model.xlsx                          # structure + values
  python3 audit_model.py model.xlsx --recalc                 # recalc in-place first
  python3 audit_model.py model.xlsx --actuals-only           # scope values/structure to Actuals
  python3 audit_model.py model.xlsx --sweep \
      --toggle "Key Assumptions!D5=1,2,3,4" --toggle "Key Assumptions!D6=Base,Bull,Other" \
      --output "Valuation Summary!G34" --output "Model!N104" \
      --default "Key Assumptions!D5=2" --default "Key Assumptions!D6=Base"

Sweep-evidence emission (D3 BU7 — feeds model_gate.py G49-G54):
  python3 audit_model.py model.xlsx --sweep \
      --emit-evidence sweep-evidence.json --config <model-config.json>

    --emit-evidence (with --sweep) writes the sweep-evidence.json that
    model_gate.py consumes. The schema MUST match what model_gate.py reads and
    what fixtures/build_fixtures.py produces:
      states            list of per-state records, EXACTLY the Cartesian product
                        of the swept axes (G49 grid-coverage gate), each with:
        axis_state        {"<Sheet>!<cell>": value} for every swept axis
        outputs           {ref: value} for every declared output (G50/G51/G52)
        errors            list of error literals found in the declared outputs +
                          min-cash/covenant refs for this state (G53). Note: the
                          scan covers declared refs only, not the full workbook
                          used-range; cells_scanned reflects the number of refs
                          actually read, not a whole-workbook count.
        cells_scanned     count of refs read for this state (outputs + min-cash
                          + covenant refs); positive int required by G49
        min_cash          numeric min-cash for the state (G54)
        covenant_flags    {ref: flag} covenant-flag refs for the state (G54)
        raw_refs          {ref: value} the min-cash + covenant refs (G54)
      default_restored  True — the sweep restored config.default_state (G49)

    When --config is given, the axes/outputs/default/min-cash/covenant refs are
    taken from the config so the emitted evidence aligns with the gate's config;
    --toggle/--output/--default still override per invocation.

Defaults for --toggle/--output match the reference architecture; override per model.
Exit codes: 0 clean, 1 critical findings, 2 environment failure.
"""
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

import excel_oracle

ERROR_LITERALS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}
BLUE = "FF0000FF"
GREEN = "FF008000"
CHECK_TOL = 0.5

DEFAULT_TOGGLES = ["Key Assumptions!D5=1,2,3,4", "Key Assumptions!D6=Base,Bull,Other"]
DEFAULT_OUTPUTS = ["Valuation Summary!G34", "Valuation Summary!G31", "Model!N104", "Model!N15"]
DEFAULT_STATE = ["Key Assumptions!D5=2", "Key Assumptions!D6=Base"]

CRIT, WARN, INFO = "CRITICAL", "WARNING", "INFO"
findings: list[tuple[str, str, str]] = []


def add(sev, where, msg):
    findings.append((sev, where, msg))


# --------------------------------------------------------------------- helpers

def is_simple_alias(formula: str) -> bool:
    """=+P26 style single-cell same-sheet ref (scenario-block alias) — blue allowed."""
    return bool(re.fullmatch(r"=\+?\$?[A-Z]{1,3}\$?\d+", formula))


def rel_pattern(formula: str, col_idx: int) -> str:
    """Normalize A1 refs to column-relative form so dragged rows compare equal."""
    from openpyxl.utils import column_index_from_string

    def sub(m):
        dollar_c, col, dollar_r, row = m.groups()
        if dollar_c:
            return f"${col}{dollar_r}{row}"
        return f"C[{column_index_from_string(col) - col_idx}]{dollar_r}{row}"

    return re.sub(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", sub, formula)


# --------------------------------------------------------------------- structure stage

def audit_structure(path: Path, only_sheet: str | None):
    wb = openpyxl.load_workbook(path, data_only=False)

    for ws in wb.worksheets:
        if only_sheet and ws.title != only_sheet:
            continue
        if ws.sheet_state != "visible":
            add(CRIT, ws.title, "hidden sheet")
        for dim in ws.row_dimensions.values():
            if dim.hidden:
                add(WARN, f"{ws.title} row {dim.index}", "hidden row")
        for key, dim in ws.column_dimensions.items():
            if dim.hidden:
                add(WARN, f"{ws.title} col {key}", "hidden column")

        rows = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                v = str(c.value)
                color = c.font.color.rgb if (c.font and c.font.color and
                                             isinstance(c.font.color.rgb, str)) else None
                is_formula = v.startswith("=")
                if is_formula:
                    if "[" in v:
                        add(CRIT, f"{ws.title}!{c.coordinate}", f"external workbook link: {v[:60]}")
                    if "!" in v and color != GREEN:
                        add(INFO, f"{ws.title}!{c.coordinate}",
                            "cross-sheet link not green font")
                    if color == BLUE and not is_simple_alias(v):
                        add(WARN, f"{ws.title}!{c.coordinate}",
                            f"blue-font formula (not a block alias): {v[:50]}")
                    rows.setdefault(c.row, []).append((c.column, v))
                else:
                    if isinstance(c.value, (int, float)) and color != BLUE:
                        add(WARN, f"{ws.title}!{c.coordinate}",
                            f"numeric constant without blue input font: {c.value}")
                    if ws.title == "Actuals" and isinstance(c.value, (int, float)) \
                            and c.comment is None:
                        add(WARN, f"{ws.title}!{c.coordinate}",
                            "Actuals input missing Source comment")

        # formula consistency within each row (dragged-row discipline)
        for r, cells in rows.items():
            if len(cells) < 3:
                continue
            pats = {}
            for col_idx, f in cells:
                pats.setdefault(rel_pattern(f, col_idx), []).append(col_idx)
            if len(pats) > 2:  # >2 distinct shapes in one row: worth an eyeball
                add(INFO, f"{ws.title} row {r}",
                    f"{len(pats)} distinct formula shapes across {len(cells)} cells "
                    "(expected: 1, plus deliberate stubs)")


# --------------------------------------------------------------------- values stage

def audit_values(path: Path, only_sheet: str | None):
    wb = openpyxl.load_workbook(path, data_only=True)
    wf = openpyxl.load_workbook(path, data_only=False)

    any_cached = False
    for ws in wb.worksheets:
        if only_sheet and ws.title != only_sheet:
            continue
        wsf = wf[ws.title]
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                any_cached = True
                if isinstance(c.value, str) and c.value in ERROR_LITERALS:
                    add(CRIT, f"{ws.title}!{c.coordinate}", f"formula error {c.value}")

        # Check rows: label 'Check' in A/B/C → all numeric cells in row must be ~0
        for r in range(1, ws.max_row + 1):
            label = next((ws.cell(row=r, column=k).value for k in (1, 2, 3)
                          if isinstance(ws.cell(row=r, column=k).value, str)), None)
            if label and label.strip().lower() == "check":
                for c in ws[r]:
                    if isinstance(c.value, (int, float)) and abs(c.value) > CHECK_TOL:
                        f = wsf[c.coordinate].value
                        add(CRIT, f"{ws.title}!{c.coordinate}",
                            f"Check row fails: {c.value:.3f} (formula: {str(f)[:60]})")

    if not any_cached:
        add(CRIT, "workbook", "no cached values — run scripts/recalc.py first (or --recalc)")


# --------------------------------------------------------------------- sweep stage

def parse_ref(spec: str):
    sheet, cell = spec.split("!")
    return sheet, cell


def _parse_axis_value(token: str):
    """Parse a single axis value token from a --toggle spec string, preserving the
    original JSON scalar type so sweep axis_state values match model_gate.py's
    declared config values exactly (G49 grid-coverage gate).

    Parsing order:
      1. integer literal  (e.g. "1", "-2")
      2. JSON scalar       (float, bool, null via json.loads — e.g. "1.5", "true")
      3. bare string       (e.g. "Base", "Bull")

    This prevents config axis values like 1.5 or true from being emitted as the
    strings "1.5" / "True" and failing G49's off-grid check."""
    t = token.strip()
    if t.lstrip("-").isdigit():
        return int(t)
    try:
        return json.loads(t.lower())   # handles true/false/null + numerics
    except (ValueError, AttributeError):
        pass
    return t


def sweep(path: Path, toggles: list[str], outputs: list[str], default_state: list[str],
          min_cash_refs: list[str] | None = None,
          covenant_refs: list[str] | None = None) -> dict:
    """Drive Excel across the full scenario grid via the OS-dispatch oracle
    (macOS AppleScript / Windows COM — excel_oracle.run_sweep). Returns a
    sweep-evidence dict in model_gate.py's schema (states + default_restored +
    oracle stamp). The structured record is a superset of the printed report;
    --emit-evidence writes the returned dict. Linux / Excel failure -> exit 2."""
    import itertools

    min_cash_refs = min_cash_refs or []
    covenant_refs = covenant_refs or []

    axes = []
    for t in toggles:
        ref, vals = t.split("=")
        sheet, cell = parse_ref(ref)
        parsed = [_parse_axis_value(v) for v in vals.split(",")]
        axes.append((sheet, cell, parsed))

    # extra refs read per state (after the outputs) for the G54 min-cash gate
    extra_refs = list(min_cash_refs) + list(covenant_refs)
    reads = list(outputs) + extra_refs

    default_parsed = []
    for d in default_state:
        ref, val = d.split("=")
        sheet, cell = parse_ref(ref)
        default_parsed.append((sheet, cell, _parse_axis_value(val)))

    # Drive Excel across the whole grid in one session (typed reads back).
    try:
        result = excel_oracle.run_sweep(path, axes, reads, default_parsed)
    except excel_oracle.NoExcelError as e:
        print(f"FATAL: {e}", file=sys.stderr)
        sys.exit(2)
    except excel_oracle.OracleError as e:
        print(f"FATAL: Excel sweep failed\n{e}", file=sys.stderr)
        sys.exit(2)

    grid = {}
    states = []
    print("\nSCENARIO SWEEP")
    print("state".ljust(28) + " | " + " | ".join(o.split("!")[1] for o in outputs))
    combos = list(itertools.product(*[vals for _, _, vals in axes]))
    for state, state_reads in zip(combos, result["states"]):
        out_vals = state_reads[:len(outputs)]
        extra_vals = state_reads[len(outputs):]
        grid[state] = tuple(out_vals)
        errs = [v for v in state_reads
                if isinstance(v, str) and v.strip() in ERROR_LITERALS]
        if errs:
            add(CRIT, f"sweep {state}", f"error value in outputs: {out_vals}")
        print(str(state).ljust(28) + " | "
              + " | ".join(str(v)[:12] for v in out_vals))

        # ---- structured record (model_gate.py sweep-evidence schema) ----
        axis_state = {f"{sheet}!{cell}": val
                      for (sheet, cell, _), val in zip(axes, state)}
        out_map = dict(zip(outputs, out_vals))
        extra_map = dict(zip(extra_refs, extra_vals))
        min_cash_vals = [extra_map[r] for r in min_cash_refs if r in extra_map]
        numeric_mc = [v for v in min_cash_vals if isinstance(v, (int, float))]
        covenant_flags = {r: extra_map.get(r) for r in covenant_refs}
        raw_refs = {r: extra_map.get(r) for r in extra_refs}
        states.append({
            "axis_state": axis_state,
            "outputs": out_map,
            "errors": [e.strip() for e in errs],
            "cells_scanned": len(outputs) + len(extra_refs),
            "min_cash": min(numeric_mc) if numeric_mc else None,
            "covenant_flags": covenant_flags,
            "raw_refs": raw_refs,
        })

    # outputs must move across states
    for i, o in enumerate(outputs):
        distinct = {g[i] for g in grid.values()}
        if len(distinct) == 1 and len(grid) > 1:
            add(CRIT, f"sweep output {o}",
                "identical across ALL states — toggle likely not wired to this output")

    print(f"restored default state {default_state}, saved.")
    return {
        "states": states,
        "default_restored": result.get("default_restored", True),
        **excel_oracle.oracle_stamp(excel_version=excel_oracle.excel_version()),
    }


# --------------------------------------------------------------------- main

def _sweep_params_from_config(config: dict):
    """Derive (toggles, outputs, default_state, min_cash_refs, covenant_refs) from a
    model config (the same config model_gate.py gates against), so the emitted
    sweep-evidence axis_state/outputs align with the gate's expectations."""
    toggles = []
    for ax in config.get("scenario_axes", []):
        vals = ",".join(str(v) for v in ax.get("values", []))
        toggles.append(f"{ax['sheet']}!{ax['cell']}={vals}")
    outputs = [o["ref"] for o in config.get("outputs", []) if o.get("ref")]
    default_state = []
    for ref, val in (config.get("default_state") or {}).items():
        default_state.append(f"{ref}={val}")
    min_cash_refs = list(config.get("min_cash_refs", []))
    covenant_refs = list(config.get("covenant_flag_refs", []))
    return toggles, outputs, default_state, min_cash_refs, covenant_refs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--recalc", action="store_true", help="run recalc.py --in-place first")
    ap.add_argument("--actuals-only", action="store_true")
    ap.add_argument("--sweep", action="store_true")
    ap.add_argument("--toggle", action="append", default=None)
    ap.add_argument("--output", action="append", default=None)
    ap.add_argument("--default", action="append", default=None)
    ap.add_argument("--config", metavar="PATH",
                    help="model config JSON; derives sweep axes/outputs/default + "
                         "min-cash/covenant refs so --emit-evidence aligns with the gate")
    ap.add_argument("--emit-evidence", metavar="PATH",
                    help="with --sweep: write sweep-evidence.json (model_gate.py G49-G54)")
    args = ap.parse_args()

    path = Path(args.workbook).expanduser().resolve()
    if not path.exists():
        print(f"FATAL: {path} not found", file=sys.stderr)
        sys.exit(2)

    if args.emit_evidence and not args.sweep:
        print("FATAL: --emit-evidence requires --sweep (the evidence is the sweep "
              "matrix; model_gate.py reads it for G49-G54)", file=sys.stderr)
        sys.exit(2)

    config = None
    if args.config:
        config = json.loads(Path(args.config).expanduser().read_text(encoding="utf-8"))

    # config-derived sweep params (CLI flags override per-axis when supplied)
    cfg_toggles = cfg_outputs = cfg_default = None
    cfg_min_cash = cfg_covenant = []
    if config is not None:
        cfg_toggles, cfg_outputs, cfg_default, cfg_min_cash, cfg_covenant = \
            _sweep_params_from_config(config)

    if args.recalc:
        rc = subprocess.run([sys.executable, str(Path(__file__).parent / "recalc.py"),
                             str(path), "--in-place"])
        if rc.returncode == 2:
            sys.exit(2)

    only = "Actuals" if args.actuals_only else None
    audit_structure(path, only)
    audit_values(path, only)

    if args.sweep:
        toggles = args.toggle or cfg_toggles or DEFAULT_TOGGLES
        sweep_outputs = args.output or cfg_outputs or DEFAULT_OUTPUTS
        default_state = args.default or cfg_default or DEFAULT_STATE
        evidence = sweep(path, toggles, sweep_outputs, default_state,
                         min_cash_refs=cfg_min_cash, covenant_refs=cfg_covenant)
        if args.emit_evidence:
            out_path = Path(args.emit_evidence).expanduser().resolve()
            out_path.write_text(json.dumps(evidence, indent=2), encoding="utf-8")
            print(f"evidence : wrote {out_path} ({len(evidence['states'])} states, "
                  f"default_restored={evidence['default_restored']})")
        # re-audit values after the sweep saved fresh state
        findings_before = len(findings)
        audit_values(path, only)
        if len(findings) > findings_before:
            add(WARN, "sweep", "new value findings after sweep save — review above")

    crit = [f for f in findings if f[0] == CRIT]
    warn = [f for f in findings if f[0] == WARN]
    info = [f for f in findings if f[0] == INFO]
    print(f"\nAUDIT: {len(crit)} critical, {len(warn)} warnings, {len(info)} info")
    for sev, where, msg in sorted(findings, key=lambda x: (x[0] != CRIT, x[0] != WARN)):
        print(f"  [{sev}] {where}: {msg}")
    print("\nVERDICT: " + ("CLEAN — ship-eligible" if not crit
                           else "BLOCKED — resolve every CRITICAL"))
    sys.exit(0 if not crit else 1)


if __name__ == "__main__":
    main()
