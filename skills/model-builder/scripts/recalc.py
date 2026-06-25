#!/usr/bin/env python3
"""recalc.py — recalculate an xlsx in real Microsoft Excel and report formula errors.

Excel is the oracle: it evaluates everything openpyxl can't (CHOOSE/VLOOKUP chains, YIELD,
the iterative-calc valuation-freeze pattern) and writes cached values back into the file so
downstream audits can read real numbers with openpyxl data_only=True.

Usage:
    python3 recalc.py <model.xlsx>              # recalc a temp copy, report (source untouched)
    python3 recalc.py <model.xlsx> --in-place   # recalc the file itself (use during builds)
    python3 recalc.py <model.xlsx> --keep       # keep the temp copy (prints its path)

Evidence emission (D3 BU7 — feeds model_gate.py G21/G22/G29):
    python3 recalc.py <model.xlsx> --in-place \
        --emit-evidence recalc-evidence.json \
        --config <model-config.json> \
        --canary "Model!N50"

    --emit-evidence writes the recalc-evidence.json that model_gate.py consumes.
    The schema MUST match what model_gate.py reads and what
    fixtures/build_fixtures.py clean_evidence() produces:
      path                       basename of the recalced (gated) workbook
      mtime_before/mtime_after   file mtime around the recalc (>=1s apart; G21)
      canary_cell                a formula cell whose value Excel re-evaluates (G22)
      canary_before/canary_after the canary value across recalc (G22 BLOCKS if equal;
                                 pick a cell whose value Excel genuinely changes on
                                 calculate-full-rebuild, e.g. in the freeze chain)
      iterative_calc_persisted   whether iterative calc is persisted in calcPr (G23)
      workbook_sha256            content hash of the recalced file (G21 binding)
      output_formulas            {ref: formula-string} per declared config output (G29)
      output_values              {ref: evaluated-value} per declared config output (G29)
      recalc_bound_default_state {ref: evaluated-value} the default state recalc saved (G29)

    --emit-evidence REQUIRES --config (to know which output refs to fingerprint)
    and works only with --in-place or --keep (the evidence must bind to the file
    model_gate.py is actually handed; a discarded temp copy cannot be gated).

Exit codes: 0 = zero formula errors; 1 = errors found; 2 = environment/Excel failure.
"""
import argparse
import hashlib
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

ERROR_LITERALS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!")

APPLESCRIPT = """
tell application "Microsoft Excel"
	open POSIX file "{path}"
	set wb to active workbook
	set iteration to true
	set max iterations to 100
	set max change to 0.001
	set calculation to calculation automatic
	calculate full rebuild
	save wb
	close wb saving no
	return "ok"
end tell
"""


def recalc_in_excel(path: Path) -> None:
    script = APPLESCRIPT.format(path=str(path))
    proc = subprocess.run(
        ["osascript", "-e", script], capture_output=True, text=True, timeout=600
    )
    if proc.returncode != 0 or "ok" not in proc.stdout:
        print(f"FATAL: Excel recalc failed\nstdout: {proc.stdout}\nstderr: {proc.stderr}",
              file=sys.stderr)
        sys.exit(2)


def scan_errors(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in ERROR_LITERALS:
                    errors.append(f"{ws.title}!{cell.coordinate} = {cell.value}")
    return errors


# --------------------------------------------------------------------- evidence

def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _split_ref(ref: str):
    """'Sheet!A1' -> ('Sheet', 'A1'). Strips $ from the coordinate."""
    sheet, coord = ref.rsplit("!", 1)
    return sheet.strip().strip("'"), coord.replace("$", "")


def _output_formula_fingerprint(path: Path, config: dict) -> dict:
    """Each declared output cell's FORMULA at recalc time, from the formula view.

    Mirrors fixtures/build_fixtures.output_formula_fingerprint EXACTLY (a non-
    formula output is omitted; model_gate.G29 fingerprints only formula outputs)."""
    wb = openpyxl.load_workbook(path, data_only=False)
    fp = {}
    for out in config.get("outputs", []):
        ref = out.get("ref")
        if not ref or "!" not in ref:
            continue
        sheet, coord = _split_ref(ref)
        if sheet in wb.sheetnames:
            v = wb[sheet][coord].value
            if isinstance(v, str) and v.startswith("="):
                fp[ref] = v
    return fp


def _output_value_fingerprint(path: Path, config: dict) -> dict:
    """Each declared output cell's EVALUATED value at recalc time, from the cached
    view. Mirrors fixtures/build_fixtures.output_value_fingerprint EXACTLY (every
    declared output is recorded, formula or not). Used for both `output_values`
    and `recalc_bound_default_state`."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ov = {}
    for out in config.get("outputs", []):
        ref = out.get("ref")
        if not ref or "!" not in ref:
            continue
        sheet, coord = _split_ref(ref)
        if sheet in wb.sheetnames:
            ov[ref] = wb[sheet][coord].value
    return ov


def _iterative_persisted(path: Path) -> bool:
    """Whether iterative calc is persisted in calcPr (G23 evidence)."""
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        cp = getattr(wb, "calculation", None)
        return bool(cp is not None and getattr(cp, "iterate", False))
    except Exception:
        return False


def _canary_value(path: Path, canary_ref: str):
    """Read the canary cell's cached (evaluated) value after recalc."""
    if not canary_ref or "!" not in canary_ref:
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet, coord = _split_ref(canary_ref)
    if sheet in wb.sheetnames:
        return wb[sheet][coord].value
    return None


def emit_evidence(target: Path, config: dict, canary_ref: str,
                  mtime_before: float, mtime_after: float,
                  canary_before, out_path: Path) -> None:
    """Write recalc-evidence.json in model_gate.py's EXACT schema.

    `target` is the recalced file model_gate.py will gate (the CLI positional
    `workbook`). The formula view and cached view are the same physical file read
    two ways — exactly the single-file production shape model_gate.py expects.

    CANARY REQUIREMENT (G22): model_gate.py BLOCKS when canary_before == canary_after.
    The canary cell must be a formula whose evaluated value changes across the recalc
    (i.e. its cached value differs between the pre-recalc and post-recalc read).
    For most production models the default `Model!N50` satisfies this because the
    iterative-calc freeze pattern always shifts it; for a fully-static model, pick
    a time-stamp cell or a cell that Excel's full-rebuild genuinely changes.
    A warning is printed below when the values match so the caller can pick a better
    canary before handing the evidence to model_gate."""
    canary_after = _canary_value(target, canary_ref)
    if canary_before is not None and canary_after is not None and canary_before == canary_after:
        print(
            f"WARNING: canary cell {canary_ref!r} value unchanged across recalc "
            f"({canary_before!r} == {canary_after!r}). "
            "model_gate.py G22 will BLOCK on this evidence. "
            "Choose a canary cell whose cached value Excel actually changes on "
            "calculate-full-rebuild (e.g. a cell in the iterative-calc freeze chain).",
            file=sys.stderr,
        )
    ov = _output_value_fingerprint(target, config)
    evidence = {
        "path": target.name,
        "mtime_before": mtime_before,
        "mtime_after": mtime_after,
        "canary_cell": canary_ref,
        "canary_before": canary_before,
        "canary_after": canary_after,
        "iterative_calc_persisted": _iterative_persisted(target),
        "workbook_sha256": _sha256_file(target),
        "output_formulas": _output_formula_fingerprint(target, config),
        "output_values": ov,
        # the recalc-bound default state IS the cached state Excel just saved,
        # so it equals output_values here (G29 default-state cross-check anchor).
        "recalc_bound_default_state": dict(ov),
    }
    out_path.write_text(json.dumps(evidence, indent=2), encoding="utf-8")
    print(f"evidence : wrote {out_path} ({len(evidence['output_formulas'])} formula outputs, "
          f"{len(evidence['output_values'])} valued outputs)")


def main() -> None:
    ap = argparse.ArgumentParser(prog="recalc.py", add_help=True,
                                 description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook")
    ap.add_argument("--in-place", action="store_true",
                    help="recalc the file itself (use during builds)")
    ap.add_argument("--keep", action="store_true",
                    help="keep the recalced temp copy (prints its path)")
    ap.add_argument("--emit-evidence", metavar="PATH",
                    help="write recalc-evidence.json (model_gate.py G21/G22/G29). "
                         "Requires --config; needs --in-place or --keep.")
    ap.add_argument("--config", metavar="PATH",
                    help="model config JSON (declares the output refs to fingerprint)")
    ap.add_argument("--canary", metavar="REF", default="Model!N50",
                    help="formula cell whose recalc Excel re-evaluates (G22 canary)")
    args = ap.parse_args()

    src = Path(args.workbook).expanduser().resolve()
    if not src.exists():
        print(f"FATAL: {src} not found", file=sys.stderr)
        sys.exit(2)

    if args.emit_evidence and not args.config:
        print("FATAL: --emit-evidence requires --config (to know which output refs "
              "to fingerprint)", file=sys.stderr)
        sys.exit(2)
    if args.emit_evidence and not (args.in_place or args.keep):
        print("FATAL: --emit-evidence needs --in-place or --keep: the evidence must "
              "bind to the file model_gate.py is handed, not a discarded temp copy",
              file=sys.stderr)
        sys.exit(2)

    if args.in_place:
        target = src
    else:
        tmp_dir = Path(tempfile.mkdtemp(prefix="recalc-"))
        target = tmp_dir / src.name
        shutil.copy2(src, target)

    # canary BEFORE recalc (read the pre-recalc cached value, if any)
    canary_before = _canary_value(target, args.canary) if args.emit_evidence else None
    mtime_before = target.stat().st_mtime

    recalc_in_excel(target)

    mtime_after = target.stat().st_mtime
    errors = scan_errors(target)

    print(f"workbook : {src.name}")
    print(f"recalced : {target}")
    print(f"errors   : {len(errors)}")
    for e in errors:
        print(f"  {e}")

    if args.emit_evidence:
        config = json.loads(Path(args.config).expanduser().read_text(encoding="utf-8"))
        emit_evidence(target, config, args.canary, mtime_before, mtime_after,
                      canary_before, Path(args.emit_evidence).expanduser().resolve())

    if target != src and not args.keep and not errors:
        shutil.rmtree(target.parent, ignore_errors=True)
    elif target != src:
        print(f"recalced copy kept at: {target}")

    print("status   : " + ("SUCCESS — zero formula errors" if not errors
                           else "FAIL — fix every error above"))
    sys.exit(0 if not errors else 1)


if __name__ == "__main__":
    main()
