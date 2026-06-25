#!/usr/bin/env python3
"""model_kit.py — openpyxl helpers encoding the Asterozoa model format grammar.

Every constant here was extracted from the FLL Full House Resorts reference model (Dec 2025; see
reference/format-grammar.md). Build with these helpers and the workbook is
format-identical to the house style by construction.

Quick demo (emits a mini-workbook exercising the grammar):
    python3 model_kit.py /tmp/kit-demo.xlsx
"""
from __future__ import annotations

import datetime as _dt

from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- color protocol
BLUE_INPUT = "FF0000FF"     # hardcoded input
GREEN_LINK = "FF008000"     # cross-sheet link
BLACK = "FF000000"          # same-sheet formula (theme default also fine)

FILL_INPUT_BLOCK = PatternFill("solid", fgColor="FFFFF2CC")   # scenario/case input zones
FILL_HEADER_BAND = PatternFill("solid", fgColor="FFEBF3FB")   # optional header banding

# ---------------------------------------------------------------- Asterozoa brand
# Palette sampled from the FLL thesis deck (see reference/brand-standards.md)
AZ_CHARCOAL = "FF2E2E2E"
AZ_TAN = "FFCBBBA1"
AZ_PANEL = "FFEEEEEE"
FILL_AZ_CHARCOAL = PatternFill("solid", fgColor=AZ_CHARCOAL)
FILL_AZ_TAN = PatternFill("solid", fgColor=AZ_TAN)
FILL_AZ_PANEL = PatternFill("solid", fgColor=AZ_PANEL)
TAN_RULE = Border(bottom=Side(style="medium", color=AZ_TAN))

CONFIDENTIAL_LINE = ("FOR INSTITUTIONAL INVESTOR USE. CONFIDENTIAL AND PROPRIETARY - "
                     "DO NOT DISTRIBUTE WITHOUT PRIOR WRITTEN PERMISSION.")

DISCLAIMER_TEXT = (
    "This document has been prepared for informational purposes only. It does not constitute "
    "investment advice, a solicitation, or an offer to buy or sell any security. Past "
    "performance is not indicative of future results. All investments involve risk, including "
    "the possible loss of principal. The information contained herein is based on sources "
    "believed to be reliable but is not guaranteed as to accuracy or completeness.\n\n"
    "This material is intended for the recipient only. Recipients should conduct their own "
    "due diligence and consult their own legal, tax, and financial advisors before making "
    "any investment decision. Nothing herein should be construed as legal, tax, or investment "
    "advice.\n\n"
    "Copyright (c) Asterozoa Capital. All rights reserved."
)

# ---------------------------------------------------------------- number formats (verbatim)
FMT_NUM = '_(* #,##0_);_(* \\(#,##0\\);_(* "   -"?_);_(@_)'
FMT_NUM1 = '_(* #,##0.0_);_(* \\(#,##0.0\\);_(* "   -"?_);_(@_)'
FMT_PCT1 = '* #,##0.0%;* \\(#,##0.0\\)%;* "   -"?_)'
FMT_PCT2 = '* #,##0.00%;* \\(#,##0.00\\)%;* "   -"?_)'
FMT_PCT0 = '* #,##0%;* \\(#,##0\\)%;* "   -"?_)'
FMT_MULT = '* #,##0.0\\x;* \\(#,##0.0\\)\\x;* "   -"?_)'
FMT_USD = '* "$"\\ #,##0_);* "$"\\ \\(#,##0\\);* """$"\\ \\ \\-""?_);_(@_)'
FMT_USD2 = '* "$"\\ #,##0.00_);* "$"\\ \\(#,##0.00\\);* """$"\\ \\ \\-""?_);_(@_)'
FMT_DATE = "m/d/yy;@"
FMT_YN = '"Y";"N";"N"'

TOP_BORDER = Border(top=Side(style="thin"))

# ---------------------------------------------------------------- cell writers

def set_input(ws, coord, value, fmt=FMT_NUM, source=None, bold=False):
    """Hardcoded input: blue font; optional Source comment (mandatory on Actuals/market data)."""
    c = ws[coord]
    c.value = value
    c.font = Font(color=BLUE_INPUT, bold=bold)
    c.number_format = fmt
    if source:
        c.comment = Comment(f"Source: {source}", "asterozoa-model-builder")
    return c


def set_formula(ws, coord, formula, fmt=FMT_NUM, bold=False):
    """Same-sheet formula: default/black font."""
    c = ws[coord]
    c.value = formula if formula.startswith("=") else "=" + formula
    c.font = Font(bold=bold)
    c.number_format = fmt
    return c


def set_link(ws, coord, formula, fmt=FMT_NUM, bold=False):
    """Cross-sheet link: green font."""
    c = ws[coord]
    c.value = formula if formula.startswith("=") else "=" + formula
    c.font = Font(color=GREEN_LINK, bold=bold)
    c.number_format = fmt
    return c


def set_label(ws, coord, text, bold=False, indent=0):
    c = ws[coord]
    c.value = ("    " * indent) + text if indent else text
    c.font = Font(bold=bold)
    return c


def set_total(ws, coord, formula, fmt=FMT_NUM):
    """Total row cell: bold + top border."""
    c = set_formula(ws, coord, formula, fmt=fmt, bold=True)
    c.border = TOP_BORDER
    return c

# ---------------------------------------------------------------- brand helpers

def brand_title(ws, company, purpose, rule_to_col="N", tab="none"):
    """Asterozoa title block: tan brand line, charcoal company + purpose, tan rule under B4.
    tab: 'assumption' (tan), 'output' (charcoal), 'none' (no tab color)."""
    b = ws["B2"]; b.value = "ASTEROZOA CAPITAL"
    b.font = Font(color=AZ_TAN, bold=True, size=10)
    c = ws["B3"]; c.value = company
    c.font = Font(color=AZ_CHARCOAL, bold=True, size=12)
    p = ws["B4"]; p.value = purpose
    p.font = Font(color=AZ_CHARCOAL, bold=True)
    from openpyxl.utils import column_index_from_string
    for j in range(2, column_index_from_string(rule_to_col) + 1):
        ws.cell(row=4, column=j).border = TAN_RULE
    if tab == "assumption":
        ws.sheet_properties.tabColor = AZ_TAN[2:]
    elif tab == "output":
        ws.sheet_properties.tabColor = AZ_CHARCOAL[2:]


def brand_band(ws, row, first_col, last_col):
    """Deck-style table header band: tan fill, white bold text (output tabs only)."""
    from openpyxl.utils import column_index_from_string
    for j in range(column_index_from_string(first_col), column_index_from_string(last_col) + 1):
        cell = ws.cell(row=row, column=j)
        cell.fill = FILL_AZ_TAN
        cell.font = Font(color="FFFFFFFF", bold=True)


def brand_footer(ws, row, col="B"):
    c = ws[f"{col}{row}"]
    c.value = "ASTEROZOA CAPITAL: CONFIDENTIAL"
    c.font = Font(color=AZ_TAN, bold=True, size=8)


def disclaimer_tab(wb, position=0):
    """Insert the standard branded DISCLAIMER tab."""
    ws = wb.create_sheet("DISCLAIMER", position)
    ws.column_dimensions["A"].width = 2.7
    ws.column_dimensions["B"].width = 200.7
    t = ws["B2"]; t.value = "Important Disclaimer"
    t.font = Font(color=AZ_CHARCOAL, bold=True, size=16)
    ws["B2"].border = TAN_RULE
    d = ws["B4"]; d.value = DISCLAIMER_TEXT
    d.alignment = d.alignment.copy(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 600
    ws.sheet_properties.tabColor = AZ_CHARCOAL[2:]
    return ws


# ---------------------------------------------------------------- sheet scaffolding

def sheet_header(ws, company, purpose, units="($ in Millions)",
                 scenario_echo=True, key_assumptions_sheet="Key Assumptions"):
    """Standard B2/B3 header block; optional live-scenario echo rows (B4/B5)."""
    set_label(ws, "B2", company, bold=True)
    set_label(ws, "B3", purpose, bold=True)
    if scenario_echo:
        set_label(ws, "B4", "Live Scenario:")
        set_link(ws, "C4", f"=+'{key_assumptions_sheet}'!$D$5", fmt="General")
        set_formula(ws, "D4",
                    f"=VLOOKUP(C4,'{key_assumptions_sheet}'!$B$9:$C$12,2,FALSE)", fmt="General")
        set_label(ws, "B5", "P&L Ramp Assumptions:")
        set_link(ws, "C5", f"='{key_assumptions_sheet}'!$D$6", fmt="General")
    set_label(ws, "B8" if scenario_echo else "B6", units)


def margin_layout(ws, label_width=33.9, notes_col=None, notes_width=50.7):
    """Column A margin (2.7) + wide label col B (+ optional far-right Notes column)."""
    ws.column_dimensions["A"].width = 2.7
    ws.column_dimensions["B"].width = label_width
    ws.column_dimensions["C"].width = 9.7
    if notes_col:
        ws.column_dimensions[notes_col].width = notes_width


def spacer(ws, col, width=1.7):
    ws.column_dimensions[col].width = width


def period_headers(ws, row, start_col, seed_date, n_periods, months=12,
                   labels=None, label_row=None, col_width=9.7):
    """EOMONTH chain off a single blue seed date; optional FY/Q label row beneath."""
    c0 = get_column_letter(start_col)
    set_input(ws, f"{c0}{row}", seed_date, fmt=FMT_DATE)
    for i in range(1, n_periods):
        col, prev = get_column_letter(start_col + i), get_column_letter(start_col + i - 1)
        set_formula(ws, f"{col}{row}", f"=EOMONTH({prev}{row},{months})", fmt=FMT_DATE)
    for i in range(n_periods):
        ws.column_dimensions[get_column_letter(start_col + i)].width = col_width
        if labels and label_row:
            set_label(ws, f"{get_column_letter(start_col + i)}{label_row}", labels[i], bold=True)
            ws[f"{get_column_letter(start_col + i)}{label_row}"].alignment = \
                ws[f"{get_column_letter(start_col + i)}{label_row}"].alignment.copy(horizontal="right")


def choose_live(ws, coord, toggle_abs, scenario_coords, fmt=FMT_NUM):
    """Live cell =CHOOSE($toggle, s1, s2, ...). scenario_coords e.g. ['G18','H18','I18','J18']."""
    return set_formula(ws, coord, f"=CHOOSE({toggle_abs},{','.join(scenario_coords)})", fmt=fmt)


def choose_live_named(ws, coord, name_toggle_abs, enum_range_abs, case_coords, fmt=FMT_NUM):
    """Live cell for a NAMED toggle: =CHOOSE(VLOOKUP(toggle, enum, 2, FALSE), c1, c2, ...)."""
    return set_formula(
        ws, coord,
        f"=CHOOSE(VLOOKUP({name_toggle_abs},{enum_range_abs},2,FALSE),{','.join(case_coords)})",
        fmt=fmt)


def check_row(ws, row, label_col, value_cols, lhs_template, rhs_template):
    """'Check' row: lhs-rhs per column; templates use {c} for the column letter.
    e.g. check_row(ws, 25, 'B', 'DEFGHIJ', '{c}24', "Actuals!{c}120")"""
    set_label(ws, f"{label_col}{row}", "Check")
    writer = set_link if ("!" in lhs_template or "!" in rhs_template) else set_formula
    for c in value_cols:
        writer(ws, f"{c}{row}", f"={lhs_template.format(c=c)}-{rhs_template.format(c=c)}")


def input_block(ws, top_left_row, top_left_col, n_rows, n_cols):
    """Apply the light-yellow fill to a scenario/case input rectangle."""
    for r in range(top_left_row, top_left_row + n_rows):
        for j in range(top_left_col, top_left_col + n_cols):
            ws.cell(row=r, column=j).fill = FILL_INPUT_BLOCK

# ---------------------------------------------------------------- demo

def _demo(path):
    """Mini-workbook exercising the grammar: inputs tab + engine tab with a check row."""
    import openpyxl
    wb = openpyxl.Workbook()

    ka = wb.active
    ka.title = "Key Assumptions"
    margin_layout(ka, label_width=30, notes_col="L")
    sheet_header(ka, "Demo Co.", "Key Assumptions", scenario_echo=False)
    set_label(ka, "B5", "Scenario Toggle:")
    set_input(ka, "D5", 1, fmt="General", source="User toggle")
    set_label(ka, "B9", "1"); set_label(ka, "C9", "Base world")
    set_label(ka, "B10", "2"); set_label(ka, "C10", "Refi world")
    set_label(ka, "B16", "Interest rate", bold=False)
    set_input(ka, "G16", 0.08, fmt=FMT_PCT2, source="Indenture")
    set_input(ka, "H16", 0.10, fmt=FMT_PCT2, source="Asterozoa estimate")
    choose_live(ka, "E16", "$D$5", ["G16", "H16"], fmt=FMT_PCT2)
    input_block(ka, 16, 7, 1, 2)

    m = wb.create_sheet("Model")
    margin_layout(m)
    sheet_header(m, "Demo Co.", "Financial Model")
    set_label(m, "B7", "Period Ended:")
    period_headers(m, 7, 4, _dt.date(2024, 12, 31), 4,
                   labels=["FY24", "FY25E", "FY26E", "FY27E"], label_row=8)
    set_label(m, "B10", "Revenue")
    set_input(m, "D10", 100.0, fmt=FMT_NUM1, source="FY24 10-K")
    for col, prev in (("E", "D"), ("F", "E"), ("G", "F")):
        set_formula(m, f"{col}10", f"=+{prev}10*1.05", fmt=FMT_NUM1)
    set_label(m, "B11", "Interest expense")
    for col in "DEFG":
        set_link(m, f"{col}11", f"=-{col}10*'Key Assumptions'!$E$16", fmt=FMT_NUM1)
    set_label(m, "B12", "Total", bold=True)
    for col in "DEFG":
        set_total(m, f"{col}12", f"=SUM({col}10:{col}11)", fmt=FMT_NUM1)
    check_row(m, 13, "B", "D", "{c}12", "({c}10+{c}11)")
    m.freeze_panes = "D9"

    wb.save(path)
    print(f"demo written: {path}")


if __name__ == "__main__":
    import sys
    _demo(sys.argv[1] if len(sys.argv) > 1 else "/tmp/model-kit-demo.xlsx")
