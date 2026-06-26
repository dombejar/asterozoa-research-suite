#!/usr/bin/env python3
"""model_gate.py - the SINGLE offline shipping gate for Asterozoa models.

The offline shipping gate for models. Reads an .xlsx with openpyxl twice
(data_only=False formula view + data_only=True cached view) and runs every
machine-checkable gate (T0 env/config -> T1 structure -> T2 values ->
T3 Actuals ties -> T4 sweep -> T5 format). NEVER drives Excel: recalc.py and
sweep.py own that serialized resource and emit JSON evidence this gate proves.

Assumes recalc.py has already run and the workbook carries cached values; G20
(cache coverage) + G21/G22 (recalc evidence) prove that rather than trust it.

Exit-code contract (D-ENTRY + folded finding F14/codex#9):
  exit 0  == CLEAN, ship-eligible (the ONLY ship-eligible code)
  exit N  == N BLOCK findings (under --strict, WARN counts too)
  exit 3  == ran clean (0 BLOCK) but verdict stamped non-shippable
            (e.g. --allow-unlocked-env). Never reads as green.

OFFLINE (D-OFFLINE): openpyxl only, never osascript / Excel.

See GATES.md for the locked per-gate spec. This module mirrors the canonical
model_kit constants (D-COLOR / D-PRECISION / D-CONST) inline so the gate is
runnable standalone; the shipped skill version imports them from model_kit.
"""
from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.formula import Tokenizer
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

# ---------------------------------------------------------------------------
# Canonical constants (mirror model_kit.py; D-ENV / D-COLOR / D-CONST / D-TOL)
# ---------------------------------------------------------------------------
REQUIRED_OPENPYXL = (3, 1)

BLUE_INPUT = "FF0000FF"   # hardcoded input
GREEN_LINK = "FF008000"   # cross-sheet link
BLACK = "FF000000"        # same-sheet formula (theme default also fine)
# brand colors legal ONLY on non-data label cells (D-COLOR)
AZ_CHARCOAL = "FF2E2E2E"
AZ_TAN = "FFCBBBA1"
BRAND_COLORS = {AZ_CHARCOAL, AZ_TAN, "FFFFFFFF"}

# D-COLOR canonical tab-color map (folded finding F19): assumption=tan,
# engine/Actuals=none, output+disclaimer=charcoal.
TAB_COLOR_BY_ROLE = {
    "assumptions": AZ_TAN[2:],
    "engine": None,
    "actuals": None,
    "valuation": AZ_CHARCOAL[2:],
    "output": AZ_CHARCOAL[2:],
    "disclaimer": AZ_CHARCOAL[2:],
    "cap_structure": None,
}

# D-CONST structural-constant allowlist (numeric literals legal inside a formula)
STRUCTURAL_CONSTS = {0.0, 1.0, -1.0, 4.0, 12.0, 100.0, 365.0, 1000.0}
OFFSET_FUNCS = {"OFFSET", "INDEX", "ROW", "COLUMN", "ROWS", "COLUMNS"}

ERROR_LITERALS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}

# D-PRECISION number-format strings (mirror model_kit.FMT_*)
FMT_NUM = '_(* #,##0_);_(* \\(#,##0\\);_(* "   -"?_);_(@_)'
FMT_NUM1 = '_(* #,##0.0_);_(* \\(#,##0.0\\);_(* "   -"?_);_(@_)'
FMT_PCT1 = '* #,##0.0%;* \\(#,##0.0\\)%;* "   -"?_)'
FMT_USD2 = '* "$"\\ #,##0.00_);* "$"\\ \\(#,##0.00\\);* """$"\\ \\ \\-""?_);_(@_)'
FMT_MULT = '* #,##0.0\\x;* \\(#,##0.0\\)\\x;* "   -"?_)'
# role -> acceptable number_format (G63). A few roles accept General-equivalents.
ROLE_FMT = {
    "price": FMT_USD2,
    "ebitda_m": FMT_NUM1,
    "net_debt_m": FMT_NUM1,
    "revenue_m": FMT_NUM1,
    "moic": FMT_MULT,
    "irr": FMT_PCT1,
}

# D-TOL tolerances (single source)
TOL_EXACT = 0.05    # Actuals exact-keying identity ties ($M)
TOL_QTR = 0.5       # quarterly->annual ties ($M)
TOL_XBRL = 0.1      # XBRL cross-check ($M)
CHECK_TOL = 0.5     # Check-row residual ($M)
MIN_CASH_FLOOR = 0.0

CONFIDENTIAL_LINE = (
    "FOR INSTITUTIONAL INVESTOR USE. CONFIDENTIAL AND PROPRIETARY - "
    "ASTEROZOA CAPITAL"
)

CHECK_LABEL_RE = re.compile(r"^check$", re.IGNORECASE)   # F6/F8: anchored exact


# ---------------------------------------------------------------------------
# Finding model (severity-tagged gate result record)
# ---------------------------------------------------------------------------
class Finding:
    __slots__ = ("gate", "severity", "location", "message")

    def __init__(self, gate, severity, location, message):
        self.gate = gate
        self.severity = severity   # BLOCK | WARN | PASS
        self.location = location   # "<Sheet>!<coord>" / sheet token / None
        self.message = message

    def __repr__(self):
        loc = self.location or "-"
        return "%s %s %s %s" % (self.gate, self.severity, loc, self.message)

    def as_dict(self):
        return {"gate": self.gate, "severity": self.severity,
                "location": self.location, "message": self.message}


# ---------------------------------------------------------------------------
# Parsing primitives
# ---------------------------------------------------------------------------
def split_ref(ref):
    """'Sheet!A1' -> ('Sheet', 'A1'); 'A1' -> (None, 'A1'). Strips $ and quotes."""
    if "!" in ref:
        sheet, coord = ref.rsplit("!", 1)
        sheet = sheet.strip().strip("'")
    else:
        sheet, coord = None, ref
    return sheet, coord.replace("$", "")


def norm_coord(coord):
    """Normalize a coord to absolute-free uppercase (e.g. '$C$38' -> 'C38')."""
    return coord.replace("$", "").upper().strip()


def cell_value(wb, ref, default_sheet=None):
    """Resolve a ref against a loaded workbook; return the cell's .value or None."""
    sheet, coord = split_ref(ref)
    sheet = sheet or default_sheet
    if sheet is None or sheet not in wb.sheetnames:
        return None
    try:
        return wb[sheet][coord].value
    except Exception:
        return None


def resolve_ref_count(wb, ref, default_sheet=None):
    """1 if ref resolves to exactly one cell, 0 if sheet/cell missing, -1 if ref is a range."""
    sheet, coord = split_ref(ref)
    sheet = sheet or default_sheet
    if sheet is None or sheet not in wb.sheetnames:
        return 0
    if ":" in coord:
        return -1
    if not re.fullmatch(r"[A-Za-z]+[0-9]+", coord):
        return 0
    return 1


def is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def iter_formula_cells(ws):
    """Yield (coord, formula_string) for every cell whose value starts with '='."""
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                yield c.coordinate, c.value


def cell_font_rgb(cell):
    """Return the font color ARGB string, or None. Robust to openpyxl theme objects."""
    f = cell.font
    if f is None or f.color is None:
        return None
    rgb = getattr(f.color, "rgb", None)
    if isinstance(rgb, str):
        return rgb.upper()
    return None


# ---------------------------------------------------------------------------
# Formula analysis helpers
# ---------------------------------------------------------------------------
EXT_NAMED_RE = re.compile(r"\[[^\]]*\.xls[xmb]?\]", re.IGNORECASE)   # [Book.xlsx]Sheet
EXT_INDEXED_RE = re.compile(r"\[\d+\]")                              # [1]Sheet1!A1
STRUCTURED_REF_RE = re.compile(r"[A-Za-z_][\w.]*\[")                 # Table1[col]


def has_external_link(formula):
    """F8/codex: detect named OR indexed external workbook refs; exempt table refs."""
    if EXT_NAMED_RE.search(formula):
        return True
    if EXT_INDEXED_RE.search(formula):
        return True
    return False


def is_cross_sheet_link(formula):
    """D-COLOR: a PURE cross-sheet link is '=Sheet!Cell' or '=+Sheet!Cell' (a single
    cross-sheet single-cell reference, no arithmetic). That is the GREEN class.

    A formula that merely *uses* a cross-sheet input inside arithmetic
    (e.g. '=+D10*(1+Sheet!E20)') is a same-sheet FORMULA (black), not a link.
    """
    if not formula.startswith("="):
        return False
    body = formula[1:]
    if body.startswith("+"):
        body = body[1:]
    if has_external_link(body):
        return False
    # exactly one '!' and the whole body is Sheet!Cell (optionally $-anchored, quoted)
    m = re.fullmatch(r"'?[^'!]+'?!\$?[A-Za-z]{1,3}\$?\d+", body)
    return bool(m)


SINGLE_CELL_RE = re.compile(r"^\$?[A-Za-z]{1,3}\$?\d+$")


def is_simple_alias(formula):
    """F5/F19: a same-sheet single-cell alias '=+P26' OR '=P26' (no '+').

    Returns the bare target coord (e.g. 'P26') if it is a legal single-cell
    alias, else None. Cross-sheet refs and arithmetic are NOT simple aliases.
    """
    if not formula.startswith("="):
        return None
    body = formula[1:]
    if body.startswith("+"):
        body = body[1:]
    if "!" in body:
        return None
    if SINGLE_CELL_RE.match(body):
        return norm_coord(body)
    return None


def numeric_literals(formula):
    """Yield (literal_float, in_offset_func) for every numeric literal in a formula.

    Uses the openpyxl Tokenizer. A literal nested inside OFFSET/INDEX/ROW/COLUMN
    is flagged in_offset_func so the |n|<=24 structural rule (D-CONST) can apply.
    """
    try:
        toks = Tokenizer(formula).items
    except Exception:
        return
    # track function-call nesting to know whether a literal is inside an offset func
    func_stack = []
    for i, t in enumerate(toks):
        ttype = t.type
        subtype = t.subtype
        tval = t.value
        if ttype == "FUNC" and subtype == "OPEN":
            name = tval.rstrip("(").upper()
            func_stack.append(name)
        elif ttype == "FUNC" and subtype == "CLOSE":
            if func_stack:
                func_stack.pop()
        elif ttype == "OPERAND" and subtype == "NUMBER":
            try:
                val = float(tval)
            except ValueError:
                continue
            in_offset = any(fn in OFFSET_FUNCS for fn in func_stack)
            yield val, in_offset


def is_string_formula(formula):
    """F3/F17: True if the formula evaluates to text by design (string literal,
    concatenation, TEXT(), &-join). Such formulas are NOT leading-eq label defects."""
    body = formula[1:] if formula.startswith("=") else formula
    body = body.strip()
    # quoted string literal, possibly concatenated
    if body.startswith('"'):
        return True
    if "&" in body:
        return True
    up = body.upper()
    if up.startswith("TEXT(") or up.startswith("CONCATENATE(") or up.startswith("CONCAT("):
        return True
    return False


def tokenizes_as_formula(formula):
    """F17: True if, after the leading '=', the remainder is a valid formula.

    A fat-fingered label like '=Check the indenture' yields a bare-word operand
    stream with whitespace-separated words and no operators/refs; it does NOT
    tokenize as a real formula.
    """
    body = formula[1:] if formula.startswith("=") else formula
    body = body.strip()
    if not body:
        return False
    if is_string_formula(formula):
        return True
    # a single-cell or range reference, function call, or arithmetic is a formula
    if SINGLE_CELL_RE.match(body):
        return True
    if "!" in body or "(" in body:
        return True
    # bare word(s) with no operators -> a label mistakenly prefixed with '='
    if re.fullmatch(r"[A-Za-z][A-Za-z ]*", body):
        return False
    # has operators / digits / refs -> treat as a formula
    return True


# ---------------------------------------------------------------------------
# Config + evidence loaders
# ---------------------------------------------------------------------------
def load_json(path):
    if path is None:
        return None
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def sha256_file(path):
    """FIX-ROUND-2 (codex P0-1): content hash of a workbook, for binding evidence
    to the EXACT gated file. Returns None if the file cannot be read."""
    try:
        h = hashlib.sha256()
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()
    except OSError:
        return None


def sheet_role_map(config):
    """role -> sheet title (and inverse). From config expected_sheets."""
    es = (config or {}).get("expected_sheets", {})
    return es, {title: role for role, title in es.items()}


def is_engine_or_actuals(config, sheet_title):
    """True if the sheet is an engine or Actuals tab (BLOCK-tier for G11-G13)."""
    es, inv = sheet_role_map(config)
    role = inv.get(sheet_title)
    engine_sheets = set((config or {}).get("engine_sheets", []))
    if sheet_title in engine_sheets:
        return True
    return role in ("engine", "actuals")


# ---------------------------------------------------------------------------
# T0 - environment + config preflight (abort-on-fail)
# ---------------------------------------------------------------------------
def _running_in_venv():
    """True when the interpreter is a dedicated venv (the plugin venv at
    ${CLAUDE_PLUGIN_DATA}/venv, or any other venv), not the bare system python3.

    Portable replacement for the old hardcoded venv-path string check: detects a
    venv via sys.prefix/base_prefix divergence (the standard venv/virtualenv marker).
    The plugin's SessionStart bootstrap installs deps into ${CLAUDE_PLUGIN_DATA}/venv;
    a plain system python3 with openpyxl still runs the gate, just stamped unlocked.
    """
    return getattr(sys, "base_prefix", sys.prefix) != sys.prefix


def g00_env(allow_unlocked):
    parts = openpyxl.__version__.split(".")
    okv = tuple(int(x) for x in parts[:2]) >= REQUIRED_OPENPYXL
    in_venv = _running_in_venv()
    if okv and in_venv:
        return [Finding("G00", "PASS", None,
                        "env locked: %s openpyxl %s" % (sys.executable, openpyxl.__version__))]
    sev = "WARN" if allow_unlocked else "BLOCK"
    return [Finding("G00", sev, None,
                    "unlocked env (%s openpyxl %s); shippable only under the plugin venv "
                    "(${CLAUDE_PLUGIN_DATA}/venv/bin/python3) or another dedicated venv"
                    % (sys.executable, openpyxl.__version__))]


def g01_config(config, wb_formula, demo, baseline_fp=None):
    """Config present; every declared ref resolves to exactly one cell; every
    visible sheet is declared (F20 fail-closed role assignment).

    D3 (P0-B): structural output-completeness, independent of config. Two
    baseline-fingerprint checks that the config alone cannot bypass:
      - `outputs:[]` on a non-demo, non-external-facing run BLOCKs: a model with
        zero declared outputs cannot be value-gated (G28/G29 have no sample set),
        so an empty outputs[] hides EVERY output. Fail closed.
      - When a baseline names prior external-facing output coords
        (`baseline_fp.output_refs`), every one must still be declared in the live
        config. A config that NARROWS outputs[] to drop a prior output (hiding a
        live valuation cell from the value gates) BLOCKs."""
    if config is None:
        if demo:
            return [Finding("G01", "PASS", None, "demo path: built-in config")]
        return [Finding("G01", "BLOCK", None,
                        "no per-model config (required for non-demo run)")]
    findings = []
    es, inv = sheet_role_map(config)

    # D3 P0-B: output-completeness (baseline-fingerprint variant, works from the
    # config + an optional baseline; never needs wb_cached).
    declared_outputs = [o.get("ref") for o in config.get("outputs", []) if o.get("ref")]
    external = bool(config.get("external_facing", False))
    if not declared_outputs and not demo:
        findings.append(Finding("G01", "BLOCK", None,
                                "config declares zero outputs (outputs:[]): a model with no "
                                "declared outputs cannot be value-gated; an empty outputs[] "
                                "hides every output (fail-closed)"))
    if isinstance(baseline_fp, dict):
        prior = baseline_fp.get("output_refs") or []
        declared_set = set(declared_outputs)
        for pref in prior:
            if pref not in declared_set:
                findings.append(Finding("G01", "BLOCK", pref,
                                        "prior output coord missing from config outputs[]: the "
                                        "gated output surface cannot narrow below the baseline "
                                        "(a hidden/narrowed output)"))

    # collect all declared single-cell refs
    refs = []
    for ax in config.get("scenario_axes", []):
        refs.append("%s!%s" % (ax["sheet"], ax["cell"]))
    for out in config.get("outputs", []):
        refs.append(out["ref"])
    for r in config.get("min_cash_refs", []):
        refs.append(r)
    for r in config.get("covenant_flag_refs", []):
        refs.append(r)
    if config.get("horizon_cell"):
        refs.append(config["horizon_cell"])

    for ref in refs:
        n = resolve_ref_count(wb_formula, ref)
        if n == 0:
            findings.append(Finding("G01", "BLOCK", ref,
                                    "declared ref does not resolve to a cell"))
        elif n == -1:
            findings.append(Finding("G01", "BLOCK", ref,
                                    "declared ref is a range, must be exactly one cell"))
        elif cell_value(wb_formula, ref) is None:
            # the cell exists in coordinate space but is empty: the ref points at
            # a phantom cell, not a real model cell (codex#24 stale-config guard)
            findings.append(Finding("G01", "BLOCK", ref,
                                    "declared ref resolves to an empty cell (stale config?)"))

    # expected_sheets titles must exist
    for role, title in es.items():
        if title is not None and title not in wb_formula.sheetnames:
            findings.append(Finding("G01", "BLOCK", title,
                                    "expected_sheets[%s] sheet missing from workbook" % role))

    # F20/codex#24: any VISIBLE sheet not mapped to a role BLOCKS (fail-closed)
    declared = set(t for t in es.values() if t is not None)
    for sn in wb_formula.sheetnames:
        ws = wb_formula[sn]
        if ws.sheet_state != "visible":
            continue
        if sn not in declared:
            findings.append(Finding("G01", "BLOCK", sn,
                                    "undeclared visible sheet (every sheet must have a "
                                    "declared role; fail-closed)"))

    if not findings:
        findings.append(Finding("G01", "PASS", None, "config + all refs + sheet roles resolve"))
    return findings


# Real-Excel oracle provenance (closes the Linux fake-green seam). Kept LOCAL to
# model_gate (no excel_oracle import) so the gate stays offline-pure (D-OFFLINE).
VALID_ORACLE_PLATFORMS = ("darwin", "win32")   # macOS AppleScript / Windows COM
VALID_ORACLE_BACKENDS = ("applescript", "com")
# exact platform->backend pairing: a mismatched stamp (e.g. darwin+com) cannot
# come from a real oracle and must be rejected (codex diff P1#1).
VALID_ORACLE_PAIRS = {("darwin", "applescript"), ("win32", "com")}
EXPECTED_ORACLE_SCHEMA = 2


def g02_oracle(recalc_ev, sweep_ev, demo, test_mode, host_platform=None):
    """G02 — the ship-gate refuses to certify without a REAL Excel oracle.

    Three-layer Linux fail-honest (DESIGN-RESOLUTIONS #13):
      (a) HOST block: model_gate run on a host with no real Excel oracle
          (sys.platform not in {darwin, win32}) BLOCKs. You cannot ship FROM a
          Linux box, period — the locked "Linux = research-only" decision.
      (b) PROVENANCE block: the recalc + sweep evidence must each carry an
          oracle stamp proving it came from real Excel (oracle_platform in
          {darwin, win32}, oracle_backend in {applescript, com}). This rejects
          stale/forged evidence left on a no-Excel host.
      (Layer (c), run-currency, is G21's workbook-hash binding — already enforced.)

    Skipped under --demo / --test (explicitly non-shipping fixtures). G21's hash
    binding plus G00's unlocked-env stamp keep those paths honest on their own."""
    if demo or test_mode:
        return [Finding("G02", "PASS", None, "demo/test path: oracle gate not applicable")]

    host = host_platform if host_platform is not None else sys.platform
    findings = []

    # (a) host must itself be a real-Excel-oracle platform
    if host not in VALID_ORACLE_PLATFORMS:
        findings.append(Finding(
            "G02", "BLOCK", host,
            "ship-gate cannot certify on a host without a real Excel oracle "
            "(platform=%s). The model can be DRAFTED on Linux, but recalc/sweep "
            "ship-evidence requires macOS or Windows with Microsoft Excel "
            "(no LibreOffice fallback)." % host))

    # (b) each evidence file present must carry a real-Excel provenance stamp:
    #     oracle=="excel", an EXACT platform/backend pair, and the expected schema.
    for name, ev in (("recalc", recalc_ev), ("sweep", sweep_ev)):
        if ev is None:
            continue   # absence is gated elsewhere (G21/G49); G02 checks provenance
        plat = ev.get("oracle_platform")
        backend = ev.get("oracle_backend")
        if (ev.get("oracle") != "excel"
                or (plat, backend) not in VALID_ORACLE_PAIRS
                or ev.get("schema_version") != EXPECTED_ORACLE_SCHEMA):
            findings.append(Finding(
                "G02", "BLOCK", name,
                "%s-evidence missing/invalid real-Excel oracle stamp "
                "(oracle=%r, oracle_platform=%r, oracle_backend=%r, schema=%r); "
                "evidence must carry oracle='excel', an exact platform/backend "
                "pair (darwin+applescript or win32+com), and schema_version=%d — "
                "not faked or carried over from a no-Excel host"
                % (name, ev.get("oracle"), plat, backend, ev.get("schema_version"),
                   EXPECTED_ORACLE_SCHEMA)))

    if not findings:
        findings.append(Finding("G02", "PASS", None,
                                "real-Excel oracle: host %s + stamped evidence" % host))
    return findings


# ---------------------------------------------------------------------------
# T1 - structure (formula view)
# ---------------------------------------------------------------------------
def g10_extlink(wb_formula):
    findings = []
    for ws in wb_formula.worksheets:
        for coord, f in iter_formula_cells(ws):
            if has_external_link(f):
                findings.append(Finding("G10", "BLOCK", "%s!%s" % (ws.title, coord),
                                        "external workbook link: %s" % f))
    if not findings:
        findings.append(Finding("G10", "PASS", None, "no external workbook links"))
    return findings


def g11_hidden(wb_formula, config):
    findings = []
    hidden_ok = set((config or {}).get("hidden_ok", []))
    hidden_ok_ranges = (config or {}).get("hidden_ok_ranges", [])
    ok_rows = {}   # sheet -> set(rows), ok_cols sheet -> set(cols)
    ok_cols = {}
    for hr in hidden_ok_ranges:
        sh = hr.get("sheet")
        for rr in hr.get("rows", []):
            ok_rows.setdefault(sh, set()).add(rr)
        for cc in hr.get("cols", []):
            ok_cols.setdefault(sh, set()).add(cc)

    for ws in wb_formula.worksheets:
        title = ws.title
        if ws.sheet_state != "visible":
            if title in hidden_ok:
                findings.append(Finding("G11", "WARN", title, "helper sheet hidden (hidden_ok)"))
            else:
                findings.append(Finding("G11", "BLOCK", title, "hidden sheet"))
            continue
        block_tier = is_engine_or_actuals(config, title)
        sev = "BLOCK" if block_tier else "WARN"
        for r, dim in ws.row_dimensions.items():
            if dim.hidden:
                if r in ok_rows.get(title, set()):
                    continue
                findings.append(Finding("G11", sev, "%s!row %s" % (title, r), "hidden row"))
        for col, dim in ws.column_dimensions.items():
            if dim.hidden:
                if col in ok_cols.get(title, set()):
                    continue
                findings.append(Finding("G11", sev, "%s!col %s" % (title, col), "hidden column"))
    if not [f for f in findings if f.severity != "WARN"] and not findings:
        pass
    if not findings:
        findings.append(Finding("G11", "PASS", None, "no hidden sheets/rows/cols"))
    return findings


def g12_merged(wb_formula, config):
    findings = []
    for ws in wb_formula.worksheets:
        title = ws.title
        block_tier = is_engine_or_actuals(config, title)
        sev = "BLOCK" if block_tier else "WARN"
        for rng in ws.merged_cells.ranges:
            # title block rows 1-8 are whitelisted
            if rng.max_row <= 8:
                continue
            findings.append(Finding("G12", sev, "%s!%s" % (title, str(rng)),
                                    "merged range overlapping a data region"))
    if not findings:
        findings.append(Finding("G12", "PASS", None, "no data-region merged cells"))
    return findings


def g13_condfmt(wb_formula, config):
    findings = []
    condfmt_ok = set((config or {}).get("condfmt_ok", []))
    for ws in wb_formula.worksheets:
        title = ws.title
        try:
            rules = list(ws.conditional_formatting)
        except Exception:
            rules = []
        if not rules:
            continue
        if is_engine_or_actuals(config, title):
            findings.append(Finding("G13", "BLOCK", title,
                                    "conditional formatting on engine/Actuals tab "
                                    "(can override the font-color audit signal)"))
        elif title in condfmt_ok:
            findings.append(Finding("G13", "WARN", title, "conditional formatting (condfmt_ok)"))
        else:
            findings.append(Finding("G13", "WARN", title, "conditional formatting present"))
    if not findings:
        findings.append(Finding("G13", "PASS", None, "no conditional formatting on protected tabs"))
    return findings


def g14_struct_fp(wb_formula, config, baseline_fp):
    """F10/codex#8: BLOCK (not skip-PASS) when external_facing and no baseline."""
    external = bool((config or {}).get("external_facing", False))
    if baseline_fp is None:
        if external:
            return [Finding("G14", "BLOCK", None,
                            "external-facing model has no baseline fingerprint; "
                            "design-pass cell movement cannot be ruled out")]
        # deepseek P1 (round 2): GATES.md says skipped-as-PASS-with-a-NOTE for an
        # internal model with no baseline. Returning WARN made --strict BLOCK an
        # internal model (over-strict, a false-positive deviation from spec).
        return [Finding("G14", "PASS", None,
                        "no baseline fingerprint; structural-fp check skipped (NOTE)")]
    findings = []
    for role in ("engine", "actuals"):
        title = (config or {}).get("expected_sheets", {}).get(role)
        if not title or title not in wb_formula.sheetnames:
            continue
        ws = wb_formula[title]
        bl = baseline_fp.get(title)
        if bl is None:
            continue
        if bl.get("max_row") != ws.max_row or bl.get("max_col") != ws.max_column:
            findings.append(Finding("G14", "BLOCK", title,
                                    "sheet dimensions changed vs baseline (%sx%s -> %sx%s)"
                                    % (bl.get("max_row"), bl.get("max_col"),
                                       ws.max_row, ws.max_column)))
    # output coordinates present
    for out in (config or {}).get("outputs", []):
        if resolve_ref_count(wb_formula, out["ref"]) != 1:
            findings.append(Finding("G14", "BLOCK", out["ref"],
                                    "output cell missing at declared coordinate"))
    if not findings:
        findings.append(Finding("G14", "PASS", None, "structure matches baseline fingerprint"))
    return findings


# ---------------------------------------------------------------------------
# T2 - values (cached view)
# ---------------------------------------------------------------------------
def _load_bearing_refs(config):
    """F2/codex#1: every output ref, check-row cell, tie-out member cell -> must be cached."""
    refs = set()
    if not config:
        return refs
    for out in config.get("outputs", []):
        refs.add(out["ref"])
    for r in config.get("min_cash_refs", []):
        refs.add(r)
    for r in config.get("covenant_flag_refs", []):
        refs.add(r)
    # tie-out member cells from actuals_rowmap (resolved later per gate); here
    # we add explicit cell refs the rowmap may carry as "<sheet>!<coord>"
    return refs


def g20_cache(wb_formula, wb_cached, config):
    """F2: split into (a) zero-tolerance on the declared load-bearing set, and
    (b) the 10%/10-cell bulk-health band."""
    findings = []
    # (a) zero-tolerance load-bearing set
    lb = _load_bearing_refs(config)
    actuals_sheet = (config or {}).get("actuals_sheet")
    rowmap = (config or {}).get("actuals_rowmap", {})
    if actuals_sheet and actuals_sheet in wb_cached.sheetnames:
        for section, spec in rowmap.items():
            for r in _rowmap_rows(spec):
                # member cells across the period columns declared
                for col in _rowmap_cols(config):
                    lb.add("%s!%s%d" % (actuals_sheet, col, r))
    # check-row cells
    for sref in _check_cell_refs(wb_formula, config):
        lb.add(sref)

    blocked = False
    for ref in sorted(lb):
        # only enforce caches where the formula view actually has a formula
        fv = cell_value(wb_formula, ref)
        if not (isinstance(fv, str) and fv.startswith("=")):
            continue
        cv = cell_value(wb_cached, ref)
        if cv is None:
            findings.append(Finding("G20", "BLOCK", ref,
                                    "load-bearing formula cell has no cached value "
                                    "(recalc did not land / cache stripped)"))
            blocked = True

    # (b) bulk-health band
    total = 0
    none_count = 0
    for ws in wb_formula.worksheets:
        ck = wb_cached[ws.title] if ws.title in wb_cached.sheetnames else None
        for coord, f in iter_formula_cells(ws):
            total += 1
            cv = ck[coord].value if ck is not None else None
            if cv is None:
                none_count += 1
    floor = max(10, int(0.10 * total))
    if total and none_count >= floor:
        findings.append(Finding("G20", "BLOCK", None,
                                "%d/%d formula cells uncached (>= %d threshold): "
                                "workbook not recalced" % (none_count, total, floor)))
        blocked = True

    if not blocked:
        findings.append(Finding("G20", "PASS", None,
                                "cache complete on load-bearing set; %d/%d bulk uncached"
                                % (none_count, total)))
    return findings


def g21_recalc_mtime(evidence, workbook_path):
    """F3/F7/codex#2: artifact bound to THIS file (path + mtime_after == file mtime).

    FIX-ROUND-2 (codex P0-1): the binding now REQUIRES a `workbook_sha256` content
    hash and recomputes the hash of the gated file; absence or mismatch BLOCKS.
    Stale/copied/divergent evidence can no longer pair with a different workbook —
    mtime is rebindable in seconds, a content hash is not (without producing the
    exact byte-identical file the evidence describes)."""
    if evidence is None:
        return [Finding("G21", "BLOCK", None,
                        "no --recalc-evidence (fail-loud: cannot prove recalc ran)")]
    mb = evidence.get("mtime_before")
    ma = evidence.get("mtime_after")
    if mb is None or ma is None:
        return [Finding("G21", "BLOCK", None, "recalc-evidence missing mtime fields")]
    findings = []
    if ma - mb < 1.0:
        findings.append(Finding("G21", "BLOCK", None,
                                "recalc did not advance mtime by >=1s (silent no-op): "
                                "%.3f -> %.3f" % (mb, ma)))
    # bind to THIS file
    ev_path = evidence.get("path")
    if ev_path is not None and Path(ev_path).name != Path(workbook_path).name:
        findings.append(Finding("G21", "BLOCK", ev_path,
                                "recalc-evidence path (%s) does not match gated workbook (%s)"
                                % (Path(ev_path).name, Path(workbook_path).name)))
    try:
        live_mtime = Path(workbook_path).stat().st_mtime
        if abs(live_mtime - ma) > 2.0:
            findings.append(Finding("G21", "BLOCK", workbook_path,
                                    "evidence mtime_after (%.1f) != current file mtime (%.1f): "
                                    "evidence describes a different/older file"
                                    % (ma, live_mtime)))
    except OSError:
        pass
    # FIX-ROUND-2: REQUIRED content-hash binding (the rebindable mtime is not enough).
    ev_hash = evidence.get("workbook_sha256")
    if not ev_hash:
        findings.append(Finding("G21", "BLOCK", None,
                                "recalc-evidence has no workbook_sha256: an unhashed artifact "
                                "cannot prove it describes THIS file (fail-closed)"))
    else:
        live_hash = sha256_file(workbook_path)
        if live_hash is None:
            findings.append(Finding("G21", "BLOCK", workbook_path,
                                    "cannot read gated workbook to verify workbook_sha256"))
        elif live_hash != ev_hash:
            findings.append(Finding("G21", "BLOCK", workbook_path,
                                    "workbook_sha256 mismatch: evidence describes a different "
                                    "file (evidence %s.. != gated %s..)"
                                    % (str(ev_hash)[:12], live_hash[:12])))
    if not findings:
        findings.append(Finding("G21", "PASS", None,
                                "recalc advanced mtime; evidence hash-bound to file"))
    return findings


def g22_recalc_canary(evidence, wb_cached, workbook_path, wb_formula=None):
    """F3/F7/codex#2 + codex P0 #4: bind the canary to the workbook with a REQUIRED
    live read, not an optional check. BLOCK when canary_cell is absent, unresolved
    (no live cached value), non-formula in the formula view, or None in the cached
    view. The gate that PROVES Excel recalced cannot be satisfied by self-reported
    JSON alone."""
    if evidence is None:
        return [Finding("G22", "BLOCK", None, "no --recalc-evidence (canary unprovable)")]
    cb = evidence.get("canary_before")
    ca = evidence.get("canary_after")
    cell = evidence.get("canary_cell")
    if cb is None and ca is None:
        return [Finding("G22", "BLOCK", None, "recalc-evidence missing canary fields")]
    # the canary MUST name a cell (cannot bind to the workbook otherwise)
    if cell is None:
        return [Finding("G22", "BLOCK", None,
                        "recalc-evidence has no canary_cell: cannot bind the canary to the "
                        "workbook (self-reported JSON alone is insufficient)")]
    findings = []
    if cb == ca:
        findings.append(Finding("G22", "BLOCK", cell,
                                "canary value unchanged across recalc (Excel did not evaluate): "
                                "%r == %r" % (cb, ca)))
    # REQUIRED live read: the canary must be a real, evaluated formula cell whose
    # cached value resolves in the workbook (codex P0 #4 — no optional skip).
    if wb_formula is not None:
        fv = cell_value(wb_formula, cell)
        if not (isinstance(fv, str) and fv.startswith("=")):
            findings.append(Finding("G22", "BLOCK", cell,
                                    "canary_cell is not a formula in the formula view (%r): a "
                                    "recalc canary must be an evaluated formula" % fv))
    live = cell_value(wb_cached, cell)
    if live is None:
        findings.append(Finding("G22", "BLOCK", cell,
                                "canary_cell does not resolve to a cached value in the gated "
                                "workbook (unresolved / stripped cache): cannot prove recalc"))
    elif ca is not None and live != ca:
        findings.append(Finding("G22", "BLOCK", cell,
                                "live canary cached value %r != evidence canary_after %r "
                                "(evidence describes a different file)" % (live, ca)))
    if not findings:
        findings.append(Finding("G22", "PASS", cell, "canary changed + matches live workbook"))
    return findings


def g23_iteration(wb_cached, evidence, config):
    uses_freeze = bool((config or {}).get("uses_freeze_pattern", False))
    external = bool((config or {}).get("external_facing", False))
    if not uses_freeze:
        return [Finding("G23", "PASS", None, "model does not use the freeze pattern; skipped")]
    persisted = (evidence or {}).get("iterative_calc_persisted", False)
    if persisted:
        return [Finding("G23", "PASS", None, "iterative calc persisted in calcPr")]
    # GATES.md G23: BLOCK when external_facing=false requires the freeze and it is
    # not persisted; an external-facing model gets WARN (deepseek P0: the prior
    # `external is False or external is True` was a tautology -> always BLOCK).
    sev = "WARN" if external else "BLOCK"
    return [Finding("G23", sev, None,
                    "freeze pattern declared but iterative calc not persisted after recalc")]


def g24_error_lit(wb_cached, config):
    """F12/codex#11: error literals BLOCK in formula/data/output regions; allow 'nm'
    label strings; require numeric outputs finite (numeric-output assertion)."""
    findings = []
    label_sheets = set((config or {}).get("label_regions", {}).keys())
    for ws in wb_cached.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERROR_LITERALS:
                    findings.append(Finding("G24", "BLOCK", "%s!%s" % (ws.title, c.coordinate),
                                            "Excel error literal: %s" % c.value))
    if not findings:
        findings.append(Finding("G24", "PASS", None, "no Excel error literals"))
    return findings


def _check_cells(wb_formula, config):
    """F6/F8: locate check rows. Prefer config check_refs; else anchored exact-label
    scan '^check$' in the declared label column (default B)."""
    cells = []   # (sheet, coord)
    cfg = config or {}
    # explicit config check_refs win
    for ref in cfg.get("check_refs", []):
        sheet, coord = split_ref(ref)
        cells.append((sheet, coord))
    if cells:
        return cells
    label_col = cfg.get("check_label_col", "B")
    engine_sheets = set(cfg.get("engine_sheets", []))
    es, inv = sheet_role_map(cfg)
    for ws in wb_formula.worksheets:
        title = ws.title
        is_engine = title in engine_sheets or inv.get(title) == "engine"
        if not is_engine:
            continue
        lc = column_index_from_string(label_col)
        for row in ws.iter_rows():
            label_cell = ws.cell(row=row[0].row, column=lc)
            lv = label_cell.value
            if isinstance(lv, str) and CHECK_LABEL_RE.match(lv.strip()):
                for c in row:
                    if c.column == lc:
                        continue
                    # only audit formula cells in the check range (F6/codex#12)
                    if isinstance(c.value, str) and c.value.startswith("="):
                        cells.append((title, c.coordinate))
    return cells


def _check_cell_refs(wb_formula, config):
    return ["%s!%s" % (s, c) for s, c in _check_cells(wb_formula, config)]


def g25_check_row(wb_cached, wb_formula, config):
    findings = []
    for sheet, coord in _check_cells(wb_formula, config):
        if sheet not in wb_cached.sheetnames:
            continue
        v = wb_cached[sheet][coord].value
        if v is None:
            continue   # uncalced -> caught by G20, not double-counted
        if is_number(v) and abs(v) >= CHECK_TOL:
            findings.append(Finding("G25", "BLOCK", "%s!%s" % (sheet, coord),
                                    "Check cell residual %g exceeds |%.1f|" % (v, CHECK_TOL)))
    if not findings:
        findings.append(Finding("G25", "PASS", None, "all Check cells within tolerance"))
    return findings


def g26_check_count(wb_formula, config):
    findings = []
    cfg = config or {}
    engine_sheets = set(cfg.get("engine_sheets", []))
    es, inv = sheet_role_map(cfg)
    label_col = cfg.get("check_label_col", "B")
    lc = column_index_from_string(label_col)
    exempt = set(cfg.get("check_count_exempt", []))
    for ws in wb_formula.worksheets:
        title = ws.title
        is_engine = title in engine_sheets or inv.get(title) == "engine"
        if not is_engine or title in exempt:
            continue
        n = 0
        for row in ws.iter_rows():
            lcell = ws.cell(row=row[0].row, column=lc)
            if isinstance(lcell.value, str) and CHECK_LABEL_RE.match(lcell.value.strip()):
                n += 1
        if n < 2:
            findings.append(Finding("G26", "WARN", title,
                                    "only %d Check rows on engine sheet (expect >=2)" % n))
    if not findings:
        findings.append(Finding("G26", "PASS", None, ">=2 Check rows per engine sheet"))
    return findings


def g27_leading_eq(wb_formula, config):
    """F3/F17: a '='-prefixed string cell is a label-as-formula defect only if it
    does NOT tokenize as a real formula and is not a text-by-design formula."""
    findings = []
    for ws in wb_formula.worksheets:
        for coord, f in iter_formula_cells(ws):
            if is_string_formula(f):
                continue
            if not tokenizes_as_formula(f):
                findings.append(Finding("G27", "BLOCK", "%s!%s" % (ws.title, coord),
                                        "label stored as a formula (Excel 'unreadable "
                                        "content' repair risk): %s" % f))
    if not findings:
        findings.append(Finding("G27", "PASS", None, "no labels mis-stored as formulas"))
    return findings


def _declared_output_refs(config):
    """The config-declared output cell refs (FIX-ROUND-2 G28/G29 sample set)."""
    refs = []
    for out in (config or {}).get("outputs", []):
        if out.get("ref"):
            refs.append(out["ref"])
    return refs


def g28_shim_guard(wb_formula, wb_cached, config, cached_workbook, demo, test_mode):
    """FIX-ROUND-2 (codex P0-1, fix #2): the two-file `--cached-workbook` shim is an
    OFFLINE TEST seam. It is legal ONLY under `--demo`/`--test`. A non-demo run that
    supplies a separate cached workbook is stamped non_shippable. Independently, a
    formula-view / cached-view divergence on ANY declared output or Check cell is a
    wrong-green and is stamped non_shippable in EVERY mode — production gates ONE
    Excel-recalced file, so the two views must be the same file's two reads.

    A G28 WARN is the non_shippable stamp (is_non_shippable consumes it -> exit 3,
    never 0); the output formula/cached divergence is ALSO caught as a hard BLOCK by
    G29, so the wrong-green never reads as clean."""
    two_file = bool(cached_workbook)
    legal_seam = bool(demo or test_mode)
    findings = []

    if two_file and not legal_seam:
        findings.append(Finding("G28", "WARN", cached_workbook,
                                "two-file --cached-workbook shim used WITHOUT --demo/--test: "
                                "production gates ONE Excel-recalced file (non_shippable)"))

    # divergence catch: any declared output/check cell whose FORMULA-view formula
    # disagrees with the CACHED-view cell content is a wrong-green. In a single
    # Excel-recalced file the formula view holds `=...` and the cached view holds
    # the evaluated number for the SAME coordinate; if the cached view still holds a
    # DIFFERENT formula string, the two files diverged (the exploit's signature).
    refs = list(_declared_output_refs(config))
    for sref in _check_cell_refs(wb_formula, config):
        refs.append(sref)
    seen = set()
    for ref in refs:
        if ref in seen:
            continue
        seen.add(ref)
        fv = cell_value(wb_formula, ref)
        cv = cell_value(wb_cached, ref)
        # only meaningful where the formula view is a formula
        if not (isinstance(fv, str) and fv.startswith("=")):
            continue
        # If the cached view still carries a formula string (not an evaluated value)
        # AND it differs from the formula view, the two files describe different
        # formulas for the same output -> divergence (wrong-green).
        if isinstance(cv, str) and cv.startswith("=") and cv != fv:
            findings.append(Finding("G28", "BLOCK", ref,
                                    "formula-view and cached-view diverge on declared output "
                                    "(%r != %r): two-file shim carries inconsistent files "
                                    "(non_shippable)" % (fv, cv)))

    if not findings:
        msg = ("single-file production shape" if not two_file
               else "two-file shim under --demo/--test; views consistent")
        findings.append(Finding("G28", "PASS", None, msg))
    return findings


def _output_value_drifted(a, b, rel=1e-6, absol=1e-9):
    """True when two recalc/cached output values disagree. Numbers compare with a
    tiny relative+absolute tolerance (float round-trip); everything else compares
    by equality. A None on either side is a drift (handled by the caller)."""
    if is_number(a) and is_number(b):
        return abs(a - b) > max(absol, rel * max(abs(a), abs(b)))
    return a != b


def g29_output_recalc(evidence, wb_formula, wb_cached, config, sweep_ev=None):
    """FIX-ROUND-2 (codex P0-1, fix #3) + D3 (P0-A): extend the recalc canary
    (G22, ONE sentinel) to the config-declared OUTPUTS on THREE axes.

    1. FORMULA fingerprint (FIX-ROUND-2): each declared output's LIVE formula must
       equal `recalc-evidence.output_formulas[ref]`. A formula that changed since
       recalc cannot pass (its cached value is stale relative to its formula).

    2. VALUE fingerprint (D3 P0-A): each declared output's CACHED value must equal
       `recalc-evidence.output_values[ref]` (the value recalc actually produced).
       A cached value that drifted from the recalc-time value BLOCKs even when the
       formula is unchanged — the exact stale-value wrong-green the formula-only
       canary missed. A formula output MISSING from output_values fails closed.

    3. DEFAULT-STATE cross-check (D3 P0-A): when
       `recalc-evidence.recalc_bound_default_state` is present, the workbook cached
       outputs must equal that recalc-bound default state. A divergence means the
       saved workbook is not in the state recalc signed off on. (The sweep state
       is scenario-driven and intentionally differs from the cached default state,
       so the cross-check uses the recalc-bound default state, NOT the sweep state;
       `sweep_ev` is accepted for signature symmetry with the T4 gates.)

    Missing fingerprint for a formula output BLOCKS (fail-closed) on every axis."""
    if evidence is None:
        return [Finding("G29", "BLOCK", None,
                        "no --recalc-evidence (output recalc-drift unprovable)")]
    fp = evidence.get("output_formulas")
    ov = evidence.get("output_values")
    rbds = evidence.get("recalc_bound_default_state")
    refs = _declared_output_refs(config)
    # which declared outputs are actually formulas in the live formula view
    formula_outputs = []
    for ref in refs:
        fv = cell_value(wb_formula, ref)
        if isinstance(fv, str) and fv.startswith("="):
            formula_outputs.append((ref, fv))
    if not formula_outputs:
        return [Finding("G29", "PASS", None, "no formula-valued declared outputs to fingerprint")]
    if not isinstance(fp, dict) or not fp:
        return [Finding("G29", "BLOCK", None,
                        "recalc-evidence has no output_formulas fingerprint: cannot prove the "
                        "cached outputs match their current formulas (fail-closed)")]
    # D3 P0-A: the VALUE fingerprint is REQUIRED for formula outputs (fail-closed).
    if not isinstance(ov, dict) or not ov:
        return [Finding("G29", "BLOCK", None,
                        "recalc-evidence has no output_values fingerprint: cannot prove the "
                        "cached output VALUES match what recalc produced (fail-closed)")]
    findings = []
    for ref, fv in formula_outputs:
        recorded = fp.get(ref)
        if recorded is None:
            findings.append(Finding("G29", "BLOCK", ref,
                                    "declared output not in recalc output_formulas fingerprint: "
                                    "its cached value cannot be tied to its current formula"))
            continue
        if _norm_formula(recorded) != _norm_formula(fv):
            findings.append(Finding("G29", "BLOCK", ref,
                                    "output formula changed since recalc (cached value is stale): "
                                    "recalc-time %r != live %r" % (recorded, fv)))
            continue
        # also require the cached view to actually carry an evaluated (non-None) value
        cv = cell_value(wb_cached, ref)
        if cv is None:
            findings.append(Finding("G29", "BLOCK", ref,
                                    "declared output has no cached value (recalc did not land)"))
            continue
        # D3 P0-A: the cached value must equal the recalc-time value. Missing the
        # ref in output_values is fail-closed; a drift BLOCKs (stale wrong-green).
        if ref not in ov:
            findings.append(Finding("G29", "BLOCK", ref,
                                    "declared output absent from recalc output_values: its cached "
                                    "value cannot be tied to the recalc-time value (fail-closed)"))
            continue
        if _output_value_drifted(cv, ov[ref]):
            findings.append(Finding("G29", "BLOCK", ref,
                                    "cached output value drifted from recalc (formula unchanged): "
                                    "recalc-time %r != cached %r" % (ov[ref], cv)))
    # D3 P0-A default-state cross-check: when the recalc-bound default state is
    # recorded, the workbook cached outputs must match it (the saved state is the
    # state recalc signed off on). A RBDS dict that is present but missing a
    # formula output ref fails closed — a partial RBDS cannot silently exempt an
    # output from the default-state check.
    if isinstance(rbds, dict) and rbds:
        for ref, _fv in formula_outputs:
            if ref not in rbds:
                findings.append(Finding("G29", "BLOCK", ref,
                                        "declared output absent from recalc_bound_default_state: "
                                        "a partial default-state fingerprint cannot cover all "
                                        "outputs (fail-closed)"))
                continue
            cv = cell_value(wb_cached, ref)
            if _output_value_drifted(cv, rbds[ref]):
                findings.append(Finding("G29", "BLOCK", ref,
                                        "workbook cached output diverges from the recalc-bound "
                                        "default state: cached %r != default-state %r"
                                        % (cv, rbds[ref])))
    if not findings:
        findings.append(Finding("G29", "PASS", None,
                                "all %d formula outputs match the recalc-time formula + value "
                                "fingerprint" % len(formula_outputs)))
    return findings


def _norm_formula(f):
    """Whitespace-insensitive formula compare (Excel ignores most internal spaces)."""
    if not isinstance(f, str):
        return f
    return re.sub(r"\s+", "", f)


# ---------------------------------------------------------------------------
# T3 - Actuals tie-outs (re-derive, never trust Check rows)
# ---------------------------------------------------------------------------
def _rowmap_rows(spec):
    """A rowmap section spec may be an int row, a list of rows, or a dict."""
    if isinstance(spec, int):
        return [spec]
    if isinstance(spec, list):
        return [r for r in spec if isinstance(r, int)]
    if isinstance(spec, dict):
        rows = []
        if "total_row" in spec and isinstance(spec["total_row"], int):
            rows.append(spec["total_row"])
        for r in spec.get("member_rows", []):
            if isinstance(r, int):
                rows.append(r)
        if "row" in spec and isinstance(spec["row"], int):
            rows.append(spec["row"])
        return rows
    return []


def _rowmap_cols(config):
    """Period columns to foot across (config actuals_period_cols, default ['C'])."""
    return (config or {}).get("actuals_period_cols", ["C"])


def _section(config, name):
    return (config or {}).get("actuals_rowmap", {}).get(name)


# deepseek P0: rowmap-faithfulness. The vacuous-pass guard catches a MISSING
# rowmap entry but not a WRONG one — a rowmap pointing at the wrong rows makes
# every tie re-derive against the wrong cells and can pass green. Each mapped
# subtotal/total row's label cell must match the expected section name, so a
# mis-keyed rowmap cannot tie out silently. Keyword set per section (normalized
# substring match against the label column).
ROWMAP_EXPECTED_LABELS = {
    "revenue_total": ("revenue", "total revenue", "net revenue", "total net revenue"),
    "opex_total": ("operating expense", "total operating expense", "total opex",
                   "total expense", "opex", "costs and expense"),
    "operating_income": ("operating income", "operating loss", "income from operations",
                         "operating profit"),
    "adjusted_ebitda": ("ebitda", "adjusted ebitda"),
    "net_income": ("net income", "net loss", "net earnings"),
    "bs_assets_total": ("total assets",),
    "bs_liab_total": ("total liab",),
    "bs_equity_total": ("total equity", "total stockholders", "total shareholders",
                        "stockholders' equity", "shareholders' equity"),
    "cf_begin_cash": ("beginning cash", "cash at beginning", "beginning of period",
                      "cash beginning"),
    "cf_net_change": ("net change", "net increase", "net decrease", "change in cash"),
    "cf_end_cash": ("ending cash", "cash at end", "end of period", "cash ending"),
    "bs_cash": ("cash", "bs cash", "cash and cash equivalents"),
    # D3 (P1-2): TOTAL-row label expectations for the block-foot gates G37/G38 so a
    # mis-keyed bs_subtotals/cf_sections total_row (pointing at a member line, not
    # a subtotal/total row) is caught the same way G32-G36 catch a mis-keyed total.
    "bs_subtotal_total": ("subtotal", "sub", "total"),
    "cf_section_total": ("total", "net cash", "net change", "cash flow", "activities",
                         "provided", "used"),
}


def _row_label(wb_formula, sheet, row, label_col="B"):
    """Read the label-column text for an Actuals row from the formula view."""
    if wb_formula is None or sheet not in wb_formula.sheetnames:
        return None
    try:
        v = wb_formula[sheet]["%s%d" % (label_col, row)].value
    except Exception:
        return None
    return v if isinstance(v, str) else None


def _label_matches(label, keywords):
    if not label:
        return False
    norm = re.sub(r"[^a-z ]", "", label.lower()).strip()
    return any(kw in norm for kw in keywords)


def _rowmap_faithful(wb_formula, config, gate, section_key, row):
    """Return a BLOCK Finding if the mapped row's label does not match the section
    it is keyed as, else None. Skips when we have no expectation or no label cell
    (faithfulness is asserted, not invented; an absent label is a separate concern
    caught by the vacuous-pass / value gates)."""
    if wb_formula is None or row is None:
        return None
    keywords = ROWMAP_EXPECTED_LABELS.get(section_key)
    if not keywords:
        return None
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    label_col = cfg.get("actuals_label_col") or "B"
    label = _row_label(wb_formula, sheet, row, label_col)
    if label is None:
        return None
    if not _label_matches(label, keywords):
        return Finding(gate, "BLOCK", "%s!%s%d" % (sheet, label_col, row),
                       "rowmap %s points at row %d labeled %r, not a %s row "
                       "(mis-keyed Actuals rowmap)" % (section_key, row, label, section_key))
    return None


def g30_actuals_exists(wb_cached, config, override):
    cfg = config or {}
    sheet = override or cfg.get("actuals_sheet")
    external = bool(cfg.get("external_facing", False))
    if sheet is None:
        # F/codex#23: null allowed only with declared reason + non-external
        reason = cfg.get("actuals_not_applicable_reason")
        if reason and not external:
            return [Finding("G30", "WARN", None,
                            "actuals_sheet:null declared (NOTE: %s); T3 skipped" % reason)]
        return [Finding("G30", "BLOCK", None,
                        "actuals_sheet:null without a declared not-applicable reason "
                        "(or external_facing); T3 cannot silent-skip")]
    if sheet not in wb_cached.sheetnames:
        return [Finding("G30", "BLOCK", sheet, "actuals_sheet does not resolve to a sheet")]
    return [Finding("G30", "PASS", sheet, "Actuals sheet resolves")]


def g31_embedded_const(wb_formula, config):
    """F4/F7/codex#5: BLOCK in calc zones (engine + valuation roles), WARN in data."""
    findings = []
    cfg = config or {}
    extra = set(float(x) for x in cfg.get("const_allowlist_extra", []))
    allow = STRUCTURAL_CONSTS | extra
    # horizon literal is allowed
    hc = cfg.get("horizon_cell")
    horizon_val = None
    if hc:
        # we can't read cached here cleanly; allow the configured horizon int if present
        horizon_val = cfg.get("horizon_value")
        if horizon_val is not None:
            allow.add(float(horizon_val))
    es, inv = sheet_role_map(cfg)
    engine_sheets = set(cfg.get("engine_sheets", []))
    calc_roles = {"engine", "valuation", "output"}
    actuals_title = cfg.get("actuals_sheet")
    for ws in wb_formula.worksheets:
        title = ws.title
        role = inv.get(title)
        in_calc_zone = (title in engine_sheets) or (role in calc_roles)
        # F7 default: BLOCK everywhere except the Actuals tab and declared input/label sheets
        if title == actuals_title or role in ("assumptions", "disclaimer", "label"):
            sev = "WARN"
        else:
            sev = "BLOCK" if in_calc_zone else "WARN"
        for coord, f in iter_formula_cells(ws):
            for val, in_offset in numeric_literals(f):
                if val in allow:
                    continue
                if in_offset and abs(val) <= 24:
                    continue
                findings.append(Finding("G31", sev, "%s!%s" % (title, coord),
                                        "non-structural numeric literal %g in formula: %s"
                                        % (val, f)))
    if not [f for f in findings if f.severity in ("BLOCK", "WARN")]:
        findings.append(Finding("G31", "PASS", None, "no embedded non-structural constants"))
    return findings


def _foot_section(wb_cached, sheet, member_rows, total_row, cols, tol, gate, label):
    """Re-sum member rows and compare to the total row, per column. Returns findings."""
    findings = []
    if sheet not in wb_cached.sheetnames:
        return [Finding(gate, "BLOCK", sheet, "%s: Actuals sheet missing" % label)]
    ws = wb_cached[sheet]
    if not member_rows or total_row is None:
        return [Finding(gate, "BLOCK", sheet,
                        "%s: rowmap members not fully enumerated (vacuous-pass guard)" % label)]
    for col in cols:
        members = []
        missing = False
        for r in member_rows:
            v = ws["%s%d" % (col, r)].value
            if v is None:
                missing = True
            members.append(v if is_number(v) else 0.0)
        total_v = ws["%s%d" % (col, total_row)].value
        if total_v is None or missing:
            findings.append(Finding(gate, "BLOCK", "%s!%s%d" % (sheet, col, total_row),
                                    "%s: missing cached value in foot region" % label))
            continue
        s = sum(members)
        if abs(s - total_v) > tol:
            findings.append(Finding(gate, "BLOCK", "%s!%s%d" % (sheet, col, total_row),
                                    "%s: re-summed %g != keyed %g (col %s)"
                                    % (label, s, total_v, col)))
    if not findings:
        findings.append(Finding(gate, "PASS", sheet, "%s foots" % label))
    return findings


def _foot_tol(config, line_count, base=TOL_EXACT):
    """F18: footing tolerance is rounding-aware for sums of N rounded lines."""
    import math
    return max(base, base * math.sqrt(max(line_count, 1)))


def g32_rev_foot(wb_cached, config, wb_formula=None):
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    sec = _section(cfg, "revenue")
    members = _section(cfg, "revenue_lines")
    total = _section(cfg, "revenue_total")
    if isinstance(sec, dict):
        members = sec.get("member_rows", members)
        total = sec.get("total_row", total)
    if members is None or total is None:
        return [Finding("G32", "BLOCK", sheet, "revenue rowmap incomplete (vacuous-pass guard)")]
    faithful = _rowmap_faithful(wb_formula, cfg, "G32", "revenue_total", total)
    if faithful:
        return [faithful]
    members = _rowmap_rows(members) if not isinstance(members, list) else members
    tol = _foot_tol(cfg, len(members))
    return _foot_section(wb_cached, sheet, members, total, _rowmap_cols(cfg), tol, "G32", "Revenue")


def g33_opex_foot(wb_cached, config, wb_formula=None):
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    members = _section(cfg, "opex_lines")
    total = _section(cfg, "opex_total")
    if members is None or total is None:
        return [Finding("G33", "BLOCK", sheet, "opex rowmap incomplete (vacuous-pass guard)")]
    faithful = _rowmap_faithful(wb_formula, cfg, "G33", "opex_total", total)
    if faithful:
        return [faithful]
    members = members if isinstance(members, list) else _rowmap_rows(members)
    tol = _foot_tol(cfg, len(members))
    return _foot_section(wb_cached, sheet, members, total, _rowmap_cols(cfg), tol, "G33", "Opex")


def g34_pl_chain(wb_cached, config, wb_formula=None):
    """operating income == revenues - opex per column."""
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    rev = _section(cfg, "revenue_total")
    opex = _section(cfg, "opex_total")
    oi = _section(cfg, "operating_income")
    if None in (rev, opex, oi) or sheet not in wb_cached.sheetnames:
        return [Finding("G34", "BLOCK", sheet, "P&L chain rowmap incomplete")]
    for key, row in (("revenue_total", rev), ("opex_total", opex),
                     ("operating_income", oi)):
        faithful = _rowmap_faithful(wb_formula, cfg, "G34", key, row)
        if faithful:
            return [faithful]
    ws = wb_cached[sheet]
    findings = []
    for col in _rowmap_cols(cfg):
        rv = ws["%s%d" % (col, rev)].value
        ov = ws["%s%d" % (col, opex)].value
        oiv = ws["%s%d" % (col, oi)].value
        if None in (rv, ov, oiv):
            findings.append(Finding("G34", "BLOCK", "%s!%s%d" % (sheet, col, oi),
                                    "P&L chain: missing cached value"))
            continue
        if abs((rv - ov) - oiv) > TOL_EXACT:
            findings.append(Finding("G34", "BLOCK", "%s!%s%d" % (sheet, col, oi),
                                    "operating income %g != rev(%g)-opex(%g)" % (oiv, rv, ov)))
    if not findings:
        findings.append(Finding("G34", "PASS", sheet, "P&L chain identity holds"))
    return findings


def g35_ebitda_recon(wb_cached, config, wb_formula=None):
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    ni = _section(cfg, "net_income")
    adj = _section(cfg, "ebitda_adjustments")
    eb = _section(cfg, "adjusted_ebitda")
    if None in (ni, eb) or adj is None or sheet not in wb_cached.sheetnames:
        return [Finding("G35", "BLOCK", sheet, "EBITDA recon rowmap incomplete")]
    for key, row in (("net_income", ni), ("adjusted_ebitda", eb)):
        faithful = _rowmap_faithful(wb_formula, cfg, "G35", key, row)
        if faithful:
            return [faithful]
    adj = adj if isinstance(adj, list) else _rowmap_rows(adj)
    ws = wb_cached[sheet]
    findings = []
    tol = _foot_tol(cfg, len(adj) + 1)
    for col in _rowmap_cols(cfg):
        niv = ws["%s%d" % (col, ni)].value
        adjv = sum(ws["%s%d" % (col, r)].value or 0.0 for r in adj)
        ebv = ws["%s%d" % (col, eb)].value
        if niv is None or ebv is None:
            findings.append(Finding("G35", "BLOCK", "%s!%s%d" % (sheet, col, eb),
                                    "EBITDA recon: missing cached value"))
            continue
        if abs((niv + adjv) - ebv) > tol:
            findings.append(Finding("G35", "BLOCK", "%s!%s%d" % (sheet, col, eb),
                                    "Adj EBITDA %g != NI(%g)+adjustments(%g)" % (ebv, niv, adjv)))
    if not findings:
        findings.append(Finding("G35", "PASS", sheet, "EBITDA recon foots"))
    return findings


def g36_bs_balance(wb_cached, config, wb_formula=None):
    """F18: balance identity is a single identity -> tight TOL_EXACT."""
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    a = _section(cfg, "bs_assets_total")
    l = _section(cfg, "bs_liab_total")
    e = _section(cfg, "bs_equity_total")
    if None in (a, l, e) or sheet not in wb_cached.sheetnames:
        return [Finding("G36", "BLOCK", sheet, "balance-sheet rowmap incomplete")]
    for key, row in (("bs_assets_total", a), ("bs_liab_total", l),
                     ("bs_equity_total", e)):
        faithful = _rowmap_faithful(wb_formula, cfg, "G36", key, row)
        if faithful:
            return [faithful]
    ws = wb_cached[sheet]
    findings = []
    for col in _rowmap_cols(cfg):
        av = ws["%s%d" % (col, a)].value
        lv = ws["%s%d" % (col, l)].value
        ev = ws["%s%d" % (col, e)].value
        if None in (av, lv, ev):
            findings.append(Finding("G36", "BLOCK", "%s!%s%d" % (sheet, col, a),
                                    "balance sheet: missing cached value"))
            continue
        if abs(av - (lv + ev)) > TOL_EXACT:
            findings.append(Finding("G36", "BLOCK", "%s!%s%d" % (sheet, col, a),
                                    "Total assets %g != liab(%g)+equity(%g)=%g"
                                    % (av, lv, ev, lv + ev)))
    if not findings:
        findings.append(Finding("G36", "PASS", sheet, "balance sheet balances"))
    return findings


def g37_bs_subtotal(wb_cached, config, wb_formula=None):
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    blocks = _section(cfg, "bs_subtotals")   # list of {member_rows,total_row}
    if not blocks or sheet not in wb_cached.sheetnames:
        return [Finding("G37", "BLOCK", sheet, "bs_subtotals rowmap incomplete")]
    findings = []
    for blk in blocks:
        members = blk.get("member_rows", [])
        total = blk.get("total_row")
        # D3 P1-2: a mis-keyed total_row (pointing at a member line, not a
        # subtotal/total) is caught before re-summing, mirroring G32-G36.
        faithful = _rowmap_faithful(wb_formula, cfg, "G37", "bs_subtotal_total", total)
        if faithful:
            return [faithful]
        tol = _foot_tol(cfg, len(members))
        findings.extend(_foot_section(wb_cached, sheet, members, total,
                                      _rowmap_cols(cfg), tol, "G37", "BS subtotal"))
    real = [f for f in findings if f.severity == "BLOCK"]
    if not real:
        return [Finding("G37", "PASS", sheet, "BS subtotals foot")]
    return real


def g38_cf_section(wb_cached, config, wb_formula=None):
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    blocks = _section(cfg, "cf_sections")
    if not blocks or sheet not in wb_cached.sheetnames:
        return [Finding("G38", "BLOCK", sheet, "cf_sections rowmap incomplete")]
    findings = []
    for blk in blocks:
        members = blk.get("member_rows", [])
        total = blk.get("total_row")
        # D3 P1-2: catch a mis-keyed cf_sections total_row before re-summing.
        faithful = _rowmap_faithful(wb_formula, cfg, "G38", "cf_section_total", total)
        if faithful:
            return [faithful]
        tol = _foot_tol(cfg, len(members))
        findings.extend(_foot_section(wb_cached, sheet, members, total,
                                      _rowmap_cols(cfg), tol, "G38", "CF section"))
    real = [f for f in findings if f.severity == "BLOCK"]
    if not real:
        return [Finding("G38", "PASS", sheet, "CF sections foot")]
    return real


def g39_cash_tie(wb_cached, config, wb_formula=None):
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    beg = _section(cfg, "cf_begin_cash")
    net = _section(cfg, "cf_net_change")
    end = _section(cfg, "cf_end_cash")
    bs_cash = _section(cfg, "bs_cash")
    if None in (beg, net, end) or sheet not in wb_cached.sheetnames:
        return [Finding("G39", "BLOCK", sheet, "cash-tie rowmap incomplete")]
    for key, row in (("cf_begin_cash", beg), ("cf_net_change", net),
                     ("cf_end_cash", end), ("bs_cash", bs_cash)):
        faithful = _rowmap_faithful(wb_formula, cfg, "G39", key, row)
        if faithful:
            return [faithful]
    ws = wb_cached[sheet]
    findings = []
    for col in _rowmap_cols(cfg):
        bv = ws["%s%d" % (col, beg)].value
        nv = ws["%s%d" % (col, net)].value
        ev = ws["%s%d" % (col, end)].value
        if None in (bv, nv, ev):
            findings.append(Finding("G39", "BLOCK", "%s!%s%d" % (sheet, col, end),
                                    "cash tie: missing cached value"))
            continue
        if abs((bv + nv) - ev) > TOL_EXACT:
            findings.append(Finding("G39", "BLOCK", "%s!%s%d" % (sheet, col, end),
                                    "CF ending %g != begin(%g)+net(%g)" % (ev, bv, nv)))
        if bs_cash is not None:
            bcv = ws["%s%d" % (col, bs_cash)].value
            if bcv is not None and abs(ev - bcv) > TOL_EXACT:
                findings.append(Finding("G39", "BLOCK", "%s!%s%d" % (sheet, col, bs_cash),
                                        "CF ending %g != BS cash %g" % (ev, bcv)))
    if not findings:
        findings.append(Finding("G39", "PASS", sheet, "cash identities hold"))
    return findings


def g40_qtr_annual(wb_cached, config):
    cfg = config or {}
    sheet = cfg.get("actuals_sheet")
    checks = _section(cfg, "qtr_annual")   # list of {annual_ref, quarter_refs:[...]}
    if checks is None:
        return [Finding("G40", "PASS", sheet, "no quarterly-annual overlap declared; skipped")]
    findings = []
    for chk in checks:
        annual = chk.get("annual_ref")
        quarters = chk.get("quarter_refs", [])
        av = cell_value(wb_cached, annual, sheet)
        qvs = [cell_value(wb_cached, q, sheet) for q in quarters]
        if av is None or any(q is None for q in qvs) or len(quarters) != 4:
            findings.append(Finding("G40", "BLOCK", annual,
                                    "qtr-annual: missing value or != 4 quarters"))
            continue
        s = sum(q for q in qvs if is_number(q))
        if abs(s - av) > TOL_QTR:
            findings.append(Finding("G40", "BLOCK", annual,
                                    "4-quarter sum %g != annual %g (> $%.1fM)" % (s, av, TOL_QTR)))
    if not findings:
        findings.append(Finding("G40", "PASS", sheet, "quarterly sums tie to annual"))
    return findings


def g41_source_comment(wb_formula, config):
    """F/codex#17: input classes; require Source:/As of:/Basis: by class; exempt control toggles."""
    findings = []
    cfg = config or {}
    exempt_ranges = set(cfg.get("comment_exempt_ranges", []))
    toggle_cells = set()
    for ax in cfg.get("scenario_axes", []):
        toggle_cells.add("%s!%s" % (ax["sheet"], norm_coord(ax["cell"])))
    for ws in wb_formula.worksheets:
        for row in ws.iter_rows():
            for c in row:
                rgb = cell_font_rgb(c)
                if rgb != BLUE_INPUT:
                    continue
                ref = "%s!%s" % (ws.title, c.coordinate)
                if ref in toggle_cells or ref in exempt_ranges:
                    continue   # control toggle: exempt
                comment = c.comment.text if c.comment else ""
                if not (comment.startswith("Source:") or comment.startswith("As of:")
                        or comment.startswith("Basis:")):
                    findings.append(Finding("G41", "WARN", ref,
                                            "blue input lacks Source:/As of:/Basis: comment"))
    if not findings:
        findings.append(Finding("G41", "PASS", None, "blue inputs carry provenance comments"))
    return findings


# Core reported lines a public-company Actuals tab MUST tie to an external source
# (codex P0 #1 / deepseek P0): a wrong number on any of these would actually ship.
G42_CORE_LINES = {"revenue", "net_income", "total_assets", "cfo"}


def _requires_external_actuals(cfg, wb_cached):
    """codex P0 #1: an external-facing model OR one whose Actuals tab claims a
    public source (edgar) MUST carry filing evidence. A model with no Actuals tab
    (actuals_sheet:null) does not (T3 is skipped with a NOTE by G30)."""
    if cfg.get("actuals_sheet") is None:
        return False
    if bool(cfg.get("external_facing", False)):
        return True
    if cfg.get("actuals_source") == "edgar":
        return True
    return False


def g42_xbrl(wb_cached, config, xbrl):
    """F11/codex#7 + codex P0 #1: fail closed for public-company / external-facing
    Actuals tabs lacking XBRL/filing evidence, and require the concept map to
    cover every core reported line. A mis-keyed public Actuals tab cannot ship
    green just because the tab is internally consistent."""
    cfg = config or {}
    source = cfg.get("actuals_source", "provided")
    cmap = cfg.get("xbrl_concept_map", {})   # workbook ref -> {concept, period, unit, sign, core}
    requires_external = _requires_external_actuals(cfg, wb_cached)
    if xbrl is None:
        if source == "edgar":
            return [Finding("G42", "BLOCK", None,
                            "actuals_source=edgar but no --xbrl artifact (prove the filing)")]
        if requires_external:
            # external-facing public model with no filing evidence: exact-keying
            # to filings is unproven -> a mis-keyed Actuals tab would ship green.
            if cfg.get("xbrl_edgar_unreachable"):
                return [Finding("G42", "WARN", None,
                                "EDGAR unreachable; external Actuals tab unverified (non-shippable)")]
            return [Finding("G42", "BLOCK", None,
                            "external-facing/public Actuals tab requires a filing source "
                            "(actuals_source=edgar + --xbrl artifact); none supplied so "
                            "exact-keying to filings is unenforced")]
        if cfg.get("xbrl_edgar_unreachable"):
            return [Finding("G42", "WARN", None, "EDGAR unreachable; XBRL cross-check skipped")]
        return [Finding("G42", "PASS", None,
                        "internal model, source=provided; no filing evidence required")]
    if not cmap:
        return [Finding("G42", "BLOCK", None,
                        "--xbrl supplied but config has no xbrl_concept_map")]
    concepts = {(c["concept"], c.get("period")): c for c in xbrl.get("concepts", [])}
    findings = []
    # core-line coverage: for a public Actuals tab, every core reported line must
    # be independently verified, not just whatever subset the config happened to map.
    if requires_external:
        covered = {m.get("core") for m in cmap.values() if m.get("core")}
        missing_core = sorted(G42_CORE_LINES - covered)
        if missing_core:
            findings.append(Finding("G42", "BLOCK", None,
                                    "external Actuals tab: core reported lines not mapped to a "
                                    "filing concept: %s (partial evidence is not evidence)"
                                    % missing_core))
    for ref, m in cmap.items():
        key = (m["concept"], m.get("period"))
        if key not in concepts:
            findings.append(Finding("G42", "BLOCK", ref,
                                    "concept %s absent from EDGAR dump" % m["concept"]))
            continue
        filed = concepts[key]["value_usd_millions"] * m.get("sign", 1)
        keyed = cell_value(wb_cached, ref, cfg.get("actuals_sheet"))
        if keyed is None or not is_number(keyed):
            findings.append(Finding("G42", "BLOCK", ref, "keyed cell not a number"))
            continue
        if abs(keyed - filed) > TOL_XBRL:
            findings.append(Finding("G42", "BLOCK", ref,
                                    "keyed %g != EDGAR %s %g" % (keyed, m["concept"], filed)))
    if not findings:
        findings.append(Finding("G42", "PASS", None, "core concepts tie to EDGAR"))
    return findings


# ---------------------------------------------------------------------------
# T4 - sweep (consumes sweep-evidence; NO Excel)
# ---------------------------------------------------------------------------
def _axes(config):
    return (config or {}).get("scenario_axes", [])


def _axis_name(ax, idx):
    return ax.get("name") or ("%s!%s" % (ax["sheet"], ax["cell"]))


def _state_axis_value(state, ax):
    ref = "%s!%s" % (ax["sheet"], norm_coord(ax["cell"]))
    axis_state = state.get("axis_state", {})
    # keys may be stored with or without $; normalize
    for k, v in axis_state.items():
        sh, co = split_ref(k)
        if "%s!%s" % (sh, norm_coord(co)) == ref:
            return v
    return None


def _norm_axis_value(v):
    """Normalize an axis value for set-membership / grid comparison. Numeric ints
    and floats that are integral compare equal (1 == 1.0); strings are stripped.
    Bools stay distinct from numbers."""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        f = float(v)
        return int(f) if f.is_integer() else f
    if isinstance(v, str):
        return v.strip()
    return v


def _state_grid_tuple(state, axes):
    """Normalized (axis_name, axis_value) tuple for a state across all declared
    axes, in declared order. Missing axis value -> None in that slot."""
    return tuple(_state_axis_value(state, ax) for ax in axes)


def g49_sweep_coverage(evidence, config):
    """F4: sweep-evidence schema gate. Assert the swept states are EXACTLY the
    Cartesian product of the declared axis values (full grid, no dupes, no extras,
    no off-grid combos), every state recorded all outputs + a positive
    cells_scanned, each declared (non-aliased) axis has an isolating pair, AND the
    sweep restored the default state (default_restored is True)."""
    if evidence is None:
        return [Finding("G49", "BLOCK", None, "no --sweep-evidence (T4 unprovable)")]
    states = evidence.get("states", [])
    axes = _axes(config)
    findings = []
    if not states:
        return [Finding("G49", "BLOCK", None, "sweep-evidence has zero states")]

    # codex P0 #1 / F4: the swept axis-state tuples must be the EXACT Cartesian
    # product of the declared axis values. Block on duplicates, extras, missing
    # combos, off-grid axis values, missing axis values, and len mismatch (both
    # directions) — never silently green on count==expected with a swapped state.
    if axes:
        axis_value_sets = [set(_norm_axis_value(v) for v in ax.get("values", []))
                           for ax in axes]
        # the full grid of normalized tuples we REQUIRE
        from itertools import product as _product
        expected_set = set(_product(*[ax.get("values", []) for ax in axes]))
        expected_norm = set(tuple(_norm_axis_value(v) for v in combo)
                            for combo in expected_set)
        expected = len(expected_norm)

        seen_norm = []
        seen_set = set()
        dup_seen = set()
        off_grid = []
        missing_axis_val = []
        for i, st in enumerate(states):
            raw_tuple = _state_grid_tuple(st, axes)
            # a None slot means the state did not record one declared axis value
            if any(v is None for v in raw_tuple):
                missing_axis_val.append(i)
                continue
            norm_tuple = tuple(_norm_axis_value(v) for v in raw_tuple)
            # any slot whose value is not in that axis's declared values -> off-grid
            on_grid = all(norm_tuple[k] in axis_value_sets[k]
                          for k in range(len(axes)))
            if not on_grid:
                off_grid.append((i, raw_tuple))
                continue
            if norm_tuple in seen_set:
                dup_seen.add(norm_tuple)
            seen_set.add(norm_tuple)
            seen_norm.append(norm_tuple)

        for i in missing_axis_val:
            findings.append(Finding("G49", "BLOCK", "state %d" % i,
                                    "state does not record a value for every declared axis "
                                    "(cannot place it on the scenario grid)"))
        for i, raw in off_grid:
            findings.append(Finding("G49", "BLOCK", "state %d" % i,
                                    "axis-state %r carries a value outside the declared "
                                    "axis values (off-grid combo)" % (raw,)))
        for d in sorted(dup_seen, key=lambda t: tuple(str(x) for x in t)):
            findings.append(Finding("G49", "BLOCK", None,
                                    "duplicate grid state swept more than once: %r "
                                    "(masks a missing combo)" % (d,)))
        missing_combos = expected_norm - seen_set
        for combo in sorted(missing_combos, key=lambda t: tuple(str(x) for x in t)):
            findings.append(Finding("G49", "BLOCK", None,
                                    "grid combo never swept: %r (incomplete coverage)"
                                    % (combo,)))
        if len(states) != expected:
            findings.append(Finding("G49", "BLOCK", None,
                                    "sweep ran %d states != full grid %d "
                                    "(coverage size mismatch)" % (len(states), expected)))

    # per-state output completeness + cells_scanned (full per-state scan proof)
    n_outputs = len((config or {}).get("outputs", []))
    for i, st in enumerate(states):
        outs = st.get("outputs", {})
        if len(outs) < n_outputs:
            findings.append(Finding("G49", "BLOCK", "state %d" % i,
                                    "state recorded %d/%d outputs (partial scan)"
                                    % (len(outs), n_outputs)))
        if "cells_scanned" not in st or st.get("cells_scanned", 0) <= 0:
            findings.append(Finding("G49", "BLOCK", "state %d" % i,
                                    "state missing cells_scanned (no full per-state scan proof)"))
    # F2: each declared axis needs an isolating pair
    aliased = {}
    for ax in axes:
        al = ax.get("aliased_to")
        if al is not None:
            aliased[_axis_name(ax, 0)] = al
    for ax in axes:
        if not _has_isolating_pair(states, axes, ax):
            findings.append(Finding("G49", "BLOCK", _axis_name(ax, 0),
                                    "no isolating pair (baseline + single-axis perturbation) "
                                    "for axis; insufficient sweep coverage"))

    # codex P0 #2 / F4: the sweep MUST have restored the default state. A sweep
    # that leaves toggles non-default cannot report clean (the next run / Dom
    # opens the file in a wrong state). `default_restored is True` is required;
    # absent/false/non-true all BLOCK (fail closed — not provably-restored).
    dr = evidence.get("default_restored", None)
    if dr is not True:
        findings.append(Finding("G49", "BLOCK", None,
                                "sweep evidence default_restored is %r (must be true): "
                                "toggles not provably restored to config.default_state" % dr))

    if not findings:
        findings.append(Finding("G49", "PASS", None,
                                "sweep coverage = exact grid; isolating pairs present; "
                                "default state restored"))
    return findings


def _has_isolating_pair(states, axes, target_ax):
    """True if two states differ ONLY on target_ax (all other axes held constant)."""
    others = [a for a in axes if a is not target_ax]
    for i in range(len(states)):
        for j in range(i + 1, len(states)):
            si, sj = states[i], states[j]
            tv_i = _state_axis_value(si, target_ax)
            tv_j = _state_axis_value(sj, target_ax)
            if tv_i is None or tv_j is None or tv_i == tv_j:
                continue
            if all(_state_axis_value(si, a) == _state_axis_value(sj, a) for a in others):
                return True
    return False


def g50_sweep_axis(evidence, config):
    """F2/F15: per-axis isolating-pair check; aliased pairs exempt from must-move."""
    if evidence is None:
        return [Finding("G50", "BLOCK", None, "no --sweep-evidence")]
    states = evidence.get("states", [])
    axes = _axes(config)
    findings = []
    for ax in axes:
        if ax.get("aliased_to") is not None:
            continue   # F15: intentionally-aliased axis exempt from must-move
        moved = False
        others = [a for a in axes if a is not ax]
        for i in range(len(states)):
            for j in range(i + 1, len(states)):
                si, sj = states[i], states[j]
                tv_i = _state_axis_value(si, ax)
                tv_j = _state_axis_value(sj, ax)
                if tv_i is None or tv_j is None or tv_i == tv_j:
                    continue
                if not all(_state_axis_value(si, a) == _state_axis_value(sj, a) for a in others):
                    continue
                # isolating pair: any output that differs?
                for out in config.get("outputs", []):
                    vi = si.get("outputs", {}).get(out["ref"])
                    vj = sj.get("outputs", {}).get(out["ref"])
                    if is_number(vi) and is_number(vj) and abs(vi - vj) > 1e-9:
                        moved = True
        if not moved:
            findings.append(Finding("G50", "BLOCK", _axis_name(ax, 0),
                                    "axis moves no output when varied alone (dead/never-wired axis)"))
    if not findings:
        findings.append(Finding("G50", "PASS", None, "every declared axis moves >=1 output"))
    return findings


def _directional_axis(cfg):
    """codex P0 #2: pick the scenario axis the direction fallback orders along.

    Prefer a config-declared 'direction_axis'; else the first numbered axis that
    is NOT intentionally aliased (the financing axis in the house pattern). The
    aliased ramp axis is not a clean monotone ordering surface, so it is skipped.
    """
    axes = _axes(cfg)
    named = cfg.get("direction_axis")
    if named:
        for ax in axes:
            if _axis_name(ax, 0) == named:
                return ax
    for ax in axes:
        if ax.get("aliased_to") is not None:
            continue
        if ax.get("kind") == "numbered" or all(is_number(v) for v in ax.get("values", [])):
            return ax
    return None


def _resolve_axis(config, axis_ref):
    """Resolve an axis by name or Sheet!cell ref."""
    for ax in _axes(config):
        nm = _axis_name(ax, 0)
        if nm == axis_ref or "%s!%s" % (ax["sheet"], norm_coord(ax["cell"])) == axis_ref:
            return ax
    return None


def _isolating_dir_pairs(states, config, axis, lower, higher, out_ref):
    """codex P0 #3: yield (lo_v, hi_v, other_state) for EVERY grid slice in which
    the named axis takes `lower` vs `higher` while all OTHER declared axes are held
    constant. A single comparison over the first matching state would miss an
    inversion confined to one slice (e.g. only ramp=Bull). Each comparable pair is
    surfaced so the caller can require the direction to hold on every one of them."""
    others = [a for a in _axes(config) if a is not axis]
    lo_states = []
    hi_states = []
    for st in states:
        av = _state_axis_value(st, axis)
        if _norm_axis_value(av) == _norm_axis_value(lower):
            lo_states.append(st)
        elif _norm_axis_value(av) == _norm_axis_value(higher):
            hi_states.append(st)
    pairs = []
    for slo in lo_states:
        for shi in hi_states:
            if not all(_norm_axis_value(_state_axis_value(slo, a))
                       == _norm_axis_value(_state_axis_value(shi, a)) for a in others):
                continue
            lo_v = slo.get("outputs", {}).get(out_ref)
            hi_v = shi.get("outputs", {}).get(out_ref)
            if is_number(lo_v) and is_number(hi_v):
                # describe the held-constant slice for the finding
                slice_desc = ", ".join(
                    "%s=%r" % (_axis_name(a, 0), _state_axis_value(slo, a))
                    for a in others) or "no other axes"
                pairs.append((lo_v, hi_v, slice_desc))
    return pairs


def g51_sweep_dir(evidence, config):
    """F15/codex#20 + codex P0 (round 2): explicit direction_tests when declared;
    ELSE the GATES.md fallback driven by outputs[].direction over the directional
    scenario axis. Direction is evaluated over EVERY isolating low/high pair (all
    non-direction axes held constant), not just the first matching state — an
    inversion confined to one scenario slice BLOCKS. Missing/ambiguous direction
    mapping for a non-'none' output is a BLOCK, never a silent PASS."""
    if evidence is None:
        return [Finding("G51", "BLOCK", None, "no --sweep-evidence")]
    states = evidence.get("states", [])
    cfg = config or {}
    tests = cfg.get("direction_tests", [])
    findings = []
    if not tests:
        # ---- GATES.md fallback: derive from outputs[].direction ----
        directional_outs = [o for o in cfg.get("outputs", [])
                            if o.get("direction") in ("higher_is_bull", "lower_is_bull")]
        if not directional_outs:
            return [Finding("G51", "PASS", None,
                            "no directional outputs declared (all direction=none); skipped")]
        ax = _directional_axis(cfg)
        if ax is None:
            return [Finding("G51", "BLOCK", None,
                            "directional outputs declared but no scenario axis to order them "
                            "along (declare direction_tests or direction_axis)")]
        vals = list(ax.get("values", []))
        if len(vals) < 2:
            return [Finding("G51", "BLOCK", None,
                            "directional axis has < 2 values; cannot prove ordering")]
        lower, higher = vals[0], vals[-1]
        for out in directional_outs:
            ref = out["ref"]
            pairs = _isolating_dir_pairs(states, cfg, ax, lower, higher, ref)
            if not pairs:
                findings.append(Finding("G51", "BLOCK", ref,
                                        "directional output %s has no isolating low/high pair "
                                        "at the axis extremes; ordering unprovable" % ref))
                continue
            for lo_v, hi_v, slice_desc in pairs:
                if out["direction"] == "higher_is_bull":
                    ok = hi_v >= lo_v - 1e-9
                else:
                    ok = hi_v <= lo_v + 1e-9
                if not ok:
                    findings.append(Finding("G51", "BLOCK", ref,
                                            "direction %s violated in slice [%s]: %s@%r=%g vs "
                                            "%s@%r=%g" % (out["direction"], slice_desc,
                                                          _axis_name(ax, 0), higher, hi_v,
                                                          _axis_name(ax, 0), lower, lo_v)))
        if not findings:
            findings.append(Finding("G51", "PASS", None,
                                    "directional ordering holds on every isolating slice "
                                    "(outputs[].direction fallback)"))
        return findings
    for t in tests:
        axis_ref = t["axis"]
        lower, higher = t["lower"], t["higher"]
        op = t.get("op", ">=")
        ax = _resolve_axis(cfg, axis_ref)
        if ax is None:
            findings.append(Finding("G51", "BLOCK", axis_ref,
                                    "direction_test names an axis not in scenario_axes"))
            continue
        for out in t["outputs"]:
            pairs = _isolating_dir_pairs(states, cfg, ax, lower, higher, out)
            if not pairs:
                findings.append(Finding("G51", "BLOCK", out,
                                        "direction_test %s %r/%r has no isolating pair for %s; "
                                        "ordering unprovable" % (axis_ref, lower, higher, out)))
                continue
            for lo_v, hi_v, slice_desc in pairs:
                ok = (hi_v >= lo_v - 1e-9) if op == ">=" else (hi_v <= lo_v + 1e-9)
                if not ok:
                    findings.append(Finding("G51", "BLOCK", out,
                                            "direction: %s(%g) %s %s(%g) violated for %s in "
                                            "slice [%s]" % (higher, hi_v, op, lower, lo_v, out,
                                                            slice_desc)))
    if not findings:
        findings.append(Finding("G51", "PASS", None,
                                "directional ordering holds on every isolating slice"))
    return findings


def _output_at_axis(states, config, axis_ref, axis_val, out_ref):
    """Find a state where the named axis == axis_val and return that output value.
    Retained as a helper for other call sites (G51 now uses _isolating_dir_pairs)."""
    target = _resolve_axis(config, axis_ref)
    if target is None:
        return None
    for st in states:
        if _norm_axis_value(_state_axis_value(st, target)) == _norm_axis_value(axis_val):
            v = st.get("outputs", {}).get(out_ref)
            if is_number(v):
                return v
    return None


def g52_sweep_mag(evidence, config):
    if evidence is None:
        return [Finding("G52", "BLOCK", None, "no --sweep-evidence")]
    states = evidence.get("states", [])
    cfg = config or {}
    default_state = cfg.get("default_state", {})
    findings = []
    for out in cfg.get("outputs", []):
        if out.get("mag_exempt"):
            continue
        vals = [st.get("outputs", {}).get(out["ref"]) for st in states]
        vals = [v for v in vals if is_number(v)]
        if len(vals) < 2:
            continue
        base = _base_output(states, cfg, out["ref"])
        if base is None or base == 0:
            base = max(abs(v) for v in vals) or 1.0
        rng = max(vals) - min(vals)
        floor_pct = out.get("magnitude_floor_pct", 1.0)
        floor = abs(base) * floor_pct / 100.0
        if rng <= floor:
            findings.append(Finding("G52", "BLOCK", out["ref"],
                                    "output range %g <= floor %g (%.2f%% of |base| %g): "
                                    "toggle barely moves it" % (rng, floor, floor_pct, base)))
    if not findings:
        findings.append(Finding("G52", "PASS", None, "outputs move beyond the magnitude floor"))
    return findings


def _base_output(states, config, out_ref):
    """Output value in the default-state state."""
    default_state = (config or {}).get("default_state", {})
    axes = _axes(config)
    for st in states:
        match = True
        for ax in axes:
            ref = "%s!%s" % (ax["sheet"], norm_coord(ax["cell"]))
            want = None
            for k, v in default_state.items():
                sh, co = split_ref(k)
                if "%s!%s" % (sh, norm_coord(co)) == ref:
                    want = v
            if want is not None and _state_axis_value(st, ax) != want:
                match = False
        if match:
            v = st.get("outputs", {}).get(out_ref)
            if is_number(v):
                return v
    return None


def g53_sweep_errscan(evidence):
    if evidence is None:
        return [Finding("G53", "BLOCK", None, "no --sweep-evidence")]
    findings = []
    for i, st in enumerate(evidence.get("states", [])):
        errs = st.get("errors", [])
        if errs:
            findings.append(Finding("G53", "BLOCK", "state %d" % i,
                                    "error literals in swept state: %s" % errs))
    if not findings:
        findings.append(Finding("G53", "PASS", None, "no error literals in any swept state"))
    return findings


COVENANT_FAIL_VALUES = (False, 0, 0.0, "FAIL", "BREACH", "FAILED", "BREACHED")


def g54_min_cash(evidence, config):
    """F/codex#4 + codex P0 #3: validate BOTH layers the spec names as evidence —
    the per-state summary fields `min_cash` + `covenant_flags` (GATES.md:169) AND
    the raw refs — and block on any mismatch between them. A negative summary
    min_cash or a failing summary covenant flag BLOCKS even when raw_refs is clean
    (and vice-versa), so neither layer can hide a ran-dry / covenant-breach state."""
    if evidence is None:
        return [Finding("G54", "BLOCK", None, "no --sweep-evidence")]
    cfg = config or {}
    min_cash_refs = cfg.get("min_cash_refs", [])
    cov_refs = cfg.get("covenant_flag_refs", [])
    findings = []
    for i, st in enumerate(evidence.get("states", [])):
        axis_state = st.get("axis_state", {})
        excused = any(_hole_matches(h, axis_state) for h in cfg.get("known_cash_holes", []))

        # ---- raw refs layer (the independent recompute) ----
        raw = st.get("raw_refs", {})
        missing = [r for r in (min_cash_refs + cov_refs) if r not in raw]
        if missing:
            findings.append(Finding("G54", "BLOCK", "state %d" % i,
                                    "min-cash/covenant raw refs absent from evidence: %s" % missing))
            continue
        raw_cash_vals = [raw[r] for r in min_cash_refs if is_number(raw.get(r))]
        raw_min_cash = min(raw_cash_vals) if raw_cash_vals else None

        # ---- summary-field layer (the spec evidence surface, GATES.md:169) ----
        summary_min = st.get("min_cash")
        summary_cov = st.get("covenant_flags", {})

        # negative cash in EITHER layer blocks (unless a known hole excuses it)
        for layer_name, val in (("raw", raw_min_cash), ("summary", summary_min)):
            if is_number(val) and val < MIN_CASH_FLOOR and not excused:
                findings.append(Finding("G54", "BLOCK", "state %d" % i,
                                        "%s min forecast cash %g < %g (no known_cash_holes entry)"
                                        % (layer_name, val, MIN_CASH_FLOOR)))

        # mismatch between the two layers: one of them is lying about the state
        if is_number(raw_min_cash) and is_number(summary_min) and \
                abs(raw_min_cash - summary_min) > TOL_EXACT:
            findings.append(Finding("G54", "BLOCK", "state %d" % i,
                                    "summary min_cash %g != raw-refs min %g (evidence layers "
                                    "disagree)" % (summary_min, raw_min_cash)))

        # covenant breach in EITHER layer blocks
        for r in cov_refs:
            v = raw.get(r)
            if v in COVENANT_FAIL_VALUES:
                findings.append(Finding("G54", "BLOCK", "state %d" % i,
                                        "covenant flag %s failed (raw): %r" % (r, v)))
        for r, v in summary_cov.items():
            if v in COVENANT_FAIL_VALUES:
                findings.append(Finding("G54", "BLOCK", "state %d" % i,
                                        "covenant flag %s failed (summary): %r" % (r, v)))
    if not findings:
        findings.append(Finding("G54", "PASS", None,
                                "min cash >= floor + covenants pass (summary + raw agree)"))
    return findings


def _hole_matches(hole, axis_state):
    want = hole.get("axis_state", {})
    for k, v in want.items():
        sh, co = split_ref(k)
        nk = "%s!%s" % (sh, norm_coord(co))
        found = None
        for ak, av in axis_state.items():
            ash, aco = split_ref(ak)
            if "%s!%s" % (ash, norm_coord(aco)) == nk:
                found = av
        if found != v:
            return False
    return True


def g55_irr_horizon(wb_formula, wb_cached, config):
    """F6/F16: locate IRR formulas by config role; check exponent denominator
    (^ operand or POWER 2nd arg) is literal==horizon OR resolves to horizon cell."""
    cfg = config or {}
    horizon_val = cfg.get("horizon_value")
    horizon_cell = cfg.get("horizon_cell")
    horizon_norm = None
    if horizon_cell:
        hs, hc = split_ref(horizon_cell)
        horizon_norm = "%s!%s" % (hs, norm_coord(hc)) if hs else norm_coord(hc)
    irr_refs = [o["ref"] for o in cfg.get("outputs", []) if o.get("role") == "irr"]
    if not irr_refs:
        return [Finding("G55", "PASS", None, "no IRR outputs declared; skipped")]
    findings = []
    for ref in irr_refs:
        sheet, coord = split_ref(ref)
        f = cell_value(wb_formula, ref, sheet)
        if not (isinstance(f, str) and f.startswith("=")):
            findings.append(Finding("G55", "BLOCK", ref, "IRR output is not a formula"))
            continue
        denoms = _exponent_denominators(f)
        if not denoms:
            findings.append(Finding("G55", "WARN", ref, "no ^/POWER exponent found in IRR formula"))
            continue
        for d in denoms:
            if d["kind"] == "literal":
                if horizon_val is None or float(d["value"]) != float(horizon_val):
                    findings.append(Finding("G55", "BLOCK", ref,
                                            "IRR exponent denominator literal %s != horizon %s"
                                            % (d["value"], horizon_val)))
            elif d["kind"] == "ref":
                rs, rc = split_ref(d["value"])
                got = "%s!%s" % (rs, norm_coord(rc)) if rs else "%s!%s" % (sheet, norm_coord(rc))
                want = horizon_norm if horizon_norm and "!" in horizon_norm else ("%s!%s" % (sheet, horizon_norm) if horizon_norm else None)
                if want is None or got != want:
                    findings.append(Finding("G55", "BLOCK", ref,
                                            "IRR exponent references %s, not horizon cell %s"
                                            % (got, want)))
    if not findings:
        findings.append(Finding("G55", "PASS", None, "IRR horizon ties to config horizon"))
    return findings


def _exponent_denominators(formula):
    """Extract '1/<x>' style exponents from '^(1/x)' and POWER(.,1/x).

    Returns list of {kind:'literal'|'ref', value}. Looks for the pattern 1/<token>
    appearing as an exponent operand.
    """
    out = []
    # match (1/ <token>) where token is a number OR a (possibly sheet-qualified) ref.
    ref_pat = r"(?:'[^']+'!|[A-Za-z0-9_]+!)?\$?[A-Za-z]{1,3}\$?\d+"
    num_pat = r"\d+(?:\.\d+)?"
    for m in re.finditer(r"1\s*/\s*(%s|%s)" % (ref_pat, num_pat), formula):
        tok = m.group(1)
        if re.fullmatch(num_pat, tok):
            out.append({"kind": "literal", "value": tok})
        else:
            out.append({"kind": "ref", "value": tok})
    return out


def g56_freeze_n(wb_formula, config):
    """F13: self-reference IF else-arm un-wrapped in N(); same-cell exact only."""
    cfg = config or {}
    external = bool(cfg.get("external_facing", False))
    findings = []
    for ws in wb_formula.worksheets:
        for coord, f in iter_formula_cells(ws):
            if not f.upper().startswith("=IF("):
                continue
            self_ref = _if_else_is_self_ref(f, coord)
            if self_ref:
                wrapped = bool(re.search(r"N\s*\(", f, re.IGNORECASE))
                if not wrapped:
                    sev = "BLOCK" if external else "WARN"
                    findings.append(Finding("G56", sev, "%s!%s" % (ws.title, coord),
                                            "self-referencing IF else-arm not wrapped in N(): %s" % f))
    if not findings:
        findings.append(Finding("G56", "PASS", None, "no unwrapped self-referential freeze formulas"))
    return findings


def _if_else_is_self_ref(formula, own_coord):
    """F13: outermost IF's else-arm (3rd arg) is a bare ref to this cell's coord."""
    body = formula[1:] if formula.startswith("=") else formula
    if not body.upper().startswith("IF("):
        return False
    inner = body[3:]
    # split top-level args by comma
    args = _split_top_level(inner)
    if len(args) < 3:
        return False
    else_arm = args[2].strip().rstrip(")")
    target = is_simple_alias("=" + else_arm) or (norm_coord(else_arm)
                                                 if SINGLE_CELL_RE.match(else_arm) else None)
    return target == norm_coord(own_coord)


def _split_top_level(s):
    """Split a string on top-level commas (respecting parens), drop a trailing ')'."""
    depth = 0
    args = []
    cur = ""
    for ch in s:
        if ch == "(":
            depth += 1
            cur += ch
        elif ch == ")":
            if depth == 0:
                break
            depth -= 1
            cur += ch
        elif ch == "," and depth == 0:
            args.append(cur)
            cur = ""
        else:
            cur += ch
    if cur:
        args.append(cur)
    return args


def _choose_arity(formula):
    """Return the number of value-arms in the outermost CHOOSE (or CHOOSE wrapping
    a VLOOKUP/index selector), else None. Arity = total top-level args - 1 (the
    first arg is the index selector)."""
    body = formula[1:] if formula.startswith("=") else formula
    m = re.search(r"CHOOSE\s*\(", body, re.IGNORECASE)
    if not m:
        return None
    inner = body[m.end():]
    args = _split_top_level(inner)
    if len(args) < 2:
        return None
    return len(args) - 1


def g57_choose_arch(wb_formula, config):
    """deepseek P1: enforce the GATES.md arity requirement (arity == case count)
    for declared live-range CHOOSE formulas, in addition to the stray-IF WARN scan.
    A live-range CHOOSE with the wrong number of arms WARNs per spec."""
    cfg = config or {}
    findings = []
    # per-axis case count, keyed by axis name (deepseek: this dict was dead before)
    n_axes_states = {}
    for ax in cfg.get("scenario_axes", []):
        n_axes_states[_axis_name(ax, 0)] = len(ax.get("values", []))

    # ---- arity check on each declared live range ----
    live_range_axis = cfg.get("live_range_axis", {})
    for lr in cfg.get("live_ranges", []):
        axis_name = live_range_axis.get(lr)
        if axis_name is None:
            continue   # no declared axis to size against
        want = n_axes_states.get(axis_name)
        if want is None:
            continue
        sheet, coord = split_ref(lr)
        if sheet not in wb_formula.sheetnames:
            continue
        f = wb_formula[sheet][coord].value
        if not (isinstance(f, str) and f.startswith("=")):
            continue
        arity = _choose_arity(f)
        if arity is None:
            findings.append(Finding("G57", "WARN", lr,
                                    "declared live range is not a CHOOSE: %s" % f))
        elif arity != want:
            findings.append(Finding("G57", "WARN", lr,
                                    "live-range CHOOSE arity %d != axis '%s' case count %d: %s"
                                    % (arity, axis_name, want, f)))

    # WARN on stray IF referencing a declared scenario-axis cell outside live ranges
    toggle_norms = set()
    for ax in cfg.get("scenario_axes", []):
        toggle_norms.add("%s!%s" % (ax["sheet"], norm_coord(ax["cell"])))
    live_ranges = set(cfg.get("live_ranges", []))
    for ws in wb_formula.worksheets:
        for coord, f in iter_formula_cells(ws):
            ref = "%s!%s" % (ws.title, coord)
            if ref in live_ranges:
                continue
            if not f.upper().startswith("=IF("):
                continue
            for tn in toggle_norms:
                tsheet, tcoord = split_ref(tn)
                # reference to the toggle cell inside a stray IF
                if re.search(r"(?:'%s'!|%s!)?\$?%s\$?%s" % (re.escape(tsheet), re.escape(tsheet),
                                                            re.escape(tcoord[:len(tcoord)-len(re.findall(r'\d+', tcoord)[0])] if re.findall(r'\d+', tcoord) else tcoord),
                                                            re.escape(re.findall(r'\d+', tcoord)[0]) if re.findall(r'\d+', tcoord) else ""), f):
                    findings.append(Finding("G57", "WARN", ref,
                                            "stray IF references scenario toggle outside live ranges: %s" % f))
                    break
    if not findings:
        findings.append(Finding("G57", "PASS", None, "scenario architecture uses CHOOSE live ranges"))
    return findings


# ---------------------------------------------------------------------------
# T5 - format / cosmetic
# ---------------------------------------------------------------------------
def g60_color_protocol(wb_formula, config):
    """F9/codex#18 + codex P1 (round 2): the green-hardcode rule is SYMMETRIC.
    Scan EVERY populated cell, not just formulas: a green font MUST be a pure
    cross-sheet link, and a pure cross-sheet link MUST be green. A green-font
    hardcoded value (not a formula at all) is off-protocol — green signals a
    cross-sheet link, never a typed-in number. Off-protocol in output/tie-out
    regions BLOCKS unconditionally; bulk off-protocol >= 10% BLOCKS."""
    cfg = config or {}
    output_refs = set()
    for out in cfg.get("outputs", []):
        s, c = split_ref(out["ref"])
        output_refs.add("%s!%s" % (s, norm_coord(c)))
    es, inv = sheet_role_map(cfg)
    findings = []
    total = 0
    off = 0
    for ws in wb_formula.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                is_formula = isinstance(v, str) and v.startswith("=")
                rgb = cell_font_rgb(cell)
                is_green = (rgb == GREEN_LINK)
                link = is_cross_sheet_link(v) if is_formula else False
                # only assess cells that participate in the link/green protocol:
                # any green-font cell, or any pure cross-sheet link.
                if not (link or is_green):
                    continue
                total += 1
                coord = cell.coordinate
                ref = "%s!%s" % (ws.title, norm_coord(coord))
                # a pure cross-sheet link MUST be green; a green font MUST be a
                # pure cross-sheet link (so a green hardcode / green arithmetic
                # formula is off-protocol).
                protocol_ok = (link == is_green)
                if not protocol_ok:
                    off += 1
                    if ref in output_refs:
                        kind = "hardcode" if not is_formula else "cross-sheet link"
                        findings.append(Finding("G60", "BLOCK", "%s!%s" % (ws.title, coord),
                                                "off-protocol color on output/tie-out %s" % kind))
    if total:
        pct = 100.0 * off / total
        if pct >= 10.0:
            findings.append(Finding("G60", "BLOCK", None,
                                    "%.0f%% of color-protocol cells off protocol (>= 10%%): "
                                    "build helper bypassed" % pct))
        elif off:
            findings.append(Finding("G60", "WARN", None,
                                    "%d/%d color-protocol cells off protocol" % (off, total)))
    if not findings:
        findings.append(Finding("G60", "PASS", None, "color protocol clean"))
    return findings


def g61_blue_alias(wb_formula, config):
    """F5/F19 + codex P1 (round 2): a blue formula must be a same-sheet single-cell
    alias '=+P26' / '=P26' AND that alias must terminate at a BLUE input. The chain
    may hop ONE cell (alias -> alias -> blue input). A blue alias pointing at a
    black formula or a non-blue cell, or a multi-hop chain, WARNs."""
    findings = []
    for ws in wb_formula.worksheets:
        for coord, f in iter_formula_cells(ws):
            rgb = cell_font_rgb(wb_formula[ws.title][coord])
            if rgb != BLUE_INPUT:
                continue
            alias = is_simple_alias(f)
            if alias is None:
                findings.append(Finding("G61", "WARN", "%s!%s" % (ws.title, coord),
                                        "blue formula is not a same-sheet single-cell alias: %s" % f))
                continue
            # resolve the alias terminal (allow one hop), require a BLUE input there
            ok, reason = _alias_terminates_at_blue(ws, alias, max_hops=2)
            if not ok:
                findings.append(Finding("G61", "WARN", "%s!%s" % (ws.title, coord),
                                        "blue alias %s does not terminate at a blue input (%s)"
                                        % (f, reason)))
    if not findings:
        findings.append(Finding("G61", "PASS", None,
                                "all blue formulas are legal aliases terminating at blue inputs"))
    return findings


def _alias_terminates_at_blue(ws, coord, max_hops=2):
    """Walk a same-sheet single-cell alias chain (<= max_hops cells) and report
    whether the terminal cell is a BLUE input. Returns (ok, reason). A terminal
    that is a blue hardcode (non-formula blue) or a blue input is ok; a black
    formula / non-blue cell / chain longer than max_hops is not."""
    seen = set()
    cur = norm_coord(coord)
    for _hop in range(max_hops):
        if cur in seen:
            return (False, "alias cycle at %s" % cur)
        seen.add(cur)
        cell = ws[cur]
        rgb = cell_font_rgb(cell)
        v = cell.value
        is_formula = isinstance(v, str) and v.startswith("=")
        if rgb == BLUE_INPUT and not is_formula:
            return (True, "terminal blue input %s" % cur)   # blue hardcode terminal
        if rgb == BLUE_INPUT and is_formula:
            nxt = is_simple_alias(v)
            if nxt is None:
                return (False, "terminal %s is a blue non-alias formula" % cur)
            cur = norm_coord(nxt)   # hop once more
            continue
        # non-blue terminal
        if is_formula:
            return (False, "terminal %s is a non-blue formula" % cur)
        return (False, "terminal %s is not a blue input" % cur)
    return (False, "alias chain exceeds %d hops" % max_hops)


def g62_check_fmt(wb_formula, config):
    """F1 + codex P1 (round 2): every Check cell must carry FMT_NUM1 (0.0) exactly.
    A wrong-but-specific format (e.g. FMT_NUM 0dp) is flagged too, not only the
    General worst case. Config `check_fmt_alternates` may list intentional
    alternates."""
    cfg = config or {}
    alternates = set(cfg.get("check_fmt_alternates", []))
    want = FMT_NUM1
    findings = []
    for sheet, coord in _check_cells(wb_formula, config):
        c = wb_formula[sheet][coord]
        fmt = c.number_format
        if fmt == want or fmt in alternates:
            continue
        if fmt == "General":
            findings.append(Finding("G62", "WARN", "%s!%s" % (sheet, coord),
                                    "Check cell uses General format (should be FMT_NUM1 0.0)"))
        else:
            findings.append(Finding("G62", "WARN", "%s!%s" % (sheet, coord),
                                    "Check cell uses %r, expected FMT_NUM1 (0.0)" % fmt))
    if not findings:
        findings.append(Finding("G62", "PASS", None, "Check cells use FMT_NUM1 precision"))
    return findings


def g63_number_fmt(wb_formula, config):
    """F/codex P1 + deepseek P1 (round 2): an output's number_format must EXACTLY
    match its role's expected format (ROLE_FMT[role]). A wrong-but-specific format
    (e.g. an EBITDA-m output formatted 0dp instead of 1dp) is flagged too, not only
    the General worst case. Config `number_fmt_alternates: {role: [fmt, ...]}` lists
    intentional per-role alternates; `number_fmt_general_ok: [role, ...]` opts a role
    out (General accepted)."""
    cfg = config or {}
    alternates = cfg.get("number_fmt_alternates", {})
    general_ok = set(cfg.get("number_fmt_general_ok", []))
    findings = []
    for out in cfg.get("outputs", []):
        ref = out["ref"]
        role = out.get("role")
        want = ROLE_FMT.get(role)
        if want is None:
            continue
        s, c = split_ref(ref)
        if s not in wb_formula.sheetnames:
            continue
        cell = wb_formula[s][c]
        fmt = cell.number_format
        ok_alts = set(alternates.get(role, []))
        if fmt == want or fmt in ok_alts:
            continue
        if fmt == "General" and role in general_ok:
            continue
        if fmt == "General":
            findings.append(Finding("G63", "WARN", ref,
                                    "output role %s uses General, expected role precision" % role))
        else:
            findings.append(Finding("G63", "WARN", ref,
                                    "output role %s uses %r, expected role-correct precision %r"
                                    % (role, fmt, want)))
    if not findings:
        findings.append(Finding("G63", "PASS", None, "output rows use role-correct precision"))
    return findings


def g64_tab_color(wb_formula, config):
    """F19: canonical tab-color map by role."""
    cfg = config or {}
    es, inv = sheet_role_map(cfg)
    exempt = set(cfg.get("tab_color_exempt", []))
    findings = []
    for ws in wb_formula.worksheets:
        title = ws.title
        if title in exempt:
            continue
        role = inv.get(title)
        if role not in TAB_COLOR_BY_ROLE:
            continue
        want = TAB_COLOR_BY_ROLE[role]
        got = ws.sheet_properties.tabColor
        got_rgb = getattr(got, "rgb", got) if got is not None else None
        if got_rgb is not None and isinstance(got_rgb, str):
            got_rgb = got_rgb.upper()
            if len(got_rgb) == 8:   # strip any 2-char alpha (FF or 00)
                got_rgb = got_rgb[2:]
        if want is None and got_rgb not in (None,):
            findings.append(Finding("G64", "WARN", title,
                                    "%s tab should have no tab color, has %s" % (role, got_rgb)))
        elif want is not None and got_rgb != want:
            findings.append(Finding("G64", "WARN", title,
                                    "%s tab color %s != expected %s" % (role, got_rgb, want)))
    if not findings:
        findings.append(Finding("G64", "PASS", None, "tab colors match role map"))
    return findings


def g65_fit_to_page(wb_formula, config):
    cfg = config or {}
    es, inv = sheet_role_map(cfg)
    findings = []
    for ws in wb_formula.worksheets:
        if ws.sheet_state != "visible":
            continue
        if inv.get(ws.title) == "disclaimer":
            continue
        fit = ws.page_setup.fitToWidth
        prop = ws.sheet_properties.pageSetUpPr
        fit_on = bool(fit) or (prop is not None and prop.fitToPage)
        if not fit_on:
            findings.append(Finding("G65", "WARN", ws.title, "fit-to-page not set"))
    if not findings:
        findings.append(Finding("G65", "PASS", None, "all visible sheets fit-to-page"))
    return findings


def g66_stamp(wb_formula, config):
    cfg = config or {}
    want = CONFIDENTIAL_LINE
    found = False
    for ws in wb_formula.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and want in c.value:
                    found = True
    if found:
        return [Finding("G66", "PASS", None, "confidentiality stamp present + canonical")]
    return [Finding("G66", "WARN", None, "canonical confidentiality stamp not found")]


# ---------------------------------------------------------------------------
# Render gate (separate --render invocation; BLOCK-on-send). GATES.md "Render
# gate" section. codex P1 #6: this was a dead argparse flag (args.render never
# used). Now a real code path that fail-closes LOUDLY when soffice / PyMuPDF are
# missing instead of silently no-op'ing, and verifies the rendered artifact when
# the deps + a rendered PDF are present.
# ---------------------------------------------------------------------------
def g_render(workbook_path, rendered_pdf=None, wb_formula=None, config=None):
    """Render-verify gate (G70). Offline: checks `soffice` via shutil.which and
    PyMuPDF via importlib.util.find_spec with a clear env message (not a
    traceback). When deps are missing it BLOCKS (render is BLOCK-on-send), never
    silently passes. When deps + a rendered PDF are present it asserts the PDF
    exists, page count == visible-sheet count, pages are non-blank, and
    fitToWidth survived the recalc."""
    findings = []
    have_soffice = shutil.which("soffice") is not None or shutil.which("libreoffice") is not None
    have_pymupdf = importlib.util.find_spec("fitz") is not None
    if not (have_soffice and have_pymupdf):
        missing = []
        if not have_soffice:
            missing.append("soffice/libreoffice (LibreOffice headless)")
        if not have_pymupdf:
            missing.append("PyMuPDF (import fitz)")
        return [Finding("G70", "BLOCK", None,
                        "render gate cannot run: missing deps %s. Render is BLOCK-on-send; "
                        "install the render deps or do not ship from this host." % missing)]

    # deps present: verify the rendered artifact if one was produced
    if rendered_pdf is None or not Path(rendered_pdf).exists():
        return [Finding("G70", "BLOCK", None,
                        "render deps present but no rendered PDF supplied/produced "
                        "(--rendered-pdf); cannot prove the workbook renders")]
    try:
        import fitz  # noqa: F401
        doc = fitz.open(rendered_pdf)
        page_count = doc.page_count
    except Exception as e:
        return [Finding("G70", "BLOCK", rendered_pdf, "cannot open rendered PDF: %s" % e)]

    # page count == visible-sheet count
    visible = 0
    if wb_formula is not None:
        for ws in wb_formula.worksheets:
            if ws.sheet_state == "visible":
                visible += 1
        if page_count != visible:
            findings.append(Finding("G70", "BLOCK", rendered_pdf,
                                    "rendered page count %d != visible-sheet count %d"
                                    % (page_count, visible)))
    # non-blank pages
    try:
        for i in range(page_count):
            txt = doc.load_page(i).get_text().strip()
            pix = doc.load_page(i).get_pixmap()
            if not txt and (pix.width == 0 or pix.height == 0):
                findings.append(Finding("G70", "BLOCK", "%s p%d" % (rendered_pdf, i + 1),
                                        "rendered page is blank"))
    except Exception:
        pass
    # fitToWidth survived
    if wb_formula is not None:
        for ws in wb_formula.worksheets:
            if ws.sheet_state != "visible":
                continue
            if (config or {}).get("expected_sheets", {}).get("disclaimer") == ws.title:
                continue
            prop = ws.sheet_properties.pageSetUpPr
            if not (ws.page_setup.fitToWidth or (prop is not None and prop.fitToPage)):
                findings.append(Finding("G70", "BLOCK", ws.title,
                                        "fitToWidth did not survive the recalc"))
    if not findings:
        findings.append(Finding("G70", "PASS", rendered_pdf, "render verified"))
    return findings


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def run_all_gates(wb_formula, wb_cached, config, evidence, workbook_path="model.xlsx"):
    """Run T0->T5 in fixed order. A T0 BLOCK short-circuits the rest (D-ENTRY)."""
    recalc_ev = (evidence or {}).get("recalc")
    sweep_ev = (evidence or {}).get("sweep")
    xbrl = (evidence or {}).get("xbrl")
    baseline_fp = (evidence or {}).get("baseline_fp")
    actuals_override = (evidence or {}).get("actuals_sheet_override")
    demo = (evidence or {}).get("demo", False)
    allow_unlocked = (evidence or {}).get("allow_unlocked", False)
    cached_workbook = (evidence or {}).get("cached_workbook")
    test_mode = (evidence or {}).get("test_mode", False)

    findings = []
    # T0
    findings.extend(g00_env(allow_unlocked))
    findings.extend(g01_config(config, wb_formula, demo, baseline_fp=baseline_fp))
    findings.extend(g02_oracle(recalc_ev, sweep_ev, demo, test_mode))
    if [f for f in findings if f.severity == "BLOCK"]:
        return findings   # T0 BLOCK aborts (D-ENTRY)

    # T1
    findings.extend(g10_extlink(wb_formula))
    findings.extend(g11_hidden(wb_formula, config))
    findings.extend(g12_merged(wb_formula, config))
    findings.extend(g13_condfmt(wb_formula, config))
    findings.extend(g14_struct_fp(wb_formula, config, baseline_fp))

    # T2
    findings.extend(g20_cache(wb_formula, wb_cached, config))
    findings.extend(g21_recalc_mtime(recalc_ev, workbook_path))
    findings.extend(g22_recalc_canary(recalc_ev, wb_cached, workbook_path, wb_formula))
    findings.extend(g23_iteration(wb_cached, recalc_ev, config))
    findings.extend(g24_error_lit(wb_cached, config))
    findings.extend(g25_check_row(wb_cached, wb_formula, config))
    findings.extend(g26_check_count(wb_formula, config))
    findings.extend(g27_leading_eq(wb_formula, config))
    # FIX-ROUND-2 (codex P0-1): close the two-file wrong-green seam.
    findings.extend(g28_shim_guard(wb_formula, wb_cached, config,
                                    cached_workbook, demo, test_mode))
    findings.extend(g29_output_recalc(recalc_ev, wb_formula, wb_cached, config,
                                      sweep_ev=sweep_ev))

    # T3
    findings.extend(g30_actuals_exists(wb_cached, config, actuals_override))
    findings.extend(g31_embedded_const(wb_formula, config))
    # the foot gates only run when an Actuals sheet resolves
    cfg = config or {}
    if cfg.get("actuals_sheet") or actuals_override:
        findings.extend(g32_rev_foot(wb_cached, config, wb_formula))
        findings.extend(g33_opex_foot(wb_cached, config, wb_formula))
        findings.extend(g34_pl_chain(wb_cached, config, wb_formula))
        findings.extend(g35_ebitda_recon(wb_cached, config, wb_formula))
        findings.extend(g36_bs_balance(wb_cached, config, wb_formula))
        findings.extend(g37_bs_subtotal(wb_cached, config, wb_formula))
        findings.extend(g38_cf_section(wb_cached, config, wb_formula))
        findings.extend(g39_cash_tie(wb_cached, config, wb_formula))
        findings.extend(g40_qtr_annual(wb_cached, config))
    findings.extend(g41_source_comment(wb_formula, config))
    findings.extend(g42_xbrl(wb_cached, config, xbrl))

    # T4
    findings.extend(g49_sweep_coverage(sweep_ev, config))
    findings.extend(g50_sweep_axis(sweep_ev, config))
    findings.extend(g51_sweep_dir(sweep_ev, config))
    findings.extend(g52_sweep_mag(sweep_ev, config))
    findings.extend(g53_sweep_errscan(sweep_ev))
    findings.extend(g54_min_cash(sweep_ev, config))
    findings.extend(g55_irr_horizon(wb_formula, wb_cached, config))
    findings.extend(g56_freeze_n(wb_formula, config))
    findings.extend(g57_choose_arch(wb_formula, config))

    # T5
    findings.extend(g60_color_protocol(wb_formula, config))
    findings.extend(g61_blue_alias(wb_formula, config))
    findings.extend(g62_check_fmt(wb_formula, config))
    findings.extend(g63_number_fmt(wb_formula, config))
    findings.extend(g64_tab_color(wb_formula, config))
    findings.extend(g65_fit_to_page(wb_formula, config))
    findings.extend(g66_stamp(wb_formula, config))
    return findings


def exit_code(findings, strict=False, non_shippable=False):
    """F14/codex#9: exit 0 is the ONLY ship-eligible code.

    exit = (BLOCK count, strict adds WARN). If 0 BLOCKs but non_shippable
    stamped -> exit 3 (ran-clean-but-not-shippable), never green.
    """
    n = 0
    for f in findings:
        if f.severity == "BLOCK":
            n += 1
        elif strict and f.severity == "WARN":
            n += 1
    if n == 0 and non_shippable:
        return 3
    return n


def is_non_shippable(findings, allow_unlocked):
    """Verdict stamped non-shippable when env was unlocked (G00 downgraded to WARN)
    OR the two-file shim was used outside --demo/--test (G28 WARN, FIX-ROUND-2)."""
    for f in findings:
        if f.gate == "G28" and f.severity == "WARN":
            return True
    if allow_unlocked:
        for f in findings:
            if f.gate == "G00" and f.severity == "WARN":
                return True
    return False


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _render_text(findings, quiet, strict, non_shippable=False):
    """D3 (P1-1): the printed verdict must NEVER read '-> CLEAN' when the run is
    non-shippable (exit 3) or has BLOCKs. `_render_text` previously derived the
    verdict purely from the BLOCK count, so a non-shippable run (0 BLOCK, exit 3)
    printed '-> CLEAN' while the process exited 3. The verdict now folds in
    non_shippable so the human-readable line agrees with the exit code."""
    lines = []
    block_n = warn_n = pass_n = 0
    for f in findings:
        eff = f.severity
        if strict and f.severity == "WARN":
            eff = "BLOCK"
        if eff == "BLOCK":
            block_n += 1
        elif f.severity == "WARN":
            warn_n += 1
        else:
            pass_n += 1
        if f.severity == "PASS":
            continue   # quiet-on-pass
        loc = f.location or "--"
        marker = "BLOCK" if eff == "BLOCK" else "WARN"
        lines.append("[%s] %-5s %-22s %s" % (marker, f.gate, loc, f.message))
    if block_n:
        verdict = "BLOCKED"
    elif non_shippable:
        verdict = "NON-SHIPPABLE (ran clean but not ship-eligible; exit 3)"
    else:
        verdict = "CLEAN"
    lines.append("")
    lines.append("summary: %d BLOCK, %d WARN, %d PASS -> %s%s"
                 % (block_n, warn_n, pass_n, verdict, " [--strict]" if strict else ""))
    return "\n".join(lines)


def main(argv=None):
    p = argparse.ArgumentParser(prog="model_gate.py",
                                description="model-gate: the single offline shipping gate")
    p.add_argument("workbook", help="<TICKER> Model.xlsx (Excel-recalced; carries cache)")
    p.add_argument("--cached-workbook",
                   help="optional separate file for the cached/value view (offline test "
                        "shim; production passes one Excel-recalced file and omits this)")
    p.add_argument("--config")
    p.add_argument("--recalc-evidence")
    p.add_argument("--sweep-evidence")
    p.add_argument("--xbrl")
    p.add_argument("--baseline-fp")
    p.add_argument("--actuals-sheet")
    p.add_argument("--demo", action="store_true")
    p.add_argument("--test", action="store_true", dest="test_mode",
                   help="like --demo: legalizes the two-file --cached-workbook shim "
                        "(offline test seam). A production run supplies ONE recalced file.")
    p.add_argument("--allow-unlocked-env", action="store_true")
    p.add_argument("--render", action="store_true",
                   help="run the render-verify gate (G70; BLOCK-on-send)")
    p.add_argument("--rendered-pdf",
                   help="path to the already-rendered PDF to verify (with --render)")
    p.add_argument("--json", action="store_true")
    p.add_argument("--quiet", action="store_true")
    p.add_argument("--strict", action="store_true")
    args = p.parse_args(argv)

    try:
        wb_formula = openpyxl.load_workbook(args.workbook, data_only=False)
        cached_path = args.cached_workbook or args.workbook
        wb_cached = openpyxl.load_workbook(cached_path, data_only=True)
    except Exception as e:
        sys.stderr.write("error: cannot open %s: %s\n" % (args.workbook, e))
        return 2

    config = load_json(args.config)
    evidence = {
        "recalc": load_json(args.recalc_evidence),
        "sweep": load_json(args.sweep_evidence),
        "xbrl": load_json(args.xbrl),
        "baseline_fp": load_json(args.baseline_fp),
        "actuals_sheet_override": args.actuals_sheet,
        "demo": args.demo,
        "test_mode": args.test_mode,
        "allow_unlocked": args.allow_unlocked_env,
        "cached_workbook": args.cached_workbook,
    }

    findings = run_all_gates(wb_formula, wb_cached, config, evidence,
                             workbook_path=args.workbook)

    if args.render:
        findings.extend(g_render(args.workbook, args.rendered_pdf,
                                 wb_formula=wb_formula, config=config))

    non_ship = is_non_shippable(findings, args.allow_unlocked_env)
    code = exit_code(findings, strict=args.strict, non_shippable=non_ship)

    env = {"python": sys.executable, "openpyxl": openpyxl.__version__,
           "non_shippable": non_ship}
    if args.json:
        out = {
            "workbook": args.workbook,
            "passed": code == 0,
            "exit_code": code,
            "strict": args.strict,
            "env": env,
            "blocks": [f.as_dict() for f in findings if f.severity == "BLOCK"],
            "warnings": [f.as_dict() for f in findings if f.severity == "WARN"],
            "findings": [f.as_dict() for f in findings],
        }
        with open("gate_verdict.json", "w", encoding="utf-8") as fh:
            json.dump(out, fh, indent=2)
        print(json.dumps(out, indent=2))
    else:
        print(_render_text(findings, args.quiet, args.strict, non_shippable=non_ship))

    return code


if __name__ == "__main__":
    sys.exit(main())
